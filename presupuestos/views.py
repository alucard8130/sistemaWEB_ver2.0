# Create your views here.
from pyexpat.errors import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
import openpyxl
from facturacion.models import (
    CobroOtrosIngresos,
    FacturaOtrosIngresos,
    Pago,
    TipoOtroIngreso,
)
from gastos.models import Gasto, GrupoGasto, PagoGasto, TipoGasto, SubgrupoGasto
from .models import Presupuesto, PresupuestoCierre, PresupuestoIngreso
from .forms import PresupuestoCargaMasivaForm, PresupuestoForm
from django.utils.timezone import now
from django.db.models import Sum, F
from empresas.models import Empresa
from calendar import month_name
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth import authenticate
import calendar
from openpyxl.styles import Font, Alignment
from django.db.models.functions import ExtractYear, ExtractMonth
from io import BytesIO
from datetime import date, datetime
from django.db.models.functions import TruncMonth, TruncYear
from .forms import PresupuestoCargaMasivaForm
from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP


@login_required
def presupuesto_nuevo(request):
    if request.method == "POST":
        form = PresupuestoForm(request.POST, user=request.user)
        if form.is_valid():
            presupuesto = form.save(commit=False)
            # Solo para usuarios normales, asigna empresa directamente
            if not request.user.is_superuser:
                presupuesto.empresa = request.user.perfilusuario.empresa
            presupuesto.save()
            form.save()
            return redirect("presupuesto_lista")
    else:
        form = PresupuestoForm(user=request.user)
    return render(request, "presupuestos/form.html", {"form": form})


@login_required
def presupuesto_editar(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    if request.method == "POST":
        form = PresupuestoForm(request.POST, instance=presupuesto, user=request.user)
        if form.is_valid():
            presupuesto = form.save(commit=False)
            if not request.user.is_superuser:
                presupuesto.empresa = request.user.perfilusuario.empresa
            presupuesto.save()
            form.save()
            return redirect("presupuesto_lista")
    else:
        form = PresupuestoForm(instance=presupuesto, user=request.user)
    return render(
        request, "presupuestos/form.html", {"form": form, "presupuesto": presupuesto}
    )


@login_required
def presupuesto_eliminar(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    if request.method == "POST":
        presupuesto.delete()
        return redirect("presupuesto_lista")
    return render(
        request, "presupuestos/confirmar_eliminar.html", {"presupuesto": presupuesto}
    )


@login_required
# dashboard.html
def dashboard_presupuestal(request):
    es_super = request.user.is_superuser
    anio_actual = now().year
    anio = int(request.GET.get("anio", anio_actual))
    # mes = int(request.GET.get("mes", 0))  # 0 = todo el año
    mes = int(
        request.GET.get("mes", datetime.now().month)
    )  # Si no se especifica, toma el mes actual

    # Filtro empresa
    if es_super:
        empresas = Empresa.objects.all()
        empresa_id = request.GET.get("empresa")
        if empresa_id:
            try:
                empresa = Empresa.objects.get(pk=int(empresa_id))
            except (Empresa.DoesNotExist, ValueError):
                empresa = empresas.first() if empresas else None
        else:
            empresa = empresas.first() if empresas else None
    else:
        empresa = request.user.perfilusuario.empresa
        empresas = Empresa.objects.filter(pk=empresa.id)

    if not empresa:
        # Si no hay empresa, devuelve página vacía o mensaje amigable
        meses_esp = [
            "01 Ene",
            "02 Feb",
            "03 Mar",
            "04 Abr",
            "05 May",
            "06 Jun",
            "07 Jul",
            "08 Ago",
            "09 Sep",
            "10 Oct",
            "11 Nov",
            "12 Dic",
        ]
        contexto = {
            "empresas": empresas,
            "empresa_id": None,
            "anio": anio,
            "mes": mes,
            "labels": [],
            "datos_presupuesto": [],
            "datos_gastado": [],
            "total_presupuestado": 0,
            "total_gastado": 0,
            "es_super": es_super,
            "meses_esp": meses_esp,
        }
        return render(request, "presupuestos/dashboard.html", contexto)

    # Presupuestos del año y empresa
    pres_qs = Presupuesto.objects.filter(empresa=empresa, anio=anio)
    if mes:
        pres_qs = pres_qs.filter(mes=mes)

    total_presupuestado = pres_qs.aggregate(total=Sum("monto"))["total"] or 0

    # Gastos reales
    gastos_qs = Gasto.objects.filter(empresa=empresa, fecha__year=anio)
    if mes:
        gastos_qs = gastos_qs.filter(fecha__month=mes)
    total_gastado = gastos_qs.aggregate(total=Sum("monto"))["total"] or 0

    # Datos para gráfico (línea presupuestos vs. gastos)
    meses = list(range(1, 13))
    meses_esp = [
        "Ene",
        "Feb",
        "Mar",
        "Abr",
        "May",
        "Jun",
        "Jul",
        "Ago",
        "Sep",
        "Oct",
        "Nov",
        "Dic",
    ]
    labels = meses_esp
    datos_presupuesto = []
    datos_gastado = []

    for m in meses:
        pres_mes = pres_qs.filter(mes=m).aggregate(total=Sum("monto"))["total"] or 0
        gasto_mes = (
            gastos_qs.filter(fecha__month=m).aggregate(total=Sum("monto"))["total"] or 0
        )
        datos_presupuesto.append(float(pres_mes))
        datos_gastado.append(float(gasto_mes))

    contexto = {
        "empresas": empresas,
        "empresa_id": empresa.id,
        "anio": anio,
        "mes": mes,
        "labels": labels,
        "datos_presupuesto": datos_presupuesto,
        "datos_gastado": datos_gastado,
        "total_presupuestado": total_presupuestado,
        "total_gastado": total_gastado,
        "es_super": es_super,
    }
    return render(request, "presupuestos/dashboard.html", contexto)


@login_required
# matriz ppto gastos
def matriz_presupuesto(request):
    anio = int(request.GET.get("anio", now().year))
    now_year = now().year
    # anios = list(range(now_year, 2021, -1))
    anios = (
        Presupuesto.objects.values_list("anio", flat=True).distinct().order_by("anio")
    )

    # Empresa y permisos
    if request.user.is_superuser:
        empresas = Empresa.objects.all()
        empresa_id = request.GET.get("empresa")
        empresa = Empresa.objects.get(pk=empresa_id) if empresa_id else empresas.first()
    else:
        empresa = request.user.perfilusuario.empresa
        empresas = None

    meses = list(range(1, 13))
    meses_nombres = [month_name[m].capitalize() for m in meses]

    # ¿Cerrado?
    cierre = PresupuestoCierre.objects.filter(empresa=empresa, anio=anio).first()
    bloqueado = cierre.cerrado if cierre else False
    pedir_superuser = False

    # ----------- NUEVO: lógica para abrir presupuesto con superuser -------------
    if bloqueado and not request.user.is_superuser:
        if request.method == "POST" and "superuser_username" in request.POST:
            username = request.POST.get("superuser_username")
            password = request.POST.get("superuser_password")
            superuser = authenticate(username=username, password=password)
            if superuser and superuser.is_superuser:
                # Reabrir presupuesto
                cierre.cerrado = False
                cierre.fecha_cierre = None
                cierre.cerrado_por = None
                cierre.save()
                messages.success(
                    request,
                    "Presupuesto reabierto correctamente. Ahora puedes editarlo.",
                )
                return redirect(
                    request.path
                    + f"?anio={anio}"
                    + (f"&empresa={empresa.id}" if empresa else "")
                )
            else:
                # pedir_superuser = True
                messages.error(
                    request, "Usuario o contraseña de superusuario incorrectos."
                )
        else:
            pedir_superuser = True

    # Ahora recalcula bloqueado
    cierre = PresupuestoCierre.objects.filter(empresa=empresa, anio=anio).first()
    bloqueado = cierre.cerrado if cierre else False

    # Asegúrate que pedir_superuser solo sea True si sigue bloqueado
    if not bloqueado:
        pedir_superuser = False

    # Edición habilitada si no está bloqueado o si eres superusuario
    edicion_habilitada = not bloqueado or request.user.is_superuser

    # --- Resto de la vista igual ---
    tipos = (
        TipoGasto.objects.filter(empresa=empresa)
        .select_related("subgrupo", "subgrupo__grupo")
        .order_by("subgrupo__grupo__nombre", "subgrupo__nombre", "nombre")
    )
    # tipos = TipoGasto.objects.filter(subgrupo__grupo__empresa=empresa).select_related("subgrupo", "subgrupo__grupo").order_by(
    #   "subgrupo__grupo__nombre", "subgrupo__nombre", "nombre"
    # )
    grupos_lista = []
    grupos_dict = defaultdict(lambda: defaultdict(list))
    for tipo in tipos:
        grupos_dict[tipo.subgrupo.grupo][tipo.subgrupo].append(tipo)
    for grupo, subgrupos in grupos_dict.items():
        subgrupos_lista = []
        for subgrupo, tipos_ in subgrupos.items():
            tipos_lista = []
            for tipo in tipos_:
                tipos_lista.append(
                    {"id": tipo.id, "nombre": tipo.nombre, "tipo_obj": tipo}
                )
            subgrupos_lista.append(
                {
                    "id": subgrupo.id,
                    "nombre": subgrupo.nombre,
                    "tipos": tipos_lista,
                    "subgrupo_obj": subgrupo,
                }
            )
        grupos_lista.append(
            {
                "id": grupo.id,
                "nombre": grupo.nombre,
                "subgrupos": subgrupos_lista,
                "grupo_obj": grupo,
            }
        )

    presupuestos = Presupuesto.objects.filter(empresa=empresa, anio=anio)
    presup_dict = {(p.tipo_gasto_id, p.mes): p for p in presupuestos}

    # Totales, subtotales
    totales_mes = []
    for mes in meses:
        total_mes = 0
        for tipo in tipos:
            p = presup_dict.get((tipo.id, mes))
            total_mes += p.monto if p else 0
        totales_mes.append(total_mes)

    subtotales_grupo = {}
    subtotales_subgrupo = {}
    subtotales_tipo = {}
    for grupo, subgrupos in grupos_dict.items():
        subtotal_grupo = [0] * 12
        for subgrupo, tipos_gasto in subgrupos.items():
            subtotal_subgrupo = [0] * 12
            for tipo in tipos_gasto:
                subtotal_tipo = []
                for i, mes in enumerate(meses):
                    p = presup_dict.get((tipo.id, mes))
                    valor = p.monto if p else 0
                    subtotal_tipo.append(valor)
                    subtotal_subgrupo[i] += valor
                    subtotal_grupo[i] += valor
                subtotales_tipo[tipo.id] = subtotal_tipo
            subtotales_subgrupo[subgrupo.id] = subtotal_subgrupo
        subtotales_grupo[grupo.id] = subtotal_grupo

    # Guardado de presupuestos (solo si habilitado)
    if request.method == "POST" and edicion_habilitada:
        if not empresa:
                        messages.error(
                            request,
                            "No hay empresa seleccionada. No se puede guardar el presupuesto de gastos.",
                        )
                        return redirect(request.path + f"?anio={anio}")  
        for tipo in tipos:
            for mes in meses:
                key = f"presupuesto_{tipo.id}_{mes}"
                monto = request.POST.get(key)
                if monto is not None:
                    monto = float(monto or 0)
                    grupo = tipo.subgrupo.grupo
                    subgrupo = tipo.subgrupo
                    obj, created = Presupuesto.objects.get_or_create(
                        empresa=empresa,
                        tipo_gasto=tipo,
                        anio=anio,
                        mes=mes,
                        defaults={"monto": monto, "grupo": grupo, "subgrupo": subgrupo},
                    )
                    if not created and obj.monto != monto:
                        obj.monto = monto
                        obj.save()
        # Cerrar presupuesto solo si corresponde
        if "cerrar_presupuesto" in request.POST and edicion_habilitada:
            cierre, created = PresupuestoCierre.objects.get_or_create(
                empresa=empresa, anio=anio
            )
            cierre.cerrado = True
            cierre.cerrado_por = request.user
            cierre.fecha_cierre = now()
            cierre.save()
            messages.success(
                request,
                "¡Presupuesto cerrado! Solo el superusuario puede volver a abrirlo.",
            )
            return redirect(
                request.path
                + f"?anio={anio}"
                + (f"&empresa={empresa.id}" if empresa else "")
            )
        else:
            messages.success(request, "Presupuestos actualizados")
        return redirect(
            request.path
            + f"?anio={anio}"
            + (f"&empresa={empresa.id}" if empresa else "")
        )

    return render(
        request,
        "presupuestos/matriz.html",
        {
            "grupos_lista": grupos_lista,
            "meses": meses,
            "meses_nombres": meses_nombres,
            "presup_dict": presup_dict,
            "anio": anio,
            "anios": anios,
            "empresas": empresas,
            "empresa": empresa,
            "totales_mes": totales_mes,
            "subtotales_grupo": subtotales_grupo,
            "subtotales_subgrupo": subtotales_subgrupo,
            "subtotales_tipo": subtotales_tipo,
            "is_super": request.user.is_superuser,
            "edicion_habilitada": edicion_habilitada,
            "pedir_superuser": pedir_superuser,
            "bloqueado": bloqueado,
            "cierre": cierre,
        },
    )


@login_required
def exportar_presupuesto_excel(request):
    anio = int(request.GET.get("anio", now().year))
    if request.user.is_superuser:
        empresas = Empresa.objects.all()
        empresa_id = request.GET.get("empresa")
        empresa = Empresa.objects.get(pk=empresa_id) if empresa_id else empresas.first()
    else:
        empresa = request.user.perfilusuario.empresa

    meses = list(range(1, 13))
    meses_nombres = [month_name[m][:3].capitalize() for m in meses]

    # Construir estructura de grupos/subgrupos/tipos
    tipos = TipoGasto.objects.select_related("subgrupo", "subgrupo__grupo").order_by(
        "subgrupo__grupo__nombre", "subgrupo__nombre", "nombre"
    )
    grupos_dict = defaultdict(lambda: defaultdict(list))
    for tipo in tipos:
        grupos_dict[tipo.subgrupo.grupo][tipo.subgrupo].append(tipo)
    presupuestos = Presupuesto.objects.filter(empresa=empresa, anio=anio)
    presup_dict = {(p.tipo_gasto_id, p.mes): p for p in presupuestos}

    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Presupuesto {anio}"

    # Escribir encabezados
    ws.append(["Grupo", "Subgrupo", "Tipo de Gasto"] + meses_nombres + ["Total"])

    for grupo, subgrupos in grupos_dict.items():
        for subgrupo, tipos in subgrupos.items():
            for tipo in tipos:
                fila = [str(grupo), str(subgrupo), str(tipo)]
                total = 0
                for mes in meses:
                    monto = (
                        presup_dict.get((tipo.id, mes)).monto
                        if presup_dict.get((tipo.id, mes))
                        else 0
                    )
                    fila.append(float(monto))
                    total += monto
                fila.append(float(total))
                ws.append(fila)

    # Opcional: Ajustar anchos de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Generar respuesta HTTP con el archivo
    nombre_archivo = f"Presupuesto_{empresa.nombre}_{anio}.xlsx".replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={nombre_archivo}"
    wb.save(response)
    return response


@login_required
# reporte_comparativo.html
# reporte comparativo presupuesto vs gasto
def reporte_presupuesto_vs_gasto(request):
    anio = int(request.GET.get("anio", datetime.now().year))
    medicion = request.GET.get(
        "medicion", "curso"
    )  # 'mes', 'semestre1', 'semestre2', 'anual', 'curso'
    meses = list(range(1, 13))
    meses_nombres = [calendar.month_name[m] for m in meses]

    # Determina el rango de meses según la medición
    if medicion == "mes":
        mes_actual = int(request.GET.get("mes", datetime.now().month))
        meses = [mes_actual]
        meses_nombres = [calendar.month_name[mes_actual]]
    elif medicion == "semestre1":
        meses = list(range(1, 7))
        meses_nombres = [calendar.month_name[m] for m in meses]
    elif medicion == "semestre2":
        meses = list(range(7, 13))
        meses_nombres = [calendar.month_name[m] for m in meses]
    elif medicion == "anual":
        meses = list(range(1, 13))
        meses_nombres = [calendar.month_name[m] for m in meses]
    else:  # 'curso' o default
        if anio == datetime.now().year:
            mes_actual = datetime.now().month
        else:
            mes_actual = 12
        meses = list(range(1, mes_actual + 1))
        meses_nombres = [calendar.month_name[m] for m in meses]

    # Empresa y permisos
    if request.user.is_superuser:
        empresas = Empresa.objects.all()
        empresa_id = request.GET.get("empresa")
        empresa = Empresa.objects.get(pk=empresa_id) if empresa_id else empresas.first()
    else:
        empresa = request.user.perfilusuario.empresa
        empresas = None

    tipos = (
        TipoGasto.objects.filter(empresa=empresa)
        .select_related("subgrupo", "subgrupo__grupo")
        .order_by("subgrupo__grupo__nombre", "subgrupo__nombre", "nombre")
    )

    # Diccionario: {(tipo_id, mes): monto}
    presupuestos = Presupuesto.objects.filter(empresa=empresa, anio=anio)
    presup_dict = {(p.tipo_gasto_id, p.mes): float(p.monto) for p in presupuestos}

    # Diccionario: {(tipo_id, mes): monto}
    gastos = (
        PagoGasto.objects.annotate(
            anio_pago=ExtractYear("fecha_pago"),
            mes_pago=ExtractMonth("fecha_pago"),
            tipo_id=F("gasto__tipo_gasto_id"),
        )
        .filter(anio_pago=anio, gasto__empresa=empresa)
        .values("tipo_id", "mes_pago")
        .annotate(total=Sum("monto"))
    )
    gastos_dict = {(g["tipo_id"], g["mes_pago"]): float(g["total"]) for g in gastos}

    # Estructura: grupos > subgrupos > tipos > meses
    grupos_dict = defaultdict(lambda: defaultdict(list))
    for tipo in tipos:
        grupos_dict[tipo.subgrupo.grupo][tipo.subgrupo].append(tipo)

    comparativo = []
    total_general_presup = 0
    total_general_real = 0
    total_general_var = 0

    num_meses = len(meses)
    meses_idx = list(range(num_meses))

    for grupo, subgrupos in grupos_dict.items():
        grupo_row = {
            "nombre": grupo.nombre,
            "subgrupos": [],
            "total": [0] * num_meses,
            "total_gasto": [0] * num_meses,
            "total_var": [0] * num_meses,
            "pct_var": [None] * num_meses,
        }
        grupo_total_presup = 0
        grupo_total_gasto = 0
        grupo_total_var = 0

        for subgrupo, tipos_ in subgrupos.items():
            subgrupo_row = {
                "nombre": subgrupo.nombre,
                "tipos": [],
                "total": [0] * num_meses,
                "total_gasto": [0] * num_meses,
                "total_var": [0] * num_meses,
                "pct_var": [None] * num_meses,
            }
            subgrupo_total_presup = 0
            subgrupo_total_gasto = 0
            subgrupo_total_var = 0

            for tipo in tipos_:
                row = {"nombre": tipo.nombre, "meses": []}
                anual_presup = 0
                anual_gasto = 0
                anual_var = 0

                for i, mes in enumerate(meses):
                    presupuesto = presup_dict.get((tipo.id, mes), 0)
                    gasto = gastos_dict.get((tipo.id, mes), 0)
                    variacion = gasto - presupuesto
                    row["meses"].append(
                        {
                            "presupuesto": presupuesto,
                            "gasto": gasto,
                            "variacion": variacion,
                            "mes": mes,
                        }
                    )
                    subgrupo_row["total"][i] += presupuesto
                    subgrupo_row["total_gasto"][i] += gasto
                    subgrupo_row["total_var"][i] += variacion
                    grupo_row["total"][i] += presupuesto
                    grupo_row["total_gasto"][i] += gasto
                    grupo_row["total_var"][i] += variacion

                    anual_presup += presupuesto
                    anual_gasto += gasto
                    anual_var += variacion

                row["total_anual_presup"] = anual_presup
                row["total_anual_gasto"] = anual_gasto
                row["total_anual_var"] = anual_var
                row["total_anual_pct"] = (
                    int(round(anual_var / anual_presup * 100)) if anual_presup else ""
                )

                subgrupo_row["tipos"].append(row)

                subgrupo_total_presup += anual_presup
                subgrupo_total_gasto += anual_gasto
                subgrupo_total_var += anual_var

            # Totales mensuales y anuales por subgrupo
            for i in range(num_meses):
                if subgrupo_row["total"][i]:
                    subgrupo_row["pct_var"][i] = round(
                        subgrupo_row["total_var"][i] / subgrupo_row["total"][i] * 100
                    )
                else:
                    subgrupo_row["pct_var"][i] = ""
            subgrupo_row["total_anual_presup"] = subgrupo_total_presup
            subgrupo_row["total_anual_gasto"] = subgrupo_total_gasto
            subgrupo_row["total_anual_var"] = subgrupo_total_var
            subgrupo_row["total_anual_pct"] = (
                int(round(subgrupo_total_var / subgrupo_total_presup * 100))
                if subgrupo_total_presup
                else ""
            )

            grupo_row["subgrupos"].append(subgrupo_row)

            grupo_total_presup += subgrupo_total_presup
            grupo_total_gasto += subgrupo_total_gasto
            grupo_total_var += subgrupo_total_var

        # Totales mensuales y anuales por grupo
        for i in range(num_meses):
            if grupo_row["total"][i]:
                grupo_row["pct_var"][i] = round(
                    grupo_row["total_var"][i] / grupo_row["total"][i] * 100
                )
            else:
                grupo_row["pct_var"][i] = ""
        grupo_row["total_anual_presup"] = grupo_total_presup
        grupo_row["total_anual_gasto"] = grupo_total_gasto
        grupo_row["total_anual_var"] = grupo_total_var
        grupo_row["total_anual_pct"] = (
            int(round(grupo_total_var / grupo_total_presup * 100))
            if grupo_total_presup
            else ""
        )

        comparativo.append(grupo_row)

        total_general_presup += grupo_total_presup
        total_general_real += grupo_total_gasto
        total_general_var += grupo_total_var

    # Exportar a Excel
    if request.GET.get("excel") == "1":
        return exportar_comparativo_excel(
            anio, empresa, meses, meses_nombres, comparativo
        )

    # Calcula los totales generales por mes (presupuesto, real, variación)
    tot_gen_presup = [0] * num_meses
    tot_gen_real = [0] * num_meses
    tot_gen_var = [0] * num_meses
    tot_gen_pct = [None] * num_meses

    for i in range(num_meses):
        for grupo in comparativo:
            tot_gen_presup[i] += grupo["total"][i]
            tot_gen_real[i] += grupo["total_gasto"][i]
            tot_gen_var[i] += grupo["total_var"][i]
        if tot_gen_presup[i]:
            tot_gen_pct[i] = round(tot_gen_var[i] / tot_gen_presup[i] * 100)
        else:
            tot_gen_pct[i] = None

    # Totales anuales generales
    total_general_pct = (
        int(round(total_general_var / total_general_presup * 100))
        if total_general_presup
        else ""
    )

    return render(
        request,
        "presupuestos/reporte_comparativo.html",
        {
            "anio": anio,
            "anios": list(range(datetime.now().year, 2021, -1)),
            "empresa": empresa,
            "empresas": empresas,
            "meses": meses,
            "meses_nombres": meses_nombres,
            "meses_idx": meses_idx,
            "medicion": medicion,
            "comparativo": comparativo,
            "tot_gen_presup": tot_gen_presup,
            "tot_gen_real": tot_gen_real,
            "tot_gen_var": tot_gen_var,
            "tot_gen_pct": tot_gen_pct,
            "total_general_presup": total_general_presup,
            "total_general_real": total_general_real,
            "total_general_var": total_general_var,
            "total_general_pct": total_general_pct,
        },
    )


def exportar_comparativo_excel(anio, empresa, meses, meses_nombres, comparativo):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto vs Gasto"

    # Encabezado principal
    encabezado = ["Grupo", "Subgrupo", "Tipo de Gasto"]
    for m in meses:
        nombre_mes = calendar.month_name[m]
        encabezado += [
            f"{nombre_mes} Presup.",
            f"{nombre_mes} Real",
            f"{nombre_mes} Var.",
            f"{nombre_mes} %Var.",
        ]
    encabezado += ["Total Presup.", "Total Real", "Total Var.", "Total %Var."]
    ws.append(encabezado)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    def porcentaje_var(presupuesto, variacion):
        try:
            return round((variacion / presupuesto) * 100) if presupuesto else ""
        except ZeroDivisionError:
            return ""

    for grupo in comparativo:
        for subgrupo in grupo["subgrupos"]:
            for tipo in subgrupo["tipos"]:
                fila = [grupo["nombre"], subgrupo["nombre"], tipo["nombre"]]
                for mes_num in meses:
                    # Busca el mes correcto por su número
                    mes_data = next(
                        (m for m in tipo["meses"] if m.get("mes") == mes_num), None
                    )
                    if mes_data:
                        presup = mes_data.get("presupuesto", 0)
                        real = mes_data.get("gasto", 0)
                        var = mes_data.get("variacion", 0)
                    else:
                        presup = real = var = 0
                    pct = porcentaje_var(presup, var)
                    fila.extend([presup, real, var, pct])
                fila.extend(
                    [
                        tipo.get("total_anual_presup", 0),
                        tipo.get("total_anual_gasto", 0),
                        tipo.get("total_anual_var", 0),
                        tipo.get("total_anual_pct", ""),
                    ]
                )
                ws.append(fila)

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f"Comparativo_{empresa}_{anio}.xlsx".replace(" ", "_")
    response = HttpResponse(
        content=output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename={nombre_archivo}"
    return response


@login_required
def comparativo_presupuesto_anio(request):
    # 1. Obtener lista de años disponibles
    anios = list(
        Presupuesto.objects.values_list("anio", flat=True).distinct().order_by("anio")
    )

    if request.user.is_superuser:
        empresas = Empresa.objects.all()
        empresa_id = request.GET.get("empresa")
        empresa = Empresa.objects.get(pk=empresa_id) if empresa_id else empresas.first()
        is_super = True
    else:
        empresa = request.user.perfilusuario.empresa
        empresas = None
        is_super = False

    anios = list(
        Presupuesto.objects.filter(empresa=empresa)
        .values_list("anio", flat=True)
        .distinct()
        .order_by("anio")
    )

    # Si no hay registros
    if not anios:
        return render(
            request,
            "presupuestos/comparativo_anio.html",
            {
                "comparativo": {},
                "anio1": None,
                "anio2": None,
                "anios_disponibles": [],
                "totales": None,
            },
        )

    # 2. Determinar años a comparar: GET o últimos dos
    default1, default2 = (
        (anios[-2], anios[-1]) if len(anios) > 1 else (anios[0], anios[0])
    )
    try:
        anio1 = int(request.GET.get("anio1", default1))
        anio2 = int(request.GET.get("anio2", default2))
    except (TypeError, ValueError):
        anio1, anio2 = default1, default2

    # 3. Consulta y agregación
    datos = (
        Presupuesto.objects.filter(empresa=empresa, anio__in=[anio1, anio2])
        .values("grupo__nombre", "subgrupo__nombre", "tipo_gasto__nombre", "anio")
        .annotate(total=Sum("monto"))
        .order_by("grupo__nombre", "subgrupo__nombre", "tipo_gasto__nombre", "anio")
    )

    # 4. Reconstruir datos en estructura anidada con subtotales
    estructura = {}
    for item in datos:
        grp = item["grupo__nombre"]
        sub = item["subgrupo__nombre"]
        tip = item["tipo_gasto__nombre"]
        monto = item["total"]
        year = item["anio"]

        # Inicializar jerarquía
        estructura.setdefault(
            grp, {"subgrupos": {}, "totales": {"valor1": 0, "valor2": 0}}
        )
        grp_obj = estructura[grp]
        grp_obj["totales"][f"valor{1 if year==anio1 else 2}"] += monto

        subs = grp_obj["subgrupos"]
        subs.setdefault(sub, {"tipos": {}, "totales": {"valor1": 0, "valor2": 0}})
        sub_obj = subs[sub]
        sub_obj["totales"][f"valor{1 if year==anio1 else 2}"] += monto

        # Detalle por tipo de gasto
        sub_obj["tipos"][tip] = sub_obj["tipos"].get(tip, {"valor1": 0, "valor2": 0})
        sub_obj["tipos"][tip][f"valor{1 if year==anio1 else 2}"] += monto

    # Calcular variaciones en cada nivel
    # Tipos
    for grp_obj in estructura.values():
        for sub_obj in grp_obj["subgrupos"].values():
            for tip, vals in sub_obj["tipos"].items():
                v1 = vals["valor1"]
                v2 = vals["valor2"]
                diff = v2 - v1
                pct = (diff / v1 * 100) if v1 else None
                vals.update({"variacion": diff, "variacion_pct": pct})
            # Subgrupo
            v1 = sub_obj["totales"]["valor1"]
            v2 = sub_obj["totales"]["valor2"]
            diff = v2 - v1
            pct = (diff / v1 * 100) if v1 else None
            sub_obj["totales"].update({"variacion": diff, "variacion_pct": pct})
        # Grupo
        v1 = grp_obj["totales"]["valor1"]
        v2 = grp_obj["totales"]["valor2"]
        diff = v2 - v1
        pct = (diff / v1 * 100) if v1 else None
        grp_obj["totales"].update({"variacion": diff, "variacion_pct": pct})

    # 5. Totales generales
    total1 = sum(g["totales"]["valor1"] for g in estructura.values())
    total2 = sum(g["totales"]["valor2"] for g in estructura.values())
    diff = total2 - total1
    pct = (diff / total1 * 100) if total1 else None
    totales = {
        "valor1": total1,
        "valor2": total2,
        "variacion": diff,
        "variacion_pct": pct,
    }

    # 6. Renderizar
    return render(
        request,
        "presupuestos/comparativo_anio.html",
        {
            "comparativo": estructura,
            "anio1": anio1,
            "anio2": anio2,
            "anios_disponibles": anios,
            "totales": totales,
            "empresa": empresa,
            "empresas": empresas,
            "is_super": is_super,
        },
    )


@login_required
def comparativo_presupuesto_vs_gastos(request):
    # 1. Años disponibles (presupuesto.anio y pago.fecha_pago__year)
    años_presupuesto = Presupuesto.objects.values_list("anio", flat=True)
    años_pagos = PagoGasto.objects.annotate(year=ExtractYear("fecha_pago")).values_list(
        "year", flat=True
    )
    años = sorted(set(años_presupuesto) | set(años_pagos))

    # 2. Año seleccionado (por GET o último)
    default_anio = años[-1] if años else date.today().year
    try:
        anio = int(request.GET.get("anio", default_anio))
    except (TypeError, ValueError):
        anio = default_anio

    # 3. Agregaciones de presupuesto por mes y jerarquía
    qs_pres = (
        Presupuesto.objects.filter(anio=anio)
        .values(
            "grupo__nombre",  # GrupoGasto.nombre
            "subgrupo__nombre",  # SubgrupoGasto.nombre (puede ser null)
            "tipo_gasto__nombre",  # TipoGasto.nombre (puede ser null)
            "mes",  # mes (1–12 o null)
        )
        .annotate(presupuesto=Sum("monto"))
    )

    # 4. Agregaciones de pagos de gasto por mes y jerarquía
    qs_pagos = (
        PagoGasto.objects.annotate(
            year=ExtractYear("fecha_pago"), mes=ExtractMonth("fecha_pago")
        )
        .filter(year=anio)
        .values(
            "gasto__tipo_gasto__subgrupo__grupo__nombre",  # GrupoGasto.nombre
            "gasto__tipo_gasto__subgrupo__nombre",  # SubgrupoGasto.nombre
            "gasto__tipo_gasto__nombre",  # TipoGasto.nombre
            "mes",  # mes (1–12)
        )
        .annotate(gasto=Sum("monto"))
    )

    # 5. Construir estructura anidada y acumular montos
    meses = list(range(1, 13))
    estructura = {}
    # Inicializar nodos
    for rec in list(qs_pres) + list(qs_pagos):
        if "grupo__nombre" in rec:
            grp = rec["grupo__nombre"]
            sub = rec["subgrupo__nombre"]
            tip = rec["tipo_gasto__nombre"]
        else:
            grp = rec["gasto__tipo_gasto__subgrupo__grupo__nombre"]
            sub = rec["gasto__tipo_gasto__subgrupo__nombre"]
            tip = rec["gasto__tipo_gasto__nombre"]
        estructura.setdefault(
            grp, {"subgrupos": {}, "totales": {"presupuesto": 0, "gasto": 0}}
        )
        estructura[grp]["subgrupos"].setdefault(
            sub, {"tipos": {}, "totales": {"presupuesto": 0, "gasto": 0}}
        )
        estructura[grp]["subgrupos"][sub]["tipos"].setdefault(
            tip, {"presupuesto": {m: 0 for m in meses}, "gasto": {m: 0 for m in meses}}
        )

    # Rellenar valores de presupuesto
    for item in qs_pres:
        grp = item["grupo__nombre"]
        sub = item["subgrupo__nombre"]
        tip = item["tipo_gasto__nombre"]
        m = item["mes"] or 0
        val = item["presupuesto"]
        estructura[grp]["subgrupos"][sub]["tipos"][tip]["presupuesto"][m] = val
        estructura[grp]["subgrupos"][sub]["totales"]["presupuesto"] += val
        estructura[grp]["totales"]["presupuesto"] += val

    # Rellenar valores de gasto
    for item in qs_pagos:
        grp = item["gasto__tipo_gasto__subgrupo__grupo__nombre"]
        sub = item["gasto__tipo_gasto__subgrupo__nombre"]
        tip = item["gasto__tipo_gasto__nombre"]
        m = item["mes"] or 0
        val = item["gasto"]
        estructura[grp]["subgrupos"][sub]["tipos"][tip]["gasto"][m] = val
        estructura[grp]["subgrupos"][sub]["totales"]["gasto"] += val
        estructura[grp]["totales"]["gasto"] += val

    # 6. Calcular totales mensuales y anuales
    totales_meses = {m: {"presupuesto": 0, "gasto": 0} for m in meses}
    tot_anual = {"presupuesto": 0, "gasto": 0}
    for grp_obj in estructura.values():
        for sub_obj in grp_obj["subgrupos"].values():
            for tipo_vals in sub_obj["tipos"].values():
                for m in meses:
                    totales_meses[m]["presupuesto"] += tipo_vals["presupuesto"][m]
                    totales_meses[m]["gasto"] += tipo_vals["gasto"][m]
    tot_anual["presupuesto"] = sum(v["presupuesto"] for v in totales_meses.values())
    tot_anual["gasto"] = sum(v["gasto"] for v in totales_meses.values())

    # 7. Renderizar
    return render(
        request,
        "presupuestos/comparativo_presupuesto_vs_gastos.html",
        {
            "estructura": estructura,
            "meses": meses,
            "totales_meses": totales_meses,
            "tot_anual": tot_anual,
            "anio": anio,
            "anios_disponibles": años,
        },
    )


@login_required
def descargar_plantilla_matriz_presupuesto(request):
    # Obtén todos los tipos de gasto ordenados por grupo y subgrupo
    tipos = TipoGasto.objects.select_related("subgrupo", "subgrupo__grupo").order_by(
        "subgrupo__grupo__nombre", "subgrupo__nombre", "nombre"
    )
    meses = list(range(1, 13))
    meses_nombres = [
        "Ene",
        "Feb",
        "Mar",
        "Abr",
        "May",
        "Jun",
        "Jul",
        "Ago",
        "Sep",
        "Oct",
        "Nov",
        "Dic",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carga Presupuesto"

    # Encabezados
    encabezados = (
        ["Grupo", "Subgrupo", "Tipo de Gasto"] + meses_nombres + ["Año", "Empresa"]
    )
    ws.append(encabezados)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Una fila por cada tipo de gasto, columnas para cada mes (importes vacíos)
    for tipo in tipos:
        fila = [
            str(tipo.subgrupo.grupo),
            str(tipo.subgrupo),
            str(tipo),
        ]
        # 12 columnas de meses vacías (el usuario debe llenarlas)
        fila += [""] * 12
        # Año sugerido (puedes cambiarlo por el año actual si quieres)
        fila.append("")
        # Empresa sugerida (el usuario debe llenarla)
        fila.append("")
        ws.append(fila)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Respuesta HTTP para descargar
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        "attachment; filename=plantilla_matriz_presupuesto.xlsx"
    )
    wb.save(response)
    return response


@login_required
# presupuesto gastos
def carga_masiva_presupuestos(request):
    if request.method == "POST":
        form = PresupuestoCargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES["archivo"]
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            errores = []
            exitos = 0
            meses = list(range(1, 13))
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Espera: Grupo, Subgrupo, Tipo de Gasto, Ene, Feb, ..., Dic, Año, Empresa
                if len(row) < 17:
                    errores.append(
                        f"Fila {i}: Número de columnas insuficiente ({len(row)})"
                    )
                    continue
                grupo_val, subgrupo_val, tipo_val = row[0], row[1], row[2]
                meses_valores = row[3:15]
                anio = row[15]
                empresa_val = row[16]
                try:
                    empresa = Empresa.objects.filter(
                        nombre__iexact=str(empresa_val).strip()
                    ).first()
                    if not empresa:
                        errores.append(
                            f"Fila {i}: Empresa '{empresa_val}' no encontrada."
                        )
                        continue

                    grupo = GrupoGasto.objects.filter(
                        nombre__iexact=str(grupo_val).strip(), empresa=empresa
                    ).first()
                    if not grupo:
                        errores.append(
                            f"Fila {i}: Grupo '{grupo_val}' no encontrado para la empresa."
                        )
                        continue

                    def extrae_nombre(valor, sep="/"):
                        # Devuelve la última parte después del separador (quita grupo/subgrupo si viene junto)."""
                        if not valor:
                            return ""
                        partes = str(valor).split(sep)
                        return partes[-1].strip()

                    def normaliza(texto):
                        return "".join(str(texto).strip().lower().split())

                    # Buscar subgrupo de forma robusta
                    subgrupo_nombre = extrae_nombre(subgrupo_val, "/")
                    subgrupo = None
                    for sg in SubgrupoGasto.objects.filter(grupo=grupo):
                        if normaliza(sg.nombre) == normaliza(subgrupo_nombre):
                            subgrupo = sg
                            break
                    if not subgrupo:
                        errores.append(
                            f"Fila {i}: Subgrupo '{subgrupo_val}' no encontrado en grupo '{grupo_val}'."
                        )
                        continue

                    # Buscar tipo de gasto de forma robusta
                    tipo_nombre = extrae_nombre(tipo_val, "-")
                    tipo_gasto = None
                    for tg in TipoGasto.objects.filter(subgrupo=subgrupo):
                        if normaliza(tg.nombre) == normaliza(tipo_nombre):
                            tipo_gasto = tg
                            break
                    if not tipo_gasto:
                        errores.append(
                            f"Fila {i}: Tipo de gasto '{tipo_val}' no encontrado en subgrupo '{subgrupo_val}'."
                        )
                        continue

                    try:
                        anio_int = int(anio)
                    except (TypeError, ValueError):
                        errores.append(f"Fila {i}: Año '{anio}' no válido.")
                        continue

                    for idx, monto in enumerate(meses_valores):
                        mes = idx + 1
                        if monto is None or monto == "":
                            continue
                        try:
                            monto_decimal = Decimal(monto)
                        except (InvalidOperation, TypeError, ValueError):
                            errores.append(
                                f"Fila {i}, mes {mes}: Monto '{monto}' no válido."
                            )
                            continue
                        obj, created = Presupuesto.objects.update_or_create(
                            empresa=empresa,
                            grupo=grupo,
                            subgrupo=subgrupo,
                            tipo_gasto=tipo_gasto,
                            anio=anio_int,
                            mes=mes,
                            defaults={"monto": monto_decimal},
                        )
                        exitos += 1
                except Exception as e:
                    errores.append(f"Fila {i}: {str(e)}")
            if exitos:
                messages.success(
                    request,
                    f"{exitos} presupuestos cargados/actualizados exitosamente.",
                )
            if errores:
                messages.error(request, "Errores:<br>" + "<br>".join(errores))
            return redirect("carga_masiva_presupuestos")
    else:
        form = PresupuestoCargaMasivaForm()
    return render(
        request, "presupuestos/carga_masiva_presupuestos.html", {"form": form}
    )


@login_required
def reporte_presupuesto_vs_ingreso(request):
    anio = int(request.GET.get("anio", datetime.now().year))
    meses = list(range(1, 13))
    meses_nombres = [calendar.month_name[m] for m in meses]
    medicion = request.GET.get(
        "medicion", "curso"
    )  # 'mes', 'curso', 'semestre1', 'semestre2', 'anual'

    # Selección de meses según filtro
    if medicion == "mes":
        mes_actual = int(request.GET.get("mes", datetime.now().month))
        meses = [mes_actual]
        meses_nombres = [calendar.month_name[mes_actual]]
    elif medicion == "semestre1":
        meses = list(range(1, 7))
        meses_nombres = [calendar.month_name[m] for m in meses]
    elif medicion == "semestre2":
        meses = list(range(7, 13))
        meses_nombres = [calendar.month_name[m] for m in meses]
    elif medicion == "anual":
        meses = list(range(1, 13))
        meses_nombres = [calendar.month_name[m] for m in meses]
    else:  # 'curso' o default
        if anio == datetime.now().year:
            mes_actual = datetime.now().month
        else:
            mes_actual = 12
        meses = list(range(1, mes_actual + 1))
        meses_nombres = [calendar.month_name[m] for m in meses]

    # Empresa y permisos
    if request.user.is_superuser:
        empresas = Empresa.objects.all()
        empresa_id = request.GET.get("empresa")
        empresa = Empresa.objects.get(pk=empresa_id) if empresa_id else empresas.first()
    else:
        empresa = request.user.perfilusuario.empresa
        empresas = None

    origenes = [
        ("local", "Locales"),
        ("area", "Áreas comunes"),
        ("otros", "Otros ingresos"),
    ]
    # tipos_otros = FacturaOtrosIngresos.TIPO_INGRESO
    tipos_otros = list(
        TipoOtroIngreso.objects.filter(empresa=empresa).values_list("id", "nombre")
    )
    # Presupuestos
    presupuestos = PresupuestoIngreso.objects.filter(empresa=empresa, anio=anio)
    presup_dict = defaultdict(lambda: defaultdict(float))  # origen/tipo -> mes -> monto

    for p in presupuestos:
        if p.origen == "otros" and p.tipo_otro:
            presup_dict[(p.origen, p.tipo_otro.id)][p.mes] = float(
                p.monto_presupuestado
            )
        else:
            presup_dict[(p.origen, None)][p.mes] = float(p.monto_presupuestado)

    # Ingresos reales
    pagos_local = Pago.objects.filter(
        factura__local__isnull=False, fecha_pago__year=anio
    )
    pagos_area = Pago.objects.filter(
        factura__area_comun__isnull=False, fecha_pago__year=anio
    )
    otros = CobroOtrosIngresos.objects.filter(fecha_cobro__year=anio)

    if empresa:
        pagos_local = pagos_local.filter(factura__empresa=empresa)
        pagos_area = pagos_area.filter(factura__empresa=empresa)
        otros = otros.filter(factura__empresa=empresa)

    reales_dict = defaultdict(lambda: defaultdict(float))  # origen/tipo -> mes -> monto

    # Locales
    for p in (
        pagos_local.annotate(mes=ExtractMonth("fecha_pago"))
        .values("mes")
        .annotate(total=Sum("monto"))
    ):
        reales_dict[("local", None)][p["mes"]] = float(p["total"])
    # Áreas comunes
    for p in (
        pagos_area.annotate(mes=ExtractMonth("fecha_pago"))
        .values("mes")
        .annotate(total=Sum("monto"))
    ):
        reales_dict[("area", None)][p["mes"]] = float(p["total"])
    # Otros ingresos por tipo
    for tipo in tipos_otros:
        tipo_id = tipo[0]
        qs = otros.filter(factura__tipo_ingreso=tipo_id)
        for p in (
            qs.annotate(mes=ExtractMonth("fecha_cobro"))
            .values("mes")
            .annotate(total=Sum("monto"))
        ):
            reales_dict[("otros", tipo_id)][p["mes"]] = float(p["total"])

    # Estructura para el comparativo
    comparativo = []

    # Locales y Áreas comunes
    for origen, nombre in origenes[:2]:
        row = {
            "nombre": nombre,
            "tipo": None,
            "meses": [],
            "total_presup": 0.0,
            "total_real": 0.0,
            "total_var": 0.0,
        }
        for mes in meses:
            presup = presup_dict[(origen, None)].get(mes, 0)
            real = reales_dict[(origen, None)].get(mes, 0)
            var = real - presup
            row["meses"].append(
                {
                    "presupuesto": presup,
                    "real": real,
                    "variacion": var,
                }
            )
            row["total_presup"] += presup
            row["total_real"] += real
            row["total_var"] += var
        comparativo.append(row)

    # Otros ingresos (desglose por tipo)
    otros_rows = []
    for tipo in tipos_otros:
        tipo_id, tipo_nombre = tipo
        row = {
            "nombre": "Otros ingresos",
            "tipo": tipo_nombre,
            "meses": [],
            "total_presup": 0.0,
            "total_real": 0.0,
            "total_var": 0.0,
        }
        for mes in meses:
            presup = presup_dict[("otros", tipo_id)].get(mes, 0)
            real = reales_dict[("otros", tipo_id)].get(mes, 0)
            var = real - presup
            row["meses"].append(
                {
                    "presupuesto": presup,
                    "real": real,
                    "variacion": var,
                }
            )
            row["total_presup"] += presup
            row["total_real"] += real
            row["total_var"] += var
        otros_rows.append(row)

    # Fila total de "Otros ingresos" sumando todos los tipos
    total_otros = {
        "nombre": "Otros ingresos (Total)",
        "tipo": None,
        "meses": [],
        "total_presup": 0.0,
        "total_real": 0.0,
        "total_var": 0.0,
    }
    for i, mes in enumerate(meses):
        presup = sum(row["meses"][i]["presupuesto"] for row in otros_rows)
        real = sum(row["meses"][i]["real"] for row in otros_rows)
        var = real - presup
        total_otros["meses"].append(
            {
                "presupuesto": presup,
                "real": real,
                "variacion": var,
            }
        )
        total_otros["total_presup"] += presup
        total_otros["total_real"] += real
        total_otros["total_var"] += var

    comparativo.append(total_otros)
    comparativo.extend(otros_rows)  # Si quieres mostrar el desglose debajo del total

    def es_numero(x):
        return isinstance(x, (int, float)) and not isinstance(x, bool)

    total_general_presup = sum(
        row["total_presup"]
        for row in comparativo
        if row["tipo"] is None and es_numero(row["total_presup"])
    )
    total_general_real = sum(
        row["total_real"]
        for row in comparativo
        if row["tipo"] is None and es_numero(row["total_real"])
    )
    total_general_var = sum(
        row["total_var"]
        for row in comparativo
        if row["tipo"] is None and es_numero(row["total_var"])
    )

    # Totales generales por mes (solo filas principales)
    filas_principales = [row for row in comparativo if row["tipo"] is None]
    tot_gen_presup = []
    tot_gen_real = []
    tot_gen_var = []
    for i in range(len(meses)):
        presup = sum(
            row["meses"][i]["presupuesto"]
            for row in filas_principales
            if es_numero(row["meses"][i]["presupuesto"])
        )
        real = sum(
            row["meses"][i]["real"]
            for row in filas_principales
            if es_numero(row["meses"][i]["real"])
        )
        var = real - presup
        tot_gen_presup.append(presup)
        tot_gen_real.append(real)
        tot_gen_var.append(var)

    # --- NUEVO: Totales como diccionario para el template ---
    totales_meses = {}
    for i, mes in enumerate(meses):
        totales_meses[mes] = {
            "presupuesto": tot_gen_presup[i],
            "real": tot_gen_real[i],
            "variacion": tot_gen_var[i],
        }

    tot_anual = {
        "presupuesto": total_general_presup,
        "real": total_general_real,
        "variacion": total_general_var,
    }
    # Exportar a Excel
    if request.GET.get("excel") == "1":
        return exportar_reporte_presupuesto_vs_ingreso(request)

    return render(
        request,
        "presupuestos/comparativo_presupuesto_vs_ingresos.html",
        {
            "anio": anio,
            "anios": list(range(datetime.now().year, 2021, -1)),
            "empresa": empresa,
            "empresas": empresas,
            "meses": meses,
            "meses_nombres": meses_nombres,
            "comparativo": comparativo,
            "totales_meses": totales_meses,
            "tot_anual": tot_anual,
            "medicion": medicion,
        },
    )


@login_required
def exportar_reporte_presupuesto_vs_ingreso(request):
    from collections import defaultdict

    anio = int(request.GET.get("anio", datetime.now().year))
    medicion = request.GET.get("medicion", "curso")
    # --- Copia la lógica de selección de meses de tu vista principal ---
    if medicion == "mes":
        mes_actual = int(request.GET.get("mes", datetime.now().month))
        meses = [mes_actual]
        meses_nombres = [calendar.month_name[mes_actual]]
    elif medicion == "semestre1":
        meses = list(range(1, 7))
        meses_nombres = [calendar.month_name[m] for m in meses]
    elif medicion == "semestre2":
        meses = list(range(7, 13))
        meses_nombres = [calendar.month_name[m] for m in meses]
    elif medicion == "anual":
        meses = list(range(1, 13))
        meses_nombres = [calendar.month_name[m] for m in meses]
    else:  # 'curso' o default
        if anio == datetime.now().year:
            mes_actual = datetime.now().month
        else:
            mes_actual = 12
        meses = list(range(1, mes_actual + 1))
        meses_nombres = [calendar.month_name[m] for m in meses]

    # Empresa y permisos
    if request.user.is_superuser:
        empresas = Empresa.objects.all()
        empresa_id = request.GET.get("empresa")
        empresa = Empresa.objects.get(pk=empresa_id) if empresa_id else empresas.first()
    else:
        empresa = request.user.perfilusuario.empresa

    origenes = [
        ("local", "Locales"),
        ("area", "Áreas comunes"),
        ("otros", "Otros ingresos"),
    ]
    tipos_otros = FacturaOtrosIngresos.TIPO_INGRESO

    # Presupuestos
    presupuestos = PresupuestoIngreso.objects.filter(empresa=empresa, anio=anio)
    presup_dict = defaultdict(lambda: defaultdict(float))
    for p in presupuestos:
        if p.origen == "otros" and p.tipo_otro:
            presup_dict[(p.origen, p.tipo_otro)][p.mes] = float(p.monto_presupuestado)
        else:
            presup_dict[(p.origen, None)][p.mes] = float(p.monto_presupuestado)

    # Ingresos reales
    pagos_local = Pago.objects.filter(
        factura__local__isnull=False, fecha_pago__year=anio
    )
    pagos_area = Pago.objects.filter(
        factura__area_comun__isnull=False, fecha_pago__year=anio
    )
    otros = CobroOtrosIngresos.objects.filter(fecha_cobro__year=anio)
    if empresa:
        pagos_local = pagos_local.filter(factura__empresa=empresa)
        pagos_area = pagos_area.filter(factura__empresa=empresa)
        otros = otros.filter(factura__empresa=empresa)

    reales_dict = defaultdict(lambda: defaultdict(float))
    for p in (
        pagos_local.annotate(mes=ExtractMonth("fecha_pago"))
        .values("mes")
        .annotate(total=Sum("monto"))
    ):
        reales_dict[("local", None)][p["mes"]] = float(p["total"])
    for p in (
        pagos_area.annotate(mes=ExtractMonth("fecha_pago"))
        .values("mes")
        .annotate(total=Sum("monto"))
    ):
        reales_dict[("area", None)][p["mes"]] = float(p["total"])
    for tipo in tipos_otros:
        tipo_id = tipo[0]
        qs = otros.filter(factura__tipo_ingreso=tipo_id)
        for p in (
            qs.annotate(mes=ExtractMonth("fecha_cobro"))
            .values("mes")
            .annotate(total=Sum("monto"))
        ):
            reales_dict[("otros", tipo_id)][p["mes"]] = float(p["total"])

    # --- Crear el archivo Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto vs Ingresos"

    # Encabezado
    encabezado = ["Origen", "Tipo"]
    for mes_nombre in meses_nombres:
        encabezado += [
            f"{mes_nombre} Presup.",
            f"{mes_nombre} Real",
            f"{mes_nombre} Var.",
        ]
    encabezado += ["Total Presup.", "Total Real", "Total Var."]
    ws.append(encabezado)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Locales y Áreas comunes
    for origen, nombre in origenes[:2]:
        fila = [nombre, ""]
        total_presup = total_real = total_var = 0
        for mes in meses:
            presup = presup_dict[(origen, None)].get(mes, 0)
            real = reales_dict[(origen, None)].get(mes, 0)
            var = real - presup
            fila += [presup, real, var]
            total_presup += presup
            total_real += real
            total_var += var
        fila += [total_presup, total_real, total_var]
        ws.append(fila)

    # Otros ingresos por tipo
    for tipo in tipos_otros:
        tipo_id, tipo_nombre = tipo
        fila = ["Otros ingresos", tipo_nombre]
        total_presup = total_real = total_var = 0
        for mes in meses:
            presup = presup_dict[("otros", tipo_id)].get(mes, 0)
            real = reales_dict[("otros", tipo_id)].get(mes, 0)
            var = real - presup
            fila += [presup, real, var]
            total_presup += presup
            total_real += real
            total_var += var
        fila += [total_presup, total_real, total_var]
        ws.append(fila)

    # Fila total de "Otros ingresos" sumando todos los tipos
    fila_total_otros = ["Otros ingresos (Total)", ""]
    total_presup = total_real = total_var = 0
    for i, mes in enumerate(meses):
        presup = sum(
            presup_dict[("otros", tipo[0])].get(mes, 0) for tipo in tipos_otros
        )
        real = sum(reales_dict[("otros", tipo[0])].get(mes, 0) for tipo in tipos_otros)
        var = real - presup
        fila_total_otros += [presup, real, var]
        total_presup += presup
        total_real += real
        total_var += var
    fila_total_otros += [total_presup, total_real, total_var]
    ws.append(fila_total_otros)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Nombre del archivo
    nombre_archivo = f"Presupuesto_vs_Ingresos_{empresa.nombre}_{anio}.xlsx".replace(
        " ", "_"
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={nombre_archivo}"
    wb.save(response)
    return response


@login_required
def matriz_presupuesto_ingresos(request):
    anio = int(request.GET.get("anio", now().year))
    now_year = now().year
    # anios = list(range(now_year, 2023, -1))
    anios = (
        PresupuestoIngreso.objects.values_list("anio", flat=True)
        .distinct()
        .order_by("anio")
    )

    # Empresa y permisos
    if request.user.is_superuser:
        empresas = Empresa.objects.all()
        empresa_id = request.GET.get("empresa")
        empresa = Empresa.objects.get(pk=empresa_id) if empresa_id else empresas.first()
    else:
        empresa = request.user.perfilusuario.empresa
        empresas = None

    meses = list(range(1, 13))
    meses_nombres = [month_name[m].capitalize() for m in meses]
    origenes = [
        ("local", "Locales"),
        ("area", "Áreas comunes"),
        ("otros", "Otros ingresos"),
    ]

    # Catálogo de tipos de otros ingresos
    # tipos_otros = FacturaOtrosIngresos.TIPO_INGRESO
    tipos_otros = list(
        TipoOtroIngreso.objects.filter(empresa=empresa).values_list("id", "nombre")
    )

    # ¿Cerrado?
    cierre = PresupuestoCierre.objects.filter(empresa=empresa, anio=anio).first()
    bloqueado = cierre.cerrado if cierre else False
    pedir_superuser = False

    # Lógica para abrir presupuesto con superuser
    if bloqueado and not request.user.is_superuser:
        if request.method == "POST" and "superuser_username" in request.POST:
            username = request.POST.get("superuser_username")
            password = request.POST.get("superuser_password")
            superuser = authenticate(username=username, password=password)
            if superuser and superuser.is_superuser:
                cierre.cerrado = False
                cierre.fecha_cierre = None
                cierre.cerrado_por = None
                cierre.save()
                messages.success(
                    request,
                    "Presupuesto reabierto correctamente. Ahora puedes editarlo.",
                )
                return redirect(
                    request.path
                    + f"?anio={anio}"
                    + (f"&empresa={empresa.id}" if empresa else "")
                )
            else:
                messages.error(
                    request, "Usuario o contraseña de superusuario incorrectos."
                )
        else:
            pedir_superuser = True

    cierre = PresupuestoCierre.objects.filter(empresa=empresa, anio=anio).first()
    bloqueado = cierre.cerrado if cierre else False
    if not bloqueado:
        pedir_superuser = False

    edicion_habilitada = not bloqueado or request.user.is_superuser

    # --- Estructura de la matriz ---
    presupuestos = PresupuestoIngreso.objects.filter(empresa=empresa, anio=anio)
    presup_dict = defaultdict(dict)
    otros_dict = defaultdict(lambda: defaultdict(float))  # tipo_otro_id -> mes -> monto
    # for p in presupuestos:
    #   presup_dict[p.origen][p.mes] = p.monto_presupuestado
    for p in presupuestos:
        if p.origen == "otros" and p.tipo_otro:
            otros_dict[p.tipo_otro.id][p.mes] = p.monto_presupuestado
        else:
            presup_dict[p.origen][p.mes] = p.monto_presupuestado

    subtotales_origen = {}
    # for origen, _ in origenes:
    #   subtotales_origen[origen] = [presup_dict[origen].get(mes, 0) for mes in meses]
    for origen, _ in origenes:
        if origen == "otros":
            # Suma de todos los tipos de otros ingresos
            subtotales_origen[origen] = [
                sum(otros_dict[tipo_val[0]].get(mes, 0) for tipo_val in tipos_otros)
                for mes in meses
            ]
        else:
            subtotales_origen[origen] = [
                presup_dict[origen].get(mes, 0) for mes in meses
            ]

    # Calcula totales_mes correctamente
    totales_mes = []
    for idx, mes in enumerate(meses):
        total_mes = sum(subtotales_origen[origen][idx] for origen, _ in origenes)
        totales_mes.append(total_mes)

    # Guardado de presupuestos (solo si habilitado)
    if request.method == "POST" and edicion_habilitada:
        for origen, _ in origenes:
            if origen != "otros":
                for mes in meses:
                    key = f"presupuesto_{origen}_{mes}"
                    monto = request.POST.get(key)
                    if monto is not None:
                        monto = float(monto or 0)
                        if not empresa:
                            messages.error(
                                request,
                                "No hay empresa seleccionada. No se puede guardar el presupuesto de ingresos.",
                            )
                            return redirect(request.path + f"?anio={anio}")
                        # Usar get_or_create para evitar duplicados
                        obj, created = PresupuestoIngreso.objects.get_or_create(
                            empresa=empresa,
                            anio=anio,
                            mes=mes,
                            origen=origen,
                            defaults={"monto_presupuestado": monto},
                        )
                        if not created and obj.monto_presupuestado != monto:
                            obj.monto_presupuestado = monto
                            obj.save()
        # Guardar tipos de otros ingresos
        for tipo in tipos_otros:
            tipo_val = tipo[0]
            tipo_otro_instance = TipoOtroIngreso.objects.get(pk=tipo_val)
            for mes in meses:
                key = f"presupuesto_otros_{tipo_val}_{mes}"
                monto = request.POST.get(key)
                if monto is not None:
                    monto = float(monto or 0)
                    obj, created = PresupuestoIngreso.objects.get_or_create(
                        empresa=empresa,
                        anio=anio,
                        mes=mes,
                        origen="otros",
                        tipo_otro=tipo_otro_instance,
                        defaults={"monto_presupuestado": monto},
                    )
                    if not created and obj.monto_presupuestado != monto:
                        obj.monto_presupuestado = monto
                        obj.save()
        # Cerrar presupuesto solo si corresponde
        if "cerrar_presupuesto" in request.POST and edicion_habilitada:
            if not empresa:
                messages.error(
                    request,
                    "No hay empresa seleccionada. No se puede cerrar el presupuesto de ingresos.",
                )
                return redirect(request.path + f"?anio={anio}")

            cierre, created = PresupuestoCierre.objects.get_or_create(
                empresa=empresa, anio=anio
            )
            cierre.cerrado = True
            cierre.cerrado_por = request.user
            cierre.fecha_cierre = now()
            cierre.save()
            messages.success(
                request,
                "¡Presupuesto cerrado! Solo el superusuario puede volver a abrirlo.",
            )
            return redirect(
                request.path
                + f"?anio={anio}"
                + (f"&empresa={empresa.id}" if empresa else "")
            )
        else:
            messages.success(request, "Presupuestos de ingresos actualizados")
        return redirect(
            request.path
            + f"?anio={anio}"
            + (f"&empresa={empresa.id}" if empresa else "")
        )

    return render(
        request,
        "presupuestos/matriz_ingresos.html",
        {
            "origenes": origenes,
            "meses": meses,
            "meses_nombres": meses_nombres,
            "presup_dict": presup_dict,
            "anio": anio,
            "anios": anios,
            "empresas": empresas,
            "empresa": empresa,
            "totales_mes": totales_mes,
            "subtotales_origen": subtotales_origen,
            "is_super": request.user.is_superuser,
            "edicion_habilitada": edicion_habilitada,
            "pedir_superuser": pedir_superuser,
            "bloqueado": bloqueado,
            "cierre": cierre,
            "tipos_otros": tipos_otros,
            "otros_dict": otros_dict,
        },
    )


@login_required
def carga_masiva_presupuesto_ingresos(request):

    if request.method == "POST":
        form = PresupuestoCargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES["archivo"]
            wb = load_workbook(archivo)
            ws = wb.active
            errores = []
            exitos = 0
            meses = list(range(1, 13))
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Espera: Origen, TipoOtroIngreso, Ene, Feb, ..., Dic, Año, Empresa
                if len(row) < 16:
                    errores.append(
                        f"Fila {i}: Número de columnas insuficiente ({len(row)})"
                    )
                    continue
                origen_val = str(row[0]).strip().lower() if row[0] else ""
                tipo_otro_val = str(row[1]).strip() if row[1] else None
                meses_valores = row[2:14]
                anio = row[14]
                empresa_val = row[15]
                try:
                    empresa = Empresa.objects.filter(
                        nombre__iexact=str(empresa_val).strip()
                    ).first()
                    if not empresa:
                        errores.append(
                            f"Fila {i}: Empresa '{empresa_val}' no encontrada."
                        )
                        continue

                    if origen_val not in ["local", "area", "otros"]:
                        errores.append(f"Fila {i}: Origen '{origen_val}' no válido.")
                        continue

                    try:
                        anio_int = int(anio)
                    except (TypeError, ValueError):
                        errores.append(f"Fila {i}: Año '{anio}' no válido.")
                        continue

                    for idx, monto in enumerate(meses_valores):
                        mes = idx + 1
                        if monto is None or monto == "":
                            continue
                        try:
                            monto_decimal = Decimal(monto)
                        except (InvalidOperation, TypeError, ValueError):
                            errores.append(
                                f"Fila {i}, mes {mes}: Monto '{monto}' no válido."
                            )
                            continue
                        if origen_val == "otros":
                            tipo_otro_obj = None
                            if tipo_otro_val:
                                tipo_otro_obj = TipoOtroIngreso.objects.filter(
                                    nombre__iexact=tipo_otro_val, empresa=empresa
                                ).first()
                                if not tipo_otro_obj:
                                    errores.append(
                                        f"Fila {i}: Tipo de otro ingreso '{tipo_otro_val}' no encontrado."
                                    )
                                    continue
                            obj, created = PresupuestoIngreso.objects.update_or_create(
                                empresa=empresa,
                                anio=anio_int,
                                mes=mes,
                                origen="otros",
                                tipo_otro=tipo_otro_obj,
                                defaults={"monto_presupuestado": monto_decimal},
                            )
                        else:
                            obj, created = PresupuestoIngreso.objects.update_or_create(
                                empresa=empresa,
                                anio=anio_int,
                                mes=mes,
                                origen=origen_val,
                                defaults={"monto_presupuestado": monto_decimal},
                            )
                        exitos += 1
                except Exception as e:
                    errores.append(f"Fila {i}: {str(e)}")
            if exitos:
                messages.success(
                    request,
                    f"{exitos} presupuestos de ingresos cargados/actualizados exitosamente.",
                )
            if errores:
                messages.error(request, "Errores:<br>" + "<br>".join(errores))
            return redirect("carga_masiva_presupuesto_ingresos")
    else:
        form = PresupuestoCargaMasivaForm()
    return render(
        request,
        "presupuestos/carga_masiva_presupuesto_ingresos.html",
        {"form": form},
    )


@login_required
def descargar_plantilla_matriz_presupuesto_ingresos(request):
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    meses_nombres = [
        "Ene",
        "Feb",
        "Mar",
        "Abr",
        "May",
        "Jun",
        "Jul",
        "Ago",
        "Sep",
        "Oct",
        "Nov",
        "Dic",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carga Presupuesto Ingresos"

    # Encabezados
    encabezados = ["Origen", "TipoOtroIngreso"] + meses_nombres + ["Año", "Empresa"]
    ws.append(encabezados)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Ejemplo de fila (puedes dejarla vacía si prefieres)
    ws.append(
        [
            "local",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "2025",
            "MiEmpresa",
        ]
    )
    ws.append(
        [
            "area",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "2025",
            "MiEmpresa",
        ]
    )
    ws.append(
        [
            "otros",
            "Renta de salón",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "2025",
            "MiEmpresa",
        ]
    )

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Respuesta HTTP para descargar
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        "attachment; filename=plantilla_matriz_presupuesto_ingresos.xlsx"
    )
    wb.save(response)
    return response


@login_required
# clonar ppto gastos
def copiar_presupuesto_gastos_a_nuevo_anio(request):
    if request.method == "POST":
        anio_actual = date.today().year
        anio_nuevo = anio_actual + 1
        existe = Presupuesto.objects.filter(anio=anio_nuevo).exists()
        if existe:
            messages.warning(
                request, f"Ya existe presupuesto para el año {anio_nuevo}."
            )
            return redirect("matriz_presupuesto")
        
         # Verifica que haya datos capturados en el año actual
        presupuestos_actuales = Presupuesto.objects.filter(anio=anio_actual)
        if not presupuestos_actuales.exists():
            messages.warning(
                request, f"No hay datos capturados para el año {anio_actual}. No se puede clonar el presupuesto."
            )
            return redirect("matriz_presupuesto")

        tipo_clon = request.POST.get("tipo_clon", "sin")
        porcentaje = request.POST.get("porcentaje", "0")
        try:
            porcentaje = Decimal(porcentaje)
        except:
            porcentaje = Decimal("0")

        presupuestos_actuales = Presupuesto.objects.filter(anio=anio_actual)
        nuevos = []
        for p in presupuestos_actuales:
            monto = p.monto
            if tipo_clon == "con" and porcentaje > 0:
                monto = (
                    Decimal(monto) * (Decimal("1") + porcentaje / Decimal("100"))
                ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            nuevos.append(
                Presupuesto(
                    grupo=p.grupo,
                    subgrupo=p.subgrupo,
                    tipo_gasto=p.tipo_gasto,
                    mes=p.mes,
                    monto=monto,
                    anio=anio_nuevo,
                    empresa=p.empresa,
                )
            )

        Presupuesto.objects.bulk_create(nuevos)
        # --- Cerrar presupuesto del año anterior ---
        for empresa in set(p.empresa for p in presupuestos_actuales):
            cierre, _ = PresupuestoCierre.objects.get_or_create(
                empresa=empresa, anio=anio_actual
            )
            cierre.cerrado = True
            cierre.cerrado_por = request.user
            cierre.fecha_cierre = now()
            cierre.save()

        if tipo_clon == "con" and porcentaje > 0:
            messages.success(
                request,
                f"Presupuesto de gastos {anio_actual} copiado a {anio_nuevo} con incremento del {porcentaje}%",
            )
        else:
            messages.success(
                request,
                f"Presupuesto de gastos {anio_actual} copiado a {anio_nuevo} sin incremento.",
            )
    return redirect("matriz_presupuesto")


@login_required
# clonar ppto ingresos
def copiar_presupuesto_ingresos_a_nuevo_anio(request):
    if request.method == "POST":
        anio_actual = date.today().year
        anio_nuevo = anio_actual + 1

        existe = PresupuestoIngreso.objects.filter(anio=anio_nuevo).exists()
        if existe:
            messages.warning(
                request, f"Ya existe presupuesto de ingresos para el año {anio_nuevo}."
            )
            return redirect("matriz_presupuesto_ingresos")
        
         # Verifica que haya datos capturados en el año actual
        presupuestos_actuales = PresupuestoIngreso.objects.filter(anio=anio_actual)
        if not presupuestos_actuales.exists():
            messages.warning(
                request, f"No hay datos capturados para el año {anio_actual}. No se puede clonar el presupuesto."
            )
            return redirect("matriz_presupuesto_ingresos")


        tipo_clon = request.POST.get("tipo_clon", "sin")
        porcentaje = request.POST.get("porcentaje", "0")
        try:
            porcentaje = Decimal(porcentaje)
        except:
            porcentaje = Decimal("0")

        actuales = PresupuestoIngreso.objects.filter(anio=anio_actual)
        nuevos = []
        for p in actuales:
            monto = p.monto_presupuestado
            if tipo_clon == "con" and porcentaje > 0:
                monto = (
                    Decimal(monto) * (Decimal("1") + porcentaje / Decimal("100"))
                ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            nuevos.append(
                PresupuestoIngreso(
                    empresa=p.empresa,
                    anio=anio_nuevo,
                    mes=p.mes,
                    origen=p.origen,
                    tipo_otro=p.tipo_otro,
                    monto_presupuestado=monto,
                )
            )
        PresupuestoIngreso.objects.bulk_create(nuevos)

        # --- Cerrar presupuesto del año anterior ---
        for empresa in set(p.empresa for p in actuales):
            cierre, _ = PresupuestoCierre.objects.get_or_create(
                empresa=empresa, anio=anio_actual
            )
            cierre.cerrado = True
            cierre.cerrado_por = request.user
            cierre.fecha_cierre = now()
            cierre.save()

        if tipo_clon == "con" and porcentaje > 0:
            messages.success(
                request,
                f"Presupuesto de ingresos {anio_actual} copiado a {anio_nuevo} con incremento del {porcentaje}%",
            )
        else:
            messages.success(
                request,
                f"Presupuesto de ingresos {anio_actual} copiado a {anio_nuevo} sin incremento.",
            )
    return redirect("matriz_presupuesto_ingresos")
