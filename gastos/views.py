
from decimal import Decimal, InvalidOperation
from io import BytesIO
import unicodedata
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook
import openpyxl
from unidecode import unidecode
from empleados.models import Empleado
from empresas.models import Empresa
from facturacion.models import Pago
from presupuestos.models import Presupuesto
from proveedores.models import Proveedor
from .forms import GastoForm, GastosCargaMasivaForm, PagoGastoForm, SubgrupoGastoForm, TipoGastoForm
from .models import Gasto, GrupoGasto, PagoGasto, SubgrupoGasto, TipoGasto
from datetime import datetime
from django.utils.timezone import localtime
from django.db.models.functions import TruncMonth
from django.db.models import Sum, Q, F, Value
import calendar
from django.db.models import Q, Sum
from django.utils.dateparse import parse_date
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import ProtectedError
from django.db.models.functions import ExtractMonth

# Create your views here.
@login_required
def subgrupo_gasto_crear(request):
    if request.method == 'POST':
        form = SubgrupoGastoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('subgrupos_gasto_lista')
    else:
        form = SubgrupoGastoForm()
    return render(request, 'gastos/subgrupo_crear.html', {'form': form})


@login_required
def subgrupo_gasto_eliminar(request, pk):
    subgrupo = get_object_or_404(SubgrupoGasto, pk=pk)
    if request.method == 'POST':
        try:
            subgrupo.delete()
            messages.success(request, "Subgrupo de gasto eliminado correctamente.")
            return redirect('subgrupos_gasto_lista')
        except ProtectedError:
            messages.error(request, "No se puede eliminar este subgrupo porque tiene Tipos de Gasto o Presupuestos relacionados. Elimina o reasigna esos registros primero.")
            return redirect('subgrupos_gasto_lista')
    return render(request, 'gastos/subgrupo_gasto_confirmar_eliminar.html', {'subgrupo': subgrupo})

@login_required
def subgrupos_gasto_lista(request):
    if request.user.is_superuser:
        subgrupos = SubgrupoGasto.objects.select_related('grupo').order_by('grupo__nombre', 'nombre')
    else:
        subgrupos = SubgrupoGasto.objects.select_related('grupo').order_by('grupo__nombre', 'nombre')
    return render(request, 'gastos/subgrupos_lista.html', {'subgrupos': subgrupos})

@login_required
def tipos_gasto_lista(request):
    if request.user.is_superuser:
        tipos = TipoGasto.objects.select_related('subgrupo__grupo').all().order_by('subgrupo__grupo__nombre')
    else:
        empresa = request.user.perfilusuario.empresa
        # Filtrar tipos de gasto por empresa
        #tipos = TipoGasto.objects.filter(subgrupo__grupo__empresa=empresa).select_related('subgrupo__grupo').all().order_by('subgrupo__grupo__nombre')
        tipos = TipoGasto.objects.filter(empresa=empresa).select_related('subgrupo__grupo').order_by('subgrupo__grupo__nombre')

    return render(request, 'gastos/tipos_gasto_lista.html', {'tipos': tipos})

@login_required
def tipo_gasto_crear(request):
    user = request.user
    perfil = getattr(user, 'perfilusuario', None)
    if request.method == 'POST':
        post = request.POST.copy()
        if not user.is_superuser and perfil and perfil.empresa:
            post['empresa'] = perfil.empresa.pk
        form = TipoGastoForm(post, user=user)
        if form.is_valid():
            tipo_gasto = form.save(commit=False)
            if not user.is_superuser and perfil and perfil.empresa:
                tipo_gasto.empresa = perfil.empresa
            tipo_gasto.save()
            return redirect('tipos_gasto_lista')
    else:
        form = TipoGastoForm(user=user)
    return render(request, 'gastos/tipo_gasto_form.html', {'form': form, 'modo': 'crear'})

@login_required
def tipo_gasto_editar(request, pk):
    tipo = get_object_or_404(TipoGasto, pk=pk)
    user = request.user
    perfil = getattr(user, 'perfilusuario', None)
    if request.method == 'POST':
        post = request.POST.copy()
        if not user.is_superuser and perfil and perfil.empresa:
            post['empresa'] = perfil.empresa.pk
        form = TipoGastoForm(post, instance=tipo, user=user)
        if form.is_valid():
            tipo_gasto = form.save(commit=False)
            if not user.is_superuser and perfil and perfil.empresa:
                tipo_gasto.empresa = perfil.empresa
            tipo_gasto.save()
            return redirect('tipos_gasto_lista')
    else:
        form = TipoGastoForm(instance=tipo, user=user)
    return render(request, 'gastos/tipo_gasto_form.html', {'form': form, 'modo': 'editar', 'tipo': tipo})

@login_required
@login_required
def tipo_gasto_eliminar(request, pk):
    tipo = get_object_or_404(TipoGasto, pk=pk)
    if request.method == 'POST':
        try:
            tipo.delete()
            messages.success(request, "Tipo de gasto eliminado correctamente.")
            return redirect('tipos_gasto_lista')
        except ProtectedError:
            messages.error(
                request,
                "No se puede eliminar este tipo de gasto porque tiene registros relacionados (por ejemplo, presupuestos o movimientos). "
                "Elimina o reasigna esos registros primero."
            )
            return redirect('tipos_gasto_lista')
    return render(request, 'gastos/tipo_gasto_confirmar_eliminar.html', {'tipo': tipo})

@login_required
def gastos_lista(request):
    if request.user.is_superuser:
        gastos = Gasto.objects.all().select_related(
            'empresa', 'proveedor', 'empleado', 'tipo_gasto', 'tipo_gasto__subgrupo', 'tipo_gasto__subgrupo__grupo'
        ).prefetch_related('pagos').order_by('-fecha')  # <-- add .prefetch_related('pagos')
        proveedores = Proveedor.objects.filter(activo=True)
        empleados = Empleado.objects.filter(activo=True)
        tipos_gasto = TipoGasto.objects.all()
    else:
        empresa = request.user.perfilusuario.empresa
        gastos = Gasto.objects.filter(empresa=empresa).select_related(
            'empresa', 'proveedor', 'empleado', 'tipo_gasto', 'tipo_gasto__subgrupo', 'tipo_gasto__subgrupo__grupo'
        ).prefetch_related('pagos').order_by('-fecha')  # <-- add .prefetch_related('pagos')
        proveedores = Proveedor.objects.filter(activo=True, empresa=empresa)
        empleados = Empleado.objects.filter(activo=True, empresa=empresa)
        tipos_gasto = TipoGasto.objects.filter(empresa=empresa)

    proveedor_id = request.GET.get('proveedor')
    empleado_id = request.GET.get('empleado')
    tipo_gasto = request.GET.get('tipo_gasto')

    if proveedor_id:
        gastos = gastos.filter(proveedor_id=proveedor_id)
    if empleado_id:
        gastos = gastos.filter(empleado_id=empleado_id)
    if tipo_gasto:
        gastos = gastos.filter(tipo_gasto=tipo_gasto)

    #paginacion
    paginator = Paginator(gastos, 25)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)    

    return render(request, 'gastos/lista.html', {
        'gastos': page_obj,                                         
        'proveedores': proveedores,
        'empleados': empleados,
        'tipos_gasto': tipos_gasto,
        'proveedor_id': proveedor_id,
        'empleado_id': empleado_id,
        'tipo_gasto_sel': tipo_gasto,
    })

@login_required
#solicitudes de pago
def gasto_nuevo(request):
    if request.method == 'POST':
        form = GastoForm(request.POST or None, request.FILES, user=request.user)
        if form.is_valid():
            gasto = form.save(commit=False)
            origen = form.cleaned_data['origen_tipo']
            if origen == 'proveedor':
                gasto.empleado = None
            elif origen == 'empleado':
                gasto.proveedor = None

            if not request.user.is_superuser:
                gasto.empresa = request.user.perfilusuario.empresa
                
            gasto.estatus = 'pendiente'    
            gasto.save()
            return redirect('gastos_lista')
    else:
        form = GastoForm(user=request.user,)
        if not request.user.is_superuser:
            form.fields['empresa'].initial = request.user.perfilusuario.empresa
    return render(request, 'gastos/form.html', {'form': form, 'modo': 'crear'})

@login_required
def gasto_editar(request, pk):
    gasto = get_object_or_404(Gasto, pk=pk)
    if not request.user.is_superuser and gasto.empresa != request.user.perfilusuario.empresa:
        return redirect('gastos_lista')
    if request.method == 'POST':
        form = GastoForm(request.POST or None, request.FILES, instance=gasto, user=request.user, modo='editar')
        if form.is_valid():
            form.save()
            return redirect('gastos_lista')
    else:
        form = GastoForm(instance=gasto, user=request.user, modo='editar')
    return render(request, 'gastos/form.html', {'form': form, 'modo': 'editar', 'gasto': gasto})

@login_required
def gasto_eliminar(request, pk):
    gasto = get_object_or_404(Gasto, pk=pk)
    if request.method == 'POST':
        gasto.delete()
        return redirect('gastos_lista')
    return render(request, 'gastos/confirmar_eliminar.html', {'gasto': gasto})

@login_required
def registrar_pago_gasto(request, gasto_id):
    gasto = get_object_or_404(Gasto, pk=gasto_id)
    pagos = gasto.pagos.all()
    saldo_restante = gasto.monto - sum([p.monto for p in pagos])

    if request.method == 'POST':
        form = PagoGastoForm(request.POST,request.FILES)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.gasto = gasto
            pago.registrado_por = request.user
            if pago.monto > saldo_restante:
                form.add_error('monto', f"El monto excede el saldo pendiente (${saldo_restante:.2f})")
            else:
                pago.save()
                gasto.actualizar_estatus()
                messages.success(request, f"Pago registrado. Saldo restante: ${gasto.saldo_restante:.2f}")
                return redirect('gastos_lista')
    else:
        form = PagoGastoForm()

    return render(request, 'gastos/registrar_pago.html', {
        'form': form,
        'gasto': gasto,
        'saldo_restante': saldo_restante
    })


@login_required
def gasto_detalle(request, pk):
    gasto = get_object_or_404(Gasto, pk=pk)
    pagos = gasto.pagos.all().order_by('fecha_pago')

    return render(request, 'gastos/gasto_detalle.html', {
        'gasto': gasto,
        'pagos': pagos,
    })


@login_required
#reporte_pagos.html
def reporte_pagos_gastos(request):
    es_super = request.user.is_superuser
    pagos = PagoGasto.objects.select_related('gasto', 'gasto__empresa', 'gasto__proveedor', 'gasto__empleado')

    # Filtros
    empresas = Empresa.objects.all() if es_super else Empresa.objects.filter(pk=request.user.perfilusuario.empresa.id)
    empresa_id = request.GET.get('empresa')
    proveedor_id = request.GET.get('proveedor')
    empleado_id = request.GET.get('empleado')
    forma_pago = request.GET.get('forma_pago')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if not es_super:
        pagos = pagos.filter(gasto__empresa=request.user.perfilusuario.empresa)
        proveedores = Proveedor.objects.filter(empresa=request.user.perfilusuario.empresa)
        empleados = Empleado.objects.filter(empresa=request.user.perfilusuario.empresa)
    else:
        if empresa_id:
            pagos = pagos.filter(gasto__empresa_id=empresa_id)
            proveedores = Proveedor.objects.filter(empresa_id=empresa_id)
            empleados = Empleado.objects.filter(empresa_id=empresa_id)
        else:
            proveedores = Proveedor.objects.all()
            empleados = Empleado.objects.all()

    if proveedor_id:
        pagos = pagos.filter(gasto__proveedor_id=proveedor_id)
    if empleado_id:
        pagos = pagos.filter(gasto__empleado_id=empleado_id)
    if forma_pago:
        pagos = pagos.filter(forma_pago=forma_pago)
    if fecha_inicio:
        pagos = pagos.filter(fecha_pago__gte=parse_date(fecha_inicio))
    if fecha_fin:
        pagos = pagos.filter(fecha_pago__lte=parse_date(fecha_fin))

    total = pagos.aggregate(total=Sum('monto'))['total'] or 0

    #proveedores = Proveedor.objects.all()
    #empleados = Empleado.objects.all()
    FORMAS_PAGO = PagoGasto._meta.get_field('forma_pago').choices

    #paginacion
    paginator = Paginator(pagos, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'gastos/reporte_pagos.html', {
        'pagos': pagos,
        'empresas': empresas,
        'empresa_id': empresa_id,
        'proveedores': proveedores,
        'empleados': empleados,
        'forma_pago_actual': forma_pago,
        'total': total,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'proveedor_id': proveedor_id,
        'empleado_id': empleado_id,
        'formas_pago': FORMAS_PAGO,
        'pagos': page_obj,
    })

@login_required
#dasboard_pagos.html
def dashboard_pagos_gastos(request):
    es_super = request.user.is_superuser
    anio_actual = datetime.now().year
    mes_actual = datetime.now().month
    anio = int(request.GET.get('anio', anio_actual))
    mes= request.GET.get('mes')    

    empresas = Empresa.objects.all() if es_super else Empresa.objects.filter(id=request.user.perfilusuario.empresa.id)
    empresa_id = request.GET.get('empresa')
    proveedor_id = request.GET.get('proveedor')
    empleado_id = request.GET.get('empleado')
    forma_pago = request.GET.get('forma_pago')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

      # NUEVOS FILTROS
    grupo_id = request.GET.get('grupo')
    subgrupo_id = request.GET.get('subgrupo')
    tipo_gasto_id = request.GET.get('tipo_gasto')
    
    if es_super:
        tipos_gasto = TipoGasto.objects.all()
    else:
        empresa = request.user.perfilusuario.empresa
        tipos_gasto = TipoGasto.objects.filter(empresa=empresa)

    # --- FILTROS BÁSICOS ---
    base_gastos = Q(fecha__year=anio)
    if es_super and empresa_id:
        base_gastos &= Q(empresa_id=empresa_id)
    elif not es_super:
        base_gastos &= Q(empresa=request.user.perfilusuario.empresa)
    if proveedor_id:
        base_gastos &= Q(proveedor_id=proveedor_id)
    if empleado_id:
        base_gastos &= Q(empleado_id=empleado_id)
    if grupo_id:
        base_gastos &= Q(tipo_gasto__subgrupo__grupo_id=grupo_id)
    if subgrupo_id:
        base_gastos &= Q(tipo_gasto__subgrupo_id=subgrupo_id)
    if tipo_gasto_id:
        base_gastos &= Q(tipo_gasto_id=tipo_gasto_id)
    if mes:
        base_gastos &= Q(fecha__month=mes)      

    # OPTIMIZACIÓN: select_related para evitar N+1 en relaciones ForeignKey
    gastos = Gasto.objects.filter(base_gastos).select_related(
        'empresa', 'proveedor', 'empleado', 'tipo_gasto', 'tipo_gasto__subgrupo', 'tipo_gasto__subgrupo__grupo'
    ).prefetch_related('pagos')
    # Gastos registrados ese año y filtro empresa
    #gastos = Gasto.objects.filter(base_gastos)

    # PAGOS
    pagos = PagoGasto.objects.filter(gasto__in=gastos).select_related(
        'gasto', 'gasto__empresa', 'gasto__proveedor', 'gasto__empleado', 'gasto__tipo_gasto'
    )
    # PAGOS
    #pagos = PagoGasto.objects.filter(gasto__in=gastos)
    
    if forma_pago:
        pagos = pagos.filter(forma_pago=forma_pago)
    if fecha_inicio:
        pagos = pagos.filter(fecha_pago__gte=fecha_inicio)
    if fecha_fin:
        pagos = pagos.filter(fecha_pago__lte=fecha_fin)

    # PAGOS POR MES
    # pagos_mes = pagos.annotate(mes=TruncMonth('fecha_pago')).values('mes').annotate(total=Sum('monto')).order_by('mes')
    # pagos_mensuales = [0] * 12
    # for p in pagos_mes:
    #     mes_idx = p['mes'].month - 1
    #     pagos_mensuales[mes_idx] = float(p['total'])
    pagos_mes = pagos.annotate(mes=ExtractMonth('fecha_pago')).values('mes').annotate(total=Sum('monto')).order_by('mes')
    pagos_mensuales = [0] * 12
    for p in pagos_mes:
        mes_idx = p['mes'] - 1
        pagos_mensuales[mes_idx] = float(p['total'])

    # GASTOS POR MES (optimizado)
    gastos_por_mes = gastos.annotate(mes=ExtractMonth('fecha')).values('mes').annotate(total=Sum('monto')).order_by('mes')
    gastos_mensuales_dict = {g['mes']: float(g['total']) for g in gastos_por_mes}

    # PAGOS POR MES (optimizado)
    pagos_gastos_por_mes = pagos.annotate(mes=ExtractMonth('gasto__fecha')).values('mes').annotate(total=Sum('monto')).order_by('mes')
    pagos_gastos_mensuales_dict = {p['mes']: float(p['total']) for p in pagos_gastos_por_mes}    

    # SALDOS PENDIENTES (por mes)
    # saldos_mes = []
    # for m in range(1, 13):
    #     # Gastos registrados ese mes
    #     gastos_mes = gastos.filter(fecha__month=m)
    #     monto_gastos_mes = gastos_mes.aggregate(total=Sum('monto'))['total'] or 0
    #     pagos_gastos_mes = PagoGasto.objects.filter(
    #         gasto__in=gastos_mes,
    #     ).aggregate(total=Sum('monto'))['total'] or 0
    #     saldo = float(monto_gastos_mes) - float(pagos_gastos_mes or 0)
    #     saldos_mes.append(saldo)

    # OPTIMIZACIÓN: calcular saldos pendientes por mes directamente
    # SALDOS PENDIENTES (por mes) - solo 2 queries
    saldos_mes = []
    for m in range(1, 13):
        monto_gastos_mes = gastos_mensuales_dict.get(m, 0)
        pagos_gastos_mes = pagos_gastos_mensuales_dict.get(m, 0)
        saldo = monto_gastos_mes - pagos_gastos_mes
        saldos_mes.append(saldo)

    # PRESUPUESTO POR MES
    # presupuesto_mes = []
    # for m in range(1, 13):
    #     pres = Presupuesto.objects.filter(empresa__in=empresas, anio=anio, mes=m)
    #     if grupo_id:
    #         pres = pres.filter(tipo_gasto__subgrupo__grupo_id=grupo_id)
    #     if subgrupo_id:
    #         pres = pres.filter(tipo_gasto__subgrupo_id=subgrupo_id)
    #     if tipo_gasto_id:
    #         pres = pres.filter(tipo_gasto_id=tipo_gasto_id)
    #     pres_total = pres.aggregate(total=Sum('monto'))['total'] or 0
    #     presupuesto_mes.append(float(pres_total))

     # PRESUPUESTO POR MES (optimizado, solo 1 query)
    presupuestos_qs = Presupuesto.objects.filter(
        empresa__in=empresas,
        anio=anio
    )
    if grupo_id:
        presupuestos_qs = presupuestos_qs.filter(tipo_gasto__subgrupo__grupo_id=grupo_id)
    if subgrupo_id:
        presupuestos_qs = presupuestos_qs.filter(tipo_gasto__subgrupo_id=subgrupo_id)
    if tipo_gasto_id:
        presupuestos_qs = presupuestos_qs.filter(tipo_gasto_id=tipo_gasto_id)

    presupuestos_por_mes = presupuestos_qs.values('mes').annotate(total=Sum('monto')).order_by('mes')
    presupuesto_mes_dict = {p['mes']: float(p['total']) for p in presupuestos_por_mes}
    presupuesto_mes = [presupuesto_mes_dict.get(m, 0.0) for m in range(1, 13)]

    # --- FILTRAR SOLO MESES TRANSCURRIDOS O EL MES SELECCIONADO ---
    if mes:
        # Si hay filtro de mes, solo muestra ese mes
        mes_int = int(mes)
        meses = [calendar.month_abbr[mes_int]]
        pagos_mensuales = [pagos_mensuales[mes_int - 1]]
        saldos_mes = [saldos_mes[mes_int - 1]]
        presupuesto_mes = [presupuesto_mes[mes_int - 1]]
        meses_mostrar = 1
    else:
        if anio == anio_actual:
            meses_mostrar = mes_actual
        else:
            meses_mostrar = 12

        meses = [calendar.month_name[m] for m in range(1, meses_mostrar + 1)]
        pagos_mensuales = pagos_mensuales[:meses_mostrar]
        saldos_mes = saldos_mes[:meses_mostrar]
        presupuesto_mes = presupuesto_mes[:meses_mostrar]    

    # KPI totales
    total_pagado = sum(pagos_mensuales)
    total_pendiente = sum(saldos_mes)
    total_presupuesto = sum(presupuesto_mes)

    # Catálogos
    proveedores = Proveedor.objects.all()
    empleados = Empleado.objects.all()
    FORMAS_PAGO = PagoGasto._meta.get_field('forma_pago').choices
    grupos = GrupoGasto.objects.all()
    subgrupos = SubgrupoGasto.objects.all()
    #tipos_gasto = TipoGasto.objects.all()

    return render(request, 'gastos/dashboard_pagos.html', {
        'empresas': empresas,
        'empresa_id': empresa_id,
        'anio': anio,
        'total_pagado': total_pagado,
        'total_pendiente': total_pendiente,
        'total_presupuesto': total_presupuesto,
        'meses': meses,
        'pagos_mensuales': pagos_mensuales,
        'saldos_mes': saldos_mes,
        'presupuesto_mes': presupuesto_mes,
        'proveedores': proveedores,
        'empleados': empleados,
        'proveedor_id': proveedor_id,
        'empleado_id': empleado_id,
        'formas_pago': FORMAS_PAGO,
        'forma_pago_actual': forma_pago,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'es_super': es_super,
        'grupos': grupos,
        'subgrupos': subgrupos,
        'tipos_gasto': tipos_gasto,
        'grupo_id': grupo_id,
        'subgrupo_id': subgrupo_id,
        'tipo_gasto_id': tipo_gasto_id,
        'mes': mes,
        'anio_actual': anio_actual,
        'mes_actual': mes_actual,
        'anio_seleccionado': anio,
        'mes_seleccionado': mes,
        'meses_mostrar': meses_mostrar,

    })

#exportar pagos de gastos a Excel
@login_required
def exportar_pagos_gastos_excel(request):
    es_super = request.user.is_superuser
    anio = request.GET.get('anio')
    if anio and anio.isdigit():
        anio = int(anio)
    else:
        anio = datetime.now().year
    empresa_id = request.GET.get('empresa')
    proveedor_id = request.GET.get('proveedor')
    empleado_id = request.GET.get('empleado')
    forma_pago = request.GET.get('forma_pago')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    
    # Filtro de empresa
    if es_super and empresa_id:
        gastos = Gasto.objects.filter(empresa_id=empresa_id, fecha__year=anio)
    else:
        gastos = Gasto.objects.filter(empresa=request.user.perfilusuario.empresa, fecha__year=anio)

    # Otros filtros
    if proveedor_id:
        gastos = gastos.filter(proveedor_id=proveedor_id)
    if empleado_id:
        gastos = gastos.filter(empleado_id=empleado_id)
    if fecha_inicio:
        gastos = gastos.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        gastos = gastos.filter(fecha__lte=fecha_fin)

    # Solo los pagos
    pagos = PagoGasto.objects.filter(gasto__in=gastos)
    if forma_pago:
        pagos = pagos.filter(forma_pago=forma_pago)

    # --- Generar Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos de Gastos"
    ws.append([
        "Fecha pago", "Empresa", "Proveedor/Empleado", "Concepto", 
        "Forma de pago", "Monto", "Estatus"
    ])

    for pago in pagos.select_related('gasto', 'gasto__empresa', 'gasto__proveedor', 'gasto__empleado'):
        gasto = pago.gasto
        # Mostrar proveedor o empleado
        origen = gasto.proveedor.nombre if gasto.proveedor else (
            gasto.empleado.nombre if gasto.empleado else ''
        )
        ws.append([
            pago.fecha_pago.strftime('%d/%m/%Y') if pago.fecha_pago else '',
            gasto.empresa.nombre if gasto.empresa else '',
            origen,
            gasto.descripcion,
            pago.get_forma_pago_display(),
            float(pago.monto),
            gasto.estatus
        ])

    # --- Respuesta HTTP ---
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"pagos_gastos_{anio}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


def buscar_por_id_o_nombre(modelo, valor, campo='nombre'):
    """Busca por ID, si falla busca por nombre (sin acentos, insensible a mayúsculas). Reporta conflicto si hay varias."""
    if not valor:
        return None
    val = str(valor).strip().replace(',', '')
    try:
        return modelo.objects.get(pk=int(val))
    except (ValueError, modelo.DoesNotExist):
        todos = modelo.objects.all()
        # Lista de coincidencias insensibles a acentos y mayúsculas
        candidatos = [
            obj for obj in todos
            if unidecode(val).lower() in unidecode(str(getattr(obj, campo))).lower()
        ]
        if len(candidatos) == 1:
            return candidatos[0]
        elif len(candidatos) > 1:
            conflicto = "; ".join([f"ID={obj.pk}, {campo}='{getattr(obj, campo)}'" for obj in candidatos])
            raise Exception(f"Conflicto: '{valor}' coincide con varios registros en {modelo.__name__}: {conflicto}")
        raise Exception(f"No se encontró '{valor}' en {modelo.__name__}")


def normaliza_texto(texto):
    if not texto:
        return ""
    #texto = texto.upper()
    texto = texto
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join([c for c in texto if not unicodedata.combining(c)])
    return texto

@login_required
def carga_masiva_gastos(request):
    if request.method == 'POST':
        form = GastosCargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            errores = []
            exitos = 0
            COLUMNAS_ESPERADAS = 13  # empresa, proveedor, empleado, rfc_empleado, grupo, subgrupo, tipo_gasto, monto, descripcion, fecha, observaciones, retencion_iva, retencion_isr
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None:
                    continue
                if len(row) != COLUMNAS_ESPERADAS:
                    errores.append(f"Fila {i}: número de columnas incorrecto ({len(row)} en vez de {COLUMNAS_ESPERADAS})")
                    continue
                empresa_val, proveedor_val, empleado_val, rfc_empleado, grupo_val, subgrupo_val, tipo_gasto_val, monto, descripcion, fecha, observaciones, retencion_iva, retencion_isr = row
                try:
                    empresa = buscar_por_id_o_nombre(Empresa, empresa_val)
                    if not empresa:
                        errores.append(f"Fila {i}: No se encontró la empresa '{empresa_val}'")
                        continue

                    proveedor = None
                    if proveedor_val:
                        proveedor, _ = Proveedor.objects.get_or_create(
                            nombre=proveedor_val,
                            empresa=empresa
                        )

                    empleado = None
                    if rfc_empleado:
                        empleado, _ = Empleado.objects.get_or_create(
                            rfc=rfc_empleado,
                            defaults={'nombre': empleado_val, 'empresa': empresa}
                        )
                    elif empleado_val:
                        empleado, _ = Empleado.objects.get_or_create(
                            nombre=empleado_val,
                            empresa=empresa
                        )

                    # Validar que al menos uno esté presente
                    if not proveedor and not empleado:
                        errores.append(f"Fila {i}: Debe especificar proveedor o empleado.")
                        continue

                    # Validar que el grupo existe
                    grupo_inst = None
                    if grupo_val:
                        grupo_nombre = normaliza_texto(grupo_val)
                        try:
                            grupo_inst = GrupoGasto.objects.get(nombre=grupo_nombre)
                        except GrupoGasto.DoesNotExist:
                            errores.append(f"Fila {i}: El grupo '{grupo_nombre}' no existe en el catálogo.")
                            continue
                    else:
                        errores.append(f"Fila {i}: El grupo es obligatorio.")
                        continue

                    # Validar que el subgrupo existe para ese grupo
                    subgrupo_inst = None
                    if subgrupo_val:
                        subgrupo_nombre = normaliza_texto(subgrupo_val)
                        try:
                            subgrupo_inst = SubgrupoGasto.objects.get(nombre=subgrupo_nombre, grupo=grupo_inst)
                        except SubgrupoGasto.DoesNotExist:
                            errores.append(f"Fila {i}: El subgrupo '{subgrupo_nombre}' no existe en el grupo '{grupo_nombre}'.")
                            continue
                    else:
                        errores.append(f"Fila {i}: El subgrupo es obligatorio.")
                        continue

                    # TIPO DE GASTO por empresa y subgrupo (se puede crear si no existe)
                    tipo_gasto_inst = None
                    if tipo_gasto_val:
                        tipo_gasto_nombre = normaliza_texto(tipo_gasto_val)
                        tipo_gasto_inst = TipoGasto.objects.filter(
                            nombre=tipo_gasto_nombre, subgrupo=subgrupo_inst, empresa=empresa
                        ).first()
                        if not tipo_gasto_inst:
                            tipo_gasto_inst = TipoGasto.objects.create(
                                nombre=tipo_gasto_nombre, subgrupo=subgrupo_inst, empresa=empresa
                            )
                    else:
                        errores.append(f"Fila {i}: El tipo de gasto es obligatorio.")
                        continue

                    try:
                        monto_decimal = Decimal(monto)
                    except (InvalidOperation, TypeError, ValueError):
                        errores.append(f"Fila {i}: El valor de monto '{monto}' no es válido.")
                        continue

                    try:
                        retencion_iva_decimal = Decimal(retencion_iva) if retencion_iva else 0
                    except (InvalidOperation, TypeError, ValueError):
                        errores.append(f"Fila {i}: El valor de retención IVA '{retencion_iva}' no es válido.")
                        continue

                    try:
                        retencion_isr_decimal = Decimal(retencion_isr) if retencion_isr else 0
                    except (InvalidOperation, TypeError, ValueError):
                        errores.append(f"Fila {i}: El valor de retención ISR '{retencion_isr}' no es válido.")
                        continue

                    # Conversión de fecha si es string
                    from datetime import datetime
                    if isinstance(fecha, str):
                        try:
                            fecha = datetime.strptime(fecha, "%Y-%m-%d").date()
                        except Exception:
                            errores.append(f"Fila {i}: El formato de fecha '{fecha}' no es válido (debe ser YYYY-MM-DD).")
                            continue

                    gasto = Gasto.objects.create(
                        empresa=empresa,
                        proveedor=proveedor,
                        empleado=empleado,
                        tipo_gasto=tipo_gasto_inst,
                        monto=monto_decimal,
                        descripcion=descripcion or "",
                        fecha=fecha,
                        observaciones=observaciones or "",
                        retencion_iva=retencion_iva_decimal,
                        retencion_isr=retencion_isr_decimal,
                        estatus='pagada',
                    )

                    # Registrar el pago automáticamente
                    PagoGasto.objects.create(
                        gasto=gasto,
                        fecha_pago=fecha,
                        monto=monto_decimal,
                        forma_pago='transferencia',
                        referencia='Carga masiva',
                        registrado_por=request.user if request.user.is_authenticated else None
                    )

                    exitos += 1
                except Exception as e:
                    import traceback
                    errores.append(f"Fila {i}: {str(e) or repr(e)}<br>{traceback.format_exc()}")

            if exitos:
                messages.success(request, f"¡{exitos} gastos cargados exitosamente!")
            if errores:
                messages.error(request, "Algunos gastos no se cargaron:<br>" + "<br>".join(errores))
            return redirect('carga_masiva_gastos')
    else:
        form = GastosCargaMasivaForm()
    return render(request, 'gastos/carga_masiva_gastos.html', {'form': form})

def descargar_plantilla_gastos(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Gastos"

    # Encabezados según el formato de carga masiva
    encabezados = [
        'empresa', 'proveedor', 'empleado', 'rfc_empleado', 'grupo', 'subgrupo',
        'tipo_gasto', 'monto', 'descripcion', 'fecha', 'observaciones', 'retencion_iva', 'retencion_isr'
    ]
    ws.append(encabezados)

    # Fila de ejemplo
    ws.append([
        'EMPRESA A.C.', 'Proveedor Ejemplo', '', '', 'Gastos Administracion', 'Papelería',
        'Copias', '1200.50', 'Compra de hojas', '2025-06-19', 'Carga inicial', '0.00', '0.00'
    ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_gastos.xlsx'
    return response

@login_required
def exportar_gastos_lista_excel(request):
    proveedor_id = request.GET.get('proveedor')
    empleado_id = request.GET.get('empleado')
    tipo_gasto = request.GET.get('tipo_gasto')

    if request.user.is_superuser:
        gastos = Gasto.objects.all().select_related('empresa', 'proveedor', 'empleado', 'tipo_gasto').order_by('-fecha')
    else:
        gastos = Gasto.objects.filter(empresa=request.user.perfilusuario.empresa).order_by('-fecha')

    if proveedor_id:
        gastos = gastos.filter(proveedor_id=proveedor_id)
    if empleado_id:
        gastos = gastos.filter(empleado_id=empleado_id)
    if tipo_gasto:
        gastos = gastos.filter(tipo_gasto=tipo_gasto)

    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    ws.append([
        "Fecha", "Empresa", "Proveedor", "Empleado", "Tipo de Gasto",
        "Monto","Saldo", "Descripción", "Estatus", "Observaciones"
    ])

    for gasto in gastos:
        ws.append([
            gasto.fecha.strftime('%Y-%m-%d') if gasto.fecha else '',
            gasto.empresa.nombre if gasto.empresa else '',
            gasto.proveedor.nombre if gasto.proveedor else '',
            gasto.empleado.nombre if gasto.empleado else '',
            gasto.tipo_gasto.nombre if gasto.tipo_gasto else '',
            float(gasto.monto),
            float(gasto.saldo_restante) if gasto.saldo_restante is not None else 0.0,
            gasto.descripcion or '',
            gasto.estatus,
            gasto.observaciones or ''
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=gastos_lista.xlsx'
    wb.save(response)
    return response