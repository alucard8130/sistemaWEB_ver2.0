# Create your views here.
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from unidecode import unidecode
from empresas.models import Empresa
from .models import Cliente
from .forms import ClienteCargaMasivaForm, ClienteForm
from django.contrib.admin.views.decorators import staff_member_required
import openpyxl
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q


@login_required
def lista_clientes(request):
    query = request.GET.get("q", "")
    if request.user.is_superuser:
        clientes = Cliente.objects.filter(activo=True).order_by("id")
    else:
        empresa = request.user.perfilusuario.empresa
        clientes = Cliente.objects.filter(empresa=empresa, activo=True).order_by("id")

    if query:
        clientes = clientes.filter(
            Q(nombre__icontains=query) | Q(rfc__icontains=query)
        )

    clientes = clientes.order_by("id")    

    # paginacion
    paginator = Paginator(clientes, 20)  # 20 áreas por página
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "clientes/lista_clientes.html",
        {"clientes": clientes, "clientes": page_obj, "q": query},
    )


@login_required
def crear_cliente(request):
    perfil = getattr(request.user, "perfilusuario", None)

    if request.method == "POST":
        form = ClienteForm(request.POST, user=request.user)
        if form.is_valid():
            cliente = form.save(commit=False)
            rfc = cliente.rfc.strip() if cliente.rfc else None
            empresa = cliente.empresa if request.user.is_superuser else perfil.empresa
            if (
                rfc
                and Cliente.objects.filter(
                    rfc__iexact=rfc, empresa=empresa, activo=True
                ).exists()
            ):
                messages.error(
                    request, "Ya existe un cliente con ese RFC en la empresa."
                )
                return render(request, "clientes/crear_cliente.html", {"form": form})
            else:
                if not request.user.is_superuser and perfil:
                    cliente.empresa = perfil.empresa
                cliente.save()
                return redirect("lista_clientes")
        else:
            messages.error(
                request,
                "No se pudo crear el cliente. Por favor revisa los datos ingresados.",
            )
    else:
        form = ClienteForm(user=request.user)
        if perfil:
            form.fields["empresa"].initial = perfil.empresa

    return render(request, "clientes/crear_cliente.html", {"form": form})


@login_required
def editar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if (
        not request.user.is_superuser
        and cliente.empresa != request.user.perfilusuario.empresa
    ):
        return redirect("lista_clientes")

    if request.method == "POST":
        form = ClienteForm(request.POST, instance=cliente, user=request.user)
        if form.is_valid():
            form.save()
            return redirect("lista_clientes")
    else:
        form = ClienteForm(instance=cliente, user=request.user)

    return render(
        request, "clientes/editar_cliente.html", {"form": form, "cliente": cliente}
    )


@login_required
def eliminar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if (
        not request.user.is_superuser
        and cliente.empresa != request.user.perfilusuario.empresa
    ):
        return redirect("lista_clientes")

    if request.method == "POST":
        cliente.activo = False
        cliente.save()
        return redirect("lista_clientes")

    return render(request, "clientes/eliminar_cliente.html", {"cliente": cliente})


def buscar_empresa(valor):
    if not valor:
        raise Exception("Empresa vacía")
    val = str(valor).strip().replace(",", "")
    try:
        return Empresa.objects.get(pk=int(val))
    except (ValueError, Empresa.DoesNotExist):
        todos = Empresa.objects.all()
        candidatos = [
            e for e in todos if unidecode(val).lower() in unidecode(e.nombre).lower()
        ]
        if len(candidatos) == 1:
            return candidatos[0]
        elif len(candidatos) > 1:
            conflicto = "; ".join(
                [f"ID={e.pk}, nombre='{e.nombre}'" for e in candidatos]
            )
            raise Exception(
                f"Conflicto: '{valor}' coincide con varios registros en Empresa: {conflicto}"
            )
        raise Exception(f"No se encontró '{valor}' en Empresa")


@staff_member_required
def carga_masiva_clientes(request):
    if request.method == "POST":
        form = ClienteCargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES["archivo"]
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            errores = []
            exitos = 0
            COLUMNAS_ESPERADAS = 5  # Cambia según tus columnas
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None:
                    continue
                if len(row) != COLUMNAS_ESPERADAS:
                    errores.append(
                        f"Fila {i}: número de columnas incorrecto ({len(row)} en vez de {COLUMNAS_ESPERADAS})"
                    )
                    continue
                empresa_val, nombre, rfc, telefono, email = row
                try:
                    empresa = buscar_empresa(empresa_val)
                    if not nombre:
                        raise Exception("Nombre vacío")
                    if (
                        rfc
                        and Cliente.objects.filter(
                            rfc__iexact=rfc.strip(), empresa=empresa
                        ).exists()
                    ):
                        raise Exception(f"RFC duplicado para empresa: {rfc}")
                    Cliente.objects.create(
                        empresa=empresa,
                        nombre=nombre,
                        rfc=rfc.strip() if rfc else None,
                        telefono=telefono,
                        email=email,
                        activo=True,
                    )
                    exitos += 1
                except Exception as e:
                    errores.append(f"Fila {i}: {e}")

            if exitos:
                messages.success(request, f"¡{exitos} clientes cargados exitosamente!")
            if errores:
                messages.error(
                    request,
                    "Algunos clientes no se cargaron:<br>" + "<br>".join(errores),
                )
            return redirect("carga_masiva_clientes")
    else:
        form = ClienteCargaMasivaForm()
    return render(request, "clientes/carga_masiva_clientes.html", {"form": form})


@staff_member_required
def plantilla_clientes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Clientes"
    ws.append(["empresa", "nombre", "rfc", "telefono", "email"])
    ws.append(
        ["Torre Reforma", "Juan Pérez", "JUPE800101ABC", "5551234567", "juan@email.com"]
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=plantilla_clientes.xlsx"
    wb.save(response)
    return response


@login_required
def clientes_inactivos(request):
    if request.user.is_superuser:
        clientes = Cliente.objects.filter(activo=False)
    else:
        empresa = request.user.perfilusuario.empresa
        clientes = Cliente.objects.filter(activo=False, empresa=empresa)
    return render(request, "clientes/clientes_inactivos.html", {"clientes": clientes})


def reactivar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk, activo=False)

    if request.method == "POST":
        cliente.activo = True
        cliente.save()
        return redirect("clientes_inactivos")

    return render(request, "clientes/reactivar_confirmacion.html", {"cliente": cliente})
