
# Create your views here.
from decimal import Decimal
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
import openpyxl
from clientes.models import Cliente
from empresas.models import Empresa
from locales.forms import LocalComercialForm
from principal.models import AuditoriaCambio
from .models import AreaComun
from .forms import AreaComunCargaMasivaForm, AreaComunForm, AsignarClienteForm
from unidecode import unidecode
from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.db.models import Q




@login_required
def lista_areas(request):
    user = request.user
    query = request.GET.get("q", "")
    if user.is_superuser:
        areas = AreaComun.objects.filter(activo=True)
    else:
        empresa = user.perfilusuario.empresa
        areas = AreaComun.objects.filter(empresa=empresa, activo=True)

    if query:
        areas = areas.filter(
            Q(numero__icontains=query) | Q(cliente__nombre__icontains=query)
        )

    areas = areas.order_by('numero')
    paginator = Paginator(areas, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'areas/lista_areas.html', {'areas': page_obj, 'q': query})

@login_required
def crear_area(request):
    user = request.user
    perfil = getattr(user, 'perfilusuario', None)

    if request.method == 'POST':
        post = request.POST.copy()
        if not user.is_superuser and perfil and perfil.empresa:
            post['empresa'] = perfil.empresa.pk
        form = AreaComunForm(post, user=user)
        if form.is_valid():
            area = form.save(commit=False)
            if not user.is_superuser and perfil and perfil.empresa:
                area.empresa = perfil.empresa
            area.save()
            messages.success(request, "Área común creada correctamente.")
            return redirect('lista_areas')
        else:
            messages.error(request, "No se pudo crear el área. Revisa los datos ingresados.")
    else:
        form = AreaComunForm(user=user)
        if not user.is_superuser and perfil and perfil.empresa:
            form.fields['empresa'].initial = perfil.empresa

    return render(request, 'areas/crear_area.html', {'form': form})

@login_required
def editar_area(request, pk):
    user = request.user
    area = get_object_or_404(AreaComun, pk=pk)
    if not user.is_superuser and area.empresa != user.perfilusuario.empresa:
        return redirect('lista_areas')

    if request.method == 'POST':
        post = request.POST.copy()
        perfil = getattr(user, 'perfilusuario', None)
        if not user.is_superuser and perfil and perfil.empresa:
            post['empresa'] = perfil.empresa.pk
        form = AreaComunForm(post, instance=area, user=user)
        if form.is_valid():
            area_original = AreaComun.objects.get(pk=pk)
            area_modificada = form.save(commit=False)
            # Asegura la empresa para usuarios normales
            if not user.is_superuser and perfil and perfil.empresa:
                area_modificada.empresa = perfil.empresa

            for field in form.changed_data:
                valor_ant = getattr(area_original, field)
                valor_nuevo = getattr(area_modificada, field)
                AuditoriaCambio.objects.create(
                    modelo='area',
                    objeto_id=area.pk,
                    campo=field,
                    valor_anterior=valor_ant,
                    valor_nuevo=valor_nuevo,
                    usuario=request.user,
                )
            area_modificada.save()
            messages.success(request, "Área común actualizada correctamente.")
            return redirect('lista_areas')
    else:
        form = AreaComunForm(instance=area, user=user)

    return render(request, 'areas/editar_area.html', {'form': form, 'area': area})

@login_required
def eliminar_area(request, pk):
    user = request.user
    area = get_object_or_404(AreaComun, pk=pk)
    if not user.is_superuser and area.empresa != user.perfilusuario.empresa:
        return redirect('lista_areas')

    if request.method == 'POST':
        area.activo = False
        area.save()
        return redirect('lista_areas')

    return render(request, 'areas/eliminar_area.html', {'area': area})

#@user_passes_test(lambda u: u.is_staff)
@login_required
def areas_inactivas(request):
    if empresa := getattr(request.user, 'perfilusuario', None):
        empresa = request.user.perfilusuario.empresa
    else:
        empresa = None
    areas = AreaComun.objects.filter(empresa=empresa,activo=False)
    return render(request, 'areas/areas_inactivas.html', {'areas': areas})

#@user_passes_test(lambda u: u.is_staff)
@login_required
def reactivar_area(request, pk):
    area = get_object_or_404(AreaComun, pk=pk, activo=False)

    if request.method == 'POST':
        area.activo = True
        area.save()
        return redirect('areas_inactivas')

    return render(request, 'areas/reactivar_confirmacion.html', {'area': area})


@login_required
def incrementar_cuotas_areas(request):
    if request.method == 'POST':
        porcentaje = request.POST.get('porcentaje')
        try:
            porcentaje = Decimal(porcentaje)
            empresa = None
            if not request.user.is_superuser and hasattr(request.user, 'perfilusuario'):
                empresa = request.user.perfilusuario.empresa
                areas = AreaComun.objects.filter(empresa=empresa, activo=True)
            else:
                areas = AreaComun.objects.filter(activo=True)

            for area in areas:
                cuota_anterior = area.cuota
                incremento = cuota_anterior * (porcentaje / Decimal('100'))
                area.cuota += incremento
                area.save()

            messages.success(request, f'Se incrementaron las cuotas en un {porcentaje}% para todas las áreas comunes activas.')
            return redirect('incrementar_c_areas')
        except:
            messages.error(request, 'Porcentaje inválido.')

    return render(request, 'areas/incrementar_c_areas.html')

@login_required
def asignar_cliente_area(request, pk):
    area = get_object_or_404(AreaComun, pk=pk, status='disponible')
    if request.method == 'POST':
        form = AsignarClienteForm(request.POST, instance=area)
        if form.is_valid():
            area = form.save(commit=False)
            area.status = 'ocupado'
            area.save()
            messages.success(request, 'Cliente asignado correctamente.')
            return redirect('lista_areas')
    else:
        form = AsignarClienteForm(instance=area)
    return render(request, 'areas/asignar_cliente.html', {'form': form, 'area': area})

def buscar_por_id_o_nombre(modelo, valor, campo='nombre'):
    if not valor:
        return None
    val = str(valor).strip().replace(',', '')
    try:
        return modelo.objects.get(pk=int(val))
    except (ValueError, modelo.DoesNotExist):
        todos = modelo.objects.all()
        candidatos = [
            obj for obj in todos
            if unidecode(val).lower() in unidecode(str(getattr(obj, campo))).lower()
        ]
        if len(candidatos) == 1:
            return candidatos[0]
        elif len(candidatos) > 1:
            conflicto = "; ".join([f"ID={obj.pk}, {campo}='{getattr(obj, campo)}'" for obj in candidatos])
            raise Exception(f"Conflicto: '{valor}' coincide con varios registros en {modelo.__name__}: {conflicto}")
        raise Exception(f"No se encontró '{valor}' en {modelo.__name__}")

@login_required
def carga_masiva_areas(request):
    if request.method == 'POST':
        form = AreaComunCargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            errores = []
            exitos = 0
            COLUMNAS_ESPERADAS = 16  # empresa, nombre_cliente, rfc_cliente, email_cliente, numero, cuota, deposito, ubicacion, superficie_m2, tipo_area, cantidad_areas, giro, status, fecha_inicial, fecha_fin, observaciones
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None:
                    continue
                if len(row) != COLUMNAS_ESPERADAS:
                    errores.append(f"Fila {i}: número de columnas incorrecto ({len(row)} en vez de {COLUMNAS_ESPERADAS})")
                    continue
                empresa_val, nombre_cliente, rfc_cliente, email_cliente, numero, cuota, deposito, ubicacion, superficie_m2, tipo_area, cantidad_areas, giro, status, fecha_inicial, fecha_fin, observaciones = row
                try:
                    empresa = buscar_por_id_o_nombre(Empresa, empresa_val)
                    if not empresa:
                        errores.append(f"Fila {i}: No se encontró la empresa '{empresa_val}'")
                        continue
                    if not numero:
                        raise Exception("Número vacío")
                    # Validar que el número de área no se repita para la empresa
                    if AreaComun.objects.filter(empresa=empresa, numero=str(numero)).exists():
                        errores.append(f"Fila {i}: El número de área '{numero}' ya existe para la empresa '{empresa}'.")
                        continue
                    # Crear cliente solo si el RFC no existe
                    cliente = None
                    if rfc_cliente:
                        cliente, creado = Cliente.objects.get_or_create(
                            rfc=rfc_cliente,
                            defaults={
                                'nombre': nombre_cliente,
                                'empresa': empresa,
                                'email': email_cliente
                            }
                        )
                    AreaComun.objects.create(
                        empresa=empresa,
                        cliente=cliente,
                        numero=str(numero),
                        cuota=Decimal(cuota),
                        deposito=Decimal(deposito) if deposito else None,
                        ubicacion=ubicacion or "",
                        superficie_m2=Decimal(superficie_m2) if superficie_m2 else None,
                        tipo_area=tipo_area or "",
                        cantidad_areas=int(cantidad_areas) if cantidad_areas else 1,
                        giro=giro or "",
                        status=status or "ocupado",
                        fecha_inicial=fecha_inicial,
                        fecha_fin=fecha_fin,
                        observaciones=observaciones or ""
                    )
                    exitos += 1
                except Exception as e:
                    import traceback
                    errores.append(f"Fila {i}: {str(e) or repr(e)}<br>{traceback.format_exc()}")

            if exitos:
                messages.success(request, f"¡{exitos} áreas cargadas exitosamente!")
            if errores:
                messages.error(request, "Algunas áreas no se cargaron:<br>" + "<br>".join(errores))
            return redirect('carga_masiva_areas')
    else:
        form = AreaComunCargaMasivaForm()
    return render(request, 'areas/carga_masiva_areas.html', {'form': form}) 

@login_required
def plantilla_areas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Áreas"
    ws.append([
        'condominio', 'cliente','rfc_cliente', 'email_cliente','numero', 'cuota','deposito', 'ubicacion', 'superficie_m2','tipo_area','cantidad_areas','giro' ,
        'status', 'fecha_inicial', 'fecha_fin', 'observaciones'
    ])
    ws.append([
        'condominio Torre Reforma AC', 'Juan Pérez','XXX-XXX-XXX','email@ejemplo.com', 'A101', '1500.00','1500', 'Roof Garden', '200.0','Modulo','1','Restaurante',
        'ocupado', '2024-07-01', '2024-12-31', 'Área exclusiva'
    ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_areas.xlsx'
    wb.save(response)
    return response  