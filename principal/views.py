from turtle import st
from uuid import uuid4
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
import openpyxl
from core import settings
from empleados.models import Empleado
from empresas.models import Empresa
from clientes.models import Cliente
from gastos.models import Gasto
from locales.models import LocalComercial
from areas.models import AreaComun
from facturacion.models import CobroOtrosIngresos, Factura, FacturaOtrosIngresos, Pago
from presupuestos.models import Presupuesto, PresupuestoIngreso
from principal.models import AuditoriaCambio
from proveedores.models import Proveedor
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.core.mail import send_mail
from django.contrib.auth.decorators import login_required
from .models import Evento, PerfilUsuario
import json
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.contrib.auth.models import User
from datetime import date
import stripe

# stripe.api_key = settings.STRIPE_SECRET_KEY
stripe.api_key = settings.STRIPE_SECRET_KEY

# Create your views here.


@login_required
def bienvenida(request):
    empresa = None
    es_demo = False
    perfil = request.user.perfilusuario
    mostrar_wizard = perfil.mostrar_wizard
    
    mensaje_pago = None
    if request.GET.get("pago") == "ok":
        mensaje_pago = "¡Tu suscripción se ha activado correctamente! Puedes empezar a usar el sistema."

    if not request.user.is_superuser:
        empresa = request.user.perfilusuario.empresa
        eventos = Evento.objects.filter(empresa=empresa).order_by("fecha")
        es_demo = request.user.perfilusuario.tipo_usuario == "demo"
    else:
        eventos = Evento.objects.all().order_by("fecha")
    return render(
        request,
        "bienvenida.html",
        {
            "empresa": empresa,
            "eventos": eventos,
            "es_demo": es_demo,
            "STRIPE_PUBLIC_KEY": settings.STRIPE_PUBLIC_KEY,
            "mostrar_wizard": mostrar_wizard,
            "mensaje_pago": mensaje_pago,
        },
    )


@staff_member_required
@login_required
def reiniciar_sistema(request):
    if request.method == "POST":
        try:
            with transaction.atomic():
                # Orden: pagos > facturas > locales/areas > clientes > empresas etc...

                Factura.objects.all().delete()
                LocalComercial.objects.all().delete()
                AreaComun.objects.all().delete()
                Cliente.objects.all().delete()
                # Empresa.objects.all().delete()
                Proveedor.objects.all().delete()
                Empleado.objects.all().delete()
                Gasto.objects.all().delete()
                Presupuesto.objects.all().delete()
                AuditoriaCambio.objects.all().delete()
                Evento.objects.all().delete()
                Pago.objects.all().delete()

            messages.success(request, "¡El sistema fue reiniciado exitosamente!")
        except Exception as e:
            messages.error(request, f"Error al reiniciar: {e}")
        return redirect("bienvenida")
    return render(request, "reiniciar_sistema.html")


@staff_member_required
def respaldo_empresa_excel(request):
    # Si no hay empresa seleccionada, muestra el formulario
    if request.method == "GET" and "empresa_id" not in request.GET:
        empresas = Empresa.objects.all()
        return render(request, "respaldo_empresas.html", {"empresas": empresas})

    empresa_id = request.GET.get("empresa_id")
    try:
        empresa = Empresa.objects.get(pk=empresa_id)
    except Empresa.DoesNotExist:
        return render(request, "respaldo_empresas.html", {
            "empresas": Empresa.objects.all(),
            "error": "Empresa no encontrada."
        })

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Borra hoja por defecto

    # CLIENTES
    ws = wb.create_sheet("Clientes")
    ws.append(["id", "nombre", "rfc", "telefono", "email", "activo"])
    for c in Cliente.objects.filter(empresa=empresa):
        ws.append([c.id, c.nombre, c.rfc, c.telefono, c.email, c.activo])

    # LOCALES
    ws = wb.create_sheet("Locales")
    ws.append([
        "id", "numero", "cliente", "ubicacion", "superficie_m2", "cuota",
        "status", "activo", "observaciones"
    ])
    for l in LocalComercial.objects.filter(empresa=empresa):
        ws.append([
            l.id,
            l.numero,
            l.cliente.nombre if l.cliente else "",
            l.ubicacion,
            l.superficie_m2,
            l.cuota,
            l.status,
            l.activo,
            l.observaciones,
        ])

    # ÁREAS COMUNES
    ws = wb.create_sheet("Áreas Comunes")
    ws.append([
        "cliente", "numero", "cuota", "ubicacion", "superficie_m2", "status",
        "fecha_inicial", "fecha_fin", "activo", "observaciones"
    ])
    for a in AreaComun.objects.filter(empresa=empresa):
        ws.append([
            a.cliente.nombre if a.cliente else "",
            a.numero,
            a.cuota,
            a.ubicacion,
            a.superficie_m2,
            a.status,
            str(a.fecha_inicial) if a.fecha_inicial else "",
            str(a.fecha_fin) if a.fecha_fin else "",
            a.activo,
            a.observaciones,
        ])

    # FACTURAS
    ws = wb.create_sheet("Facturas")
    ws.append([
        "folio", "cliente", "local", "area_comun", "monto",
        "fecha_emision", "fecha_vencimiento", "estatus"
    ])
    for f in Factura.objects.filter(empresa=empresa):
        ws.append([
            f.folio,
            f.cliente.nombre if f.cliente else "",
            f.local.numero if f.local else "",
            f.area_comun.numero if f.area_comun else "",
            float(f.monto),
            str(f.fecha_emision),
            str(f.fecha_vencimiento),
            f.estatus,
        ])

    # PAGOS
    ws = wb.create_sheet("Pagos")
    ws.append(["id", "factura", "fecha_pago", "monto", "registrado_por"])
    for p in Pago.objects.filter(factura__empresa=empresa):
        ws.append([
            p.id,
            p.factura.folio if p.factura else "",
            str(p.fecha_pago),
            float(p.monto),
            p.registrado_por.get_full_name() if p.registrado_por else "",
        ])

    # GASTOS
    ws = wb.create_sheet("Gastos")
    ws.append([
        "id", "proveedor", "empleado", "descripcion", "monto", "tipo_gasto", "fecha"
    ])
    for g in Gasto.objects.filter(empresa=empresa):
        ws.append([
            g.id,
            str(g.proveedor) if g.proveedor else "",
            str(g.empleado) if g.empleado else "",
            g.descripcion,
            float(g.monto),
            str(g.tipo_gasto) if g.tipo_gasto else "",
            str(g.fecha),
        ])

    # PAGOS GASTOS
    ws = wb.create_sheet("Pagos Gastos")
    ws.append(["id", "referencia", "fecha_pago", "monto", "registrado_por"])
    for g in Gasto.objects.filter(empresa=empresa):
        for pago in g.pagos.all():
            ws.append([
                pago.id,
                pago.referencia,
                str(pago.fecha_pago),
                float(pago.monto),
                pago.registrado_por.get_full_name() if pago.registrado_por else "",
            ])

    # PRESUPUESTOS
    ws = wb.create_sheet("Presupuestos")
    ws.append([
        "id", "empresa", "grupo", "subgrupo", "tipo_gasto", "anio", "mes", "monto"
    ])
    for p in Presupuesto.objects.filter(empresa=empresa):
        ws.append([
            p.id,
            p.empresa.nombre if p.empresa else "",
            str(p.grupo) if p.grupo else "",
            str(p.subgrupo) if p.subgrupo else "",
            str(p.tipo_gasto) if p.tipo_gasto else "",
            p.anio,
            p.mes,
            float(p.monto),
        ])
        # PRESUPUESTOS INGRESOS
    ws = wb.create_sheet("Presupuestos Ingresos")
    ws.append([
        "id", "empresa", "tipo_ingreso", "anio", "mes", "monto"
    ])
    for p in PresupuestoIngreso.objects.filter(empresa=empresa):
        ws.append([
            p.id,
            p.empresa.nombre if p.empresa else "",
            str(p.tipo_ingreso) if hasattr(p, "tipo_ingreso") else "",
            p.anio,
            p.mes,
            float(p.monto),
        ])

    # EMPLEADOS
    ws = wb.create_sheet("Empleados")
    ws.append(["id", "nombre", "email", "telefono", "puesto", "activo"])
    for e in Empleado.objects.filter(empresa=empresa):
        ws.append([e.id, e.nombre, e.email, e.telefono, e.puesto, e.activo])

    # PROVEEDORES
    ws = wb.create_sheet("Proveedores")
    ws.append(["id", "nombre", "rfc", "telefono", "email", "activo"])
    for p in Proveedor.objects.filter(empresa=empresa):
        ws.append([
            p.id,
            str(p.nombre),
            p.rfc,
            p.telefono,
            p.email,
            p.activo
        ])

    # OTROS INGRESOS
    ws = wb.create_sheet("Otros Ingresos")
    ws.append([
        "id", "folio", "cliente", "tipo_ingreso", "monto", "saldo", "fecha_emision",
        "fecha_vencimiento", "estatus", "observaciones"
    ])
    for f in FacturaOtrosIngresos.objects.filter(empresa=empresa):
        ws.append([
            f.id,
            f.folio,
            f.cliente.nombre if f.cliente else "",
            str(f.tipo_ingreso) if hasattr(f, "tipo_ingreso") else "",
            float(f.monto),
            float(f.saldo),
            str(f.fecha_emision),
            str(f.fecha_vencimiento) if f.fecha_vencimiento else "",
            f.estatus,
            f.observaciones or "",
        ])

    # PAGOS OTROS INGRESOS
    ws = wb.create_sheet("Pagos Otros Ingresos")
    ws.append(["id", "factura", "fecha_pago", "monto", "registrado_por"])
    for p in CobroOtrosIngresos.objects.filter(factura__empresa=empresa):
        ws.append([
            p.id,
            p.factura.folio if p.factura else "",
            str(p.fecha_pago),
            float(p.monto),
            p.registrado_por.get_full_name() if p.registrado_por else "",
        ])    

    # Responde el archivo Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f"attachment; filename=respaldo_empresa_{empresa.nombre}.xlsx"
    )
    wb.save(response)
    return response


@staff_member_required  # Solo para admin/superusuario, opcional
def reporte_auditoria(request):
    modelo = request.GET.get("modelo")
    queryset = AuditoriaCambio.objects.all().order_by("-fecha_cambio")
    if modelo in ["local", "area", "factura"]:
        queryset = queryset.filter(modelo=modelo)
    return render(
        request, "auditoria/reporte.html", {"auditorias": queryset, "modelo": modelo}
    )


@csrf_exempt
@login_required
def crear_evento(request):
    if request.method == "POST":
        empresa = request.user.perfilusuario.empresa
        data = json.loads(request.body)
        evento = Evento.objects.create(
            empresa=empresa,
            titulo=data.get("titulo"),
            fecha=data.get("fecha"),
            descripcion=data.get("descripcion"),
            creado_por=request.user,
        )

        evento.save()
        return JsonResponse({"ok": True, "id": evento.id})
    return JsonResponse({"ok": False}, status=400)


@csrf_exempt
@login_required
def eliminar_evento(request, evento_id):
    if request.method == "POST":
        try:
            evento = Evento.objects.get(
                id=evento_id, empresa=request.user.perfilusuario.empresa
            )
            evento.delete()
            return JsonResponse({"ok": True})
        except Evento.DoesNotExist:
            return JsonResponse({"ok": False, "error": "No encontrado"}, status=404)
    return JsonResponse({"ok": False}, status=400)


@csrf_exempt
@login_required
def enviar_correo_evento(request, evento_id):
    if request.method == "POST":
        correo_destino = request.POST.get("correo")
        archivos = request.FILES.getlist("archivos")
        try:
            evento = Evento.objects.get(
                id=evento_id, empresa=request.user.perfilusuario.empresa
            )
            if correo_destino:
                cuerpo_html = render_to_string(
                    "correo_evento.html", {"evento": evento, "empresa": evento.empresa}
                )
                email = EmailMessage(
                    subject=f"Nuevo evento: {evento.titulo}",
                    body=cuerpo_html,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[correo_destino],
                )
                email.content_subtype = "html"
                for archivo in archivos:
                    email.attach(archivo.name, archivo.read(), archivo.content_type)
                email.send(fail_silently=False)
                evento.enviado_correo = True
                evento.save()
                return JsonResponse({"ok": True})
            else:
                return JsonResponse(
                    {"ok": False, "error": "Correo no proporcionado"}, status=400
                )
        except Evento.DoesNotExist:
            return JsonResponse(
                {"ok": False, "error": "Evento no encontrado"}, status=404
            )
    return JsonResponse({"ok": False}, status=400)


def registro_usuario(request):
    mensaje = ""
    if request.method == "POST":
        nombre = request.POST["nombre"]
        username = request.POST["username"]
        password = request.POST["password"]
        email = request.POST["email"]
        # telefono = request.POST['telefono']

        if User.objects.filter(username=username).exists():
            mensaje = "El nombre de usuario ya está en uso. Por favor elige otro."
        else:
            user = User.objects.create_user(
                username=username, password=password, email=email, first_name=nombre
            )
            empresa = Empresa.objects.create(
                nombre="EMPRESA DEMO AC", rfc=f"DEMO{uuid4().hex[:8].upper()}"
            )
            # PerfilUsuario.objects.create(
            #     usuario=user, empresa=empresa, tipo_usuario="demo"
            # )
            # Asigna la empresa y tipo_usuario al perfil creado por la señal
            perfil = user.perfilusuario
            perfil.empresa = empresa
            if not user.is_superuser:
                perfil.tipo_usuario = "demo"
            perfil.save()
            return redirect("login")
    return render(request, "registro.html", {"mensaje": mensaje})


@staff_member_required
@login_required
def usuarios_demo(request):
    usuarios = User.objects.filter(perfilusuario__tipo_usuario="demo")
    usuarios_info = []
    for user in usuarios:
        dias = (date.today() - user.date_joined.date()).days
        usuarios_info.append(
            {
                "id": user.id,
                "username": user.username,
                "first_name": user.first_name,
                "email": user.email,
                "is_active": user.is_active,
                "dias_demo": dias,
            }
        )
    if request.method == "POST":
        accion = request.POST.get("accion")
        if accion == "inactivar":
            ids = request.POST.getlist("inactivar")
            User.objects.filter(id__in=ids).update(is_active=False)
        elif accion == "reactivar":
            ids = request.POST.getlist("reactivar")
            User.objects.filter(id__in=ids).update(is_active=True)
        elif accion == "reactivar_todos":
            User.objects.filter(perfilusuario__tipo_usuario="demo").update(
                is_active=True
            )
        return redirect("usuarios_demo")
    return render(request, "usuarios_demo.html", {"usuarios": usuarios_info})



@csrf_exempt
def stripe_webhook(request):
    import stripe
    from django.http import HttpResponse
    from django.conf import settings
    from .models import PerfilUsuario
    from django.contrib.auth.models import User

    payload = request.body
    sig_header = request.META.get('HTTP_STRIPE_SIGNATURE')
    endpoint_secret = settings.STRIPE_ENDPOINT_SECRET

    try:
        event = stripe.Webhook.construct_event(
            payload, sig_header, endpoint_secret
        )
    except Exception as e:
        print("Error Stripe:", e)
        return HttpResponse(status=400)

    # Activación inicial de suscripción
    if event['type'] == 'checkout.session.completed':
        session = event['data']['object']
        customer_id = session.get('customer')
        subscription_id = session.get('subscription')
        email = session.get('customer_details', {}).get('email')
        if customer_id:
            try:
                perfil = PerfilUsuario.objects.get(stripe_customer_id=customer_id)
                user = perfil.usuario
                user.is_active = True
                perfil.tipo_usuario = 'pago'
                # Solo mostrar el wizard si nunca ha tenido suscripción
                if not perfil.stripe_subscription_id:
                    perfil.mostrar_wizard = True
                if subscription_id:
                    perfil.stripe_subscription_id = subscription_id
                perfil.save()
                user.save()
            except PerfilUsuario.DoesNotExist:
                if email:
                    try:
                        user = User.objects.get(email=email)
                        perfil = PerfilUsuario.objects.get(usuario=user)
                        perfil.stripe_customer_id = customer_id
                        perfil.tipo_usuario = 'pago'
                        # Solo mostrar el wizard si nunca ha tenido suscripción
                        if not perfil.stripe_subscription_id:
                            perfil.mostrar_wizard = True
                        if subscription_id:
                            perfil.stripe_subscription_id = subscription_id
                        perfil.save()
                        user.is_active = True
                        user.save()
                    except (User.DoesNotExist, PerfilUsuario.DoesNotExist):
                        print("No se pudo encontrar el usuario asociado al pago.")

    # Renovación automática de suscripción
    if event['type'] == 'invoice.payment_succeeded':
        invoice = event['data']['object']
        customer_id = invoice.get('customer')
        if customer_id:
            try:
                perfil = PerfilUsuario.objects.get(stripe_customer_id=customer_id)
                perfil.tipo_usuario = 'pago'
                perfil.save()
                print("Renovación automática: acceso renovado.")
            except PerfilUsuario.DoesNotExist:
                print("No se encontró perfil para renovar acceso.")

    return HttpResponse(status=200)


@login_required
def crear_sesion_pago(request):
    session = stripe.checkout.Session.create(
        payment_method_types=["card"],
        line_items=[
            {
                "price": "price_1RqexnPYnlfwKZQHILP9tgW5",
                "quantity": 1,
            }
        ],
        mode="subscription",
        success_url=request.build_absolute_uri("/bienvenida/?pago=ok"),
        cancel_url=request.build_absolute_uri("/bienvenida/"),
        client_reference_id=str(request.user.id),  # Para identificar al usuario
        customer_email=request.user.email,
    )
    return JsonResponse({"id": session.id})


@login_required
def cancelar_suscripcion(request):
    perfil = request.user.perfilusuario
    subscription_id = perfil.stripe_subscription_id
    if subscription_id:
        stripe.api_key = settings.STRIPE_SECRET_KEY
        try:
            stripe.Subscription.delete(subscription_id)
            perfil.tipo_usuario = 'demo'  # O el estado que corresponda
            perfil.save()
            return JsonResponse({'status': 'cancelada'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'detail': str(e)}, status=400)
    return JsonResponse({'status': 'no encontrada'}, status=404)

@require_POST
@login_required
def guardar_datos_empresa(request):
    perfil = request.user.perfilusuario
    empresa = perfil.empresa
    nuevo_rfc = request.POST.get('rfc_empresa', '').strip()

    if Empresa.objects.filter(rfc=nuevo_rfc).exclude(id=empresa.id).exists():
        messages.error(request, "El RFC ingresado ya está registrado en otra empresa.")
        return redirect('bienvenida')
    
    empresa.nombre = request.POST.get('nombre_empresa', '')
    empresa.rfc = nuevo_rfc
    empresa.direccion = request.POST.get('direccion_empresa', '')
    empresa.email = request.POST.get('email_empresa', '')
    empresa.telefono = request.POST.get('telefono_empresa', '')
    empresa.cuenta_bancaria = request.POST.get('cuenta_bancaria', '')
    empresa.numero_cuenta = request.POST.get('numero_cuenta', '')
    try:
        empresa.saldo_inicial = float(request.POST.get('saldo_inicial', 0.00))
    except ValueError:
        empresa.saldo_inicial = 0.00
    empresa.save()
    perfil.mostrar_wizard = False
    perfil.save()
    messages.success(request, "¡Datos de empresa actualizados correctamente!")
    return redirect('bienvenida')

# @login_required
# def prueba_error(request):
#     # Esto lanzará un error que Sentry debe capturar
#     raise Exception("¡Este es un error de prueba para Sentry!")
#     return HttpResponse("No deberías ver esto.")