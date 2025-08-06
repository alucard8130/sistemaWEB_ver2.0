
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
import openpyxl
#import unidecode
from clientes.models import Cliente
from empresas.models import Empresa
import locales
from principal.models import AuditoriaCambio
from .models import LocalComercial
from .forms import LocalCargaMasivaForm, LocalComercialForm
from django.contrib.admin.views.decorators import staff_member_required
from unidecode import unidecode
from django.core.paginator import Paginator
from django.db.models import Q


# Create your views here.

@login_required
def lista_locales(request):
    user = request.user
    query = request.GET.get("q", "")
    if user.is_superuser:
        #locales = LocalComercial.objects.all()
        locales = LocalComercial.objects.filter(activo=True).order_by('numero')
    else:
        #empresa = getattr(request.user.perfilusuario, 'empresa', None)
        empresa = user.perfilusuario.empresa
        locales = LocalComercial.objects.filter(empresa=empresa, activo=True).order_by('numero')

    if query:
        locales = locales.filter(
            Q(numero__icontains=query) | Q(cliente__nombre__icontains=query) | Q(cliente__rfc__icontains=query)
        )

    locales = locales.order_by('numero')

    # Paginación
    paginator = Paginator(locales, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'locales/lista_locales.html', {'locales': locales, 'locales': page_obj, 'q': query})


@login_required
def crear_local(request):
    user = request.user
    perfil = getattr(user, 'perfilusuario', None)
    
    if request.method == 'POST':
        form = LocalComercialForm(request.POST, user=user)
        if form.is_valid():
            local = form.save(commit=False)
            # Si no es superusuario, asignamos su empresa
            if not user.is_superuser and perfil and perfil.empresa:
                local.empresa = perfil.empresa
            local.save()
            messages.success(request, "Local creado correctamente.")
            return redirect('lista_locales')
        else:
            messages.error(request, "No se pudo crear el local. Revisa los datos ingresados.")
    else:
        form = LocalComercialForm(user=user)
        # Si no es superusuario, asignamos la empresa inicial al form
        if not user.is_superuser and perfil and perfil.empresa:
            form.fields['empresa'].initial = perfil.empresa

    return render(request, 'locales/crear_local.html', {'form': form})

@login_required
def editar_local(request, pk):
    user = request.user
    local= get_object_or_404(LocalComercial, pk=pk)
 
    if not user.is_superuser and local.empresa != user.perfilusuario.empresa:
        return redirect('lista_locales')

    if request.method == 'POST':
        form = LocalComercialForm(request.POST, instance=local, user=user)
        if form.is_valid():
            local_original = LocalComercial.objects.get(pk=pk)
            local_modificado = form.save(commit=False)
            for field in form.changed_data:
                valor_anterior = getattr(local_original, field)
                valor_nuevo = getattr(local_modificado, field)
                AuditoriaCambio.objects.create(
                    modelo='local',
                    objeto_id=local.pk,
                    campo=field,
                    valor_anterior=valor_anterior,
                    valor_nuevo=valor_nuevo,
                    usuario=request.user,
                )
            form.save()
            messages.success(request, "Local actualizado correctamente.")
            return redirect('lista_locales')
    else:
        form = LocalComercialForm(instance=local, user=user)

    return render(request, 'locales/editar_local.html', {'form': form, 'local': local})

@login_required
def eliminar_local(request, pk):
    user = request.user
    local= get_object_or_404(LocalComercial, pk=pk)
    if not user.is_superuser and local.empresa != user.perfilusuario.empresa:
        return redirect('lista_locales')

    if request.method == 'POST':
        local.activo = False
        local.save()
        return redirect('lista_locales')

    return render(request, 'locales/eliminar_local.html', {'local': local})

#@staff_member_required
#@user_passes_test(lambda u: u.is_staff)
def locales_inactivos(request):
    if empresa := getattr(request.user, 'perfilusuario', None):
        empresa = request.user.perfilusuario.empresa
    else:
        empresa = None
    locales = LocalComercial.objects.filter(empresa=empresa,activo=False)
    return render(request, 'locales/locales_inactivos.html', {'locales': locales})

#@staff_member_required
#@user_passes_test(lambda u: u.is_staff)
def reactivar_local(request, pk):
    local = get_object_or_404(LocalComercial, pk=pk, activo=False)

    if request.method == 'POST':
        local.activo = True
        local.save()
        return redirect('locales_inactivos')

    return render(request, 'locales/reactivar_confirmacion.html', {'local': local})

@login_required
def incrementar_cuotas_locales(request):
    if request.method == 'POST':
        porcentaje = request.POST.get('porcentaje')
        try:
            porcentaje = Decimal(porcentaje)
            empresa = None
            if not request.user.is_superuser and hasattr(request.user, 'perfilusuario'):
                empresa = request.user.perfilusuario.empresa
                locales = LocalComercial.objects.filter(empresa=empresa, activo=True)
            else:
                locales = LocalComercial.objects.filter(activo=True)

            for local in locales:
                cuota_anterior = local.cuota
                incremento = cuota_anterior * (porcentaje / Decimal('100'))
                local.cuota += incremento
                local.save()

            messages.success(request, f'Se incrementaron las cuotas en un {porcentaje}% para todos los locales activos.')
            return redirect('incrementar_c_locales')
        except:
            messages.error(request, 'Porcentaje inválido.')
    
    return render(request, 'locales/incrementar_c_locales.html')

def buscar_por_id_o_nombre(modelo, valor, campo='nombre'):
    if not valor:
        return None
    val = str(valor).strip().replace(',', '')
    try:
        return modelo.objects.get(pk=int(val))
    except (ValueError, modelo.DoesNotExist):
        todos = modelo.objects.all()
        candidatos = [
            obj for obj in todos
            if unidecode(val).lower() in unidecode(str(getattr(obj, campo))).lower()
        ]
        if len(candidatos) == 1:
            return candidatos[0]
        elif len(candidatos) > 1:
            conflicto = "; ".join([f"ID={obj.pk}, {campo}='{getattr(obj, campo)}'" for obj in candidatos])
            raise Exception(f"Conflicto: '{valor}' coincide con varios registros en {modelo.__name__}: {conflicto}")
        raise Exception(f"No se encontró '{valor}' en {modelo.__name__}")

@login_required
def carga_masiva_locales(request):
    if request.method == 'POST':
        form = LocalCargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            errores = []
            exitos = 0
            COLUMNAS_ESPERADAS = 12  # Ajusta según tus columnas
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None:
                    continue
                if len(row) != COLUMNAS_ESPERADAS:
                    errores.append(f"Fila {i}: número de columnas incorrecto ({len(row)} en vez de {COLUMNAS_ESPERADAS})")
                    continue
                empresa_val, propietario_val, nombre_cliente, rfc_cliente, email_cliente, numero, cuota, ubicacion, superficie_m2, giro, status, observaciones = row
                try:
                    empresa = buscar_por_id_o_nombre(Empresa, empresa_val)
                    if not empresa:
                        errores.append(f"Fila {i}: No se encontró la empresa '{empresa_val}'")
                        continue
                    if not numero:
                        raise Exception("Número vacío")
                    # Validar que el número de local no se repita para la empresa
                    if LocalComercial.objects.filter(empresa=empresa, numero=str(numero)).exists():
                        errores.append(f"Fila {i}: El número de local '{numero}' ya existe para la empresa '{empresa}'.")
                        continue
                    # Validar y convertir cuota
                    try:
                        cuota_decimal = Decimal(cuota)
                    except (InvalidOperation, TypeError, ValueError):
                        errores.append(f"Fila {i}: El valor de cuota '{cuota}' no es válido.")
                        continue
                    # Crear cliente solo si el RFC no existe
                    cliente = None
                    if rfc_cliente:
                        cliente, creado = Cliente.objects.get_or_create(
                            rfc=rfc_cliente,
                            defaults={
                                'nombre': nombre_cliente,
                                'empresa': empresa,
                                'email': email_cliente
                            }
                        )
                    LocalComercial.objects.create(
                        empresa=empresa,
                        propietario=propietario_val or "",
                        cliente=cliente,
                        numero=str(numero),
                        cuota= cuota_decimal,
                        ubicacion=ubicacion or "",
                        superficie_m2=Decimal(superficie_m2) if superficie_m2 else None,
                        giro=giro or "",
                        status=status or "ocupado",
                        observaciones=observaciones or ""
                    )
                    exitos += 1
                except Exception as e:
                    import traceback
                    errores.append(f"Fila {i}: {str(e) or repr(e)}<br>{traceback.format_exc()}")

            if exitos:
                messages.success(request, f"¡{exitos} locales cargados exitosamente!")
            if errores:
                messages.error(request, "Algunos locales no se cargaron:<br>" + "<br>".join(errores))
            return redirect('carga_masiva_locales')
    else:
        form = LocalCargaMasivaForm()
    return render(request, 'locales/carga_masiva_locales.html', {'form': form})

@login_required
def plantilla_locales_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Locales"
    ws.append([
        'condominio','propietario', 'cliente', 'rfc','email','numero',  'cuota','ubicacion', 'superficie_m2','giro', 'status', 'observaciones'
    ])
    ws.append([
        'plaza en condominio AC','Tiendas Soriana SA de CV','Juan Pérez','XXX-XXX-XXX','email@ejemplo.com', '101', '120.3', 'planta baja', '30.5','venta ropa', 'ocupado', 'carga inicial'
    ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_locales.xlsx'
    wb.save(response)
    return response
