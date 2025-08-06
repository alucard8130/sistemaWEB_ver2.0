
# Create your views here.
from django.forms import CharField
from django.shortcuts import render, redirect,get_object_or_404
from areas.models import AreaComun
from clientes.models import Cliente
from empresas.models import Empresa
from locales.models import LocalComercial
from .forms import FacturaEditForm, FacturaForm, FacturaOtrosIngresosForm, PagoForm,FacturaCargaMasivaForm, CobroForm, TipoOtroIngresoForm
from .models import CobroOtrosIngresos, Factura, FacturaOtrosIngresos, Pago, TipoOtroIngreso
from principal.models import AuditoriaCambio, PerfilUsuario
from django.utils.timezone import now
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from datetime import date, timedelta
from django.db.models import Q,  Value, Case, When,  CharField,FloatField
from django.db.models import F, OuterRef, Subquery, Sum, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce
#from django.db.models import Case, When
import openpyxl
from django.http import HttpResponse
from django.db.models import Q
from facturacion.models import Pago
from decimal import Decimal, InvalidOperation
from unidecode import unidecode
from django.db.models.functions import TruncMonth, TruncYear
from datetime import datetime
from django.db.models import Sum
from django.db import transaction
from django.contrib.auth import authenticate
from django.core.paginator import Paginator
import io
from django.utils.dateformat import DateFormat
from presupuestos.models import PresupuestoIngreso
from collections import defaultdict
import json
from django.http import JsonResponse

@login_required
def crear_factura(request):
    from django.db import transaction
   
    conflicto = False
    conflicto_tipo = ""
    superuser_auth_ok = False

    if request.method == 'POST':
        form = FacturaForm(request.POST, request.FILES, user=request.user)
        superuser_username = request.POST.get('superuser_username')
        superuser_password = request.POST.get('superuser_password')

        if form.is_valid():
            factura = form.save(commit=False)
            tipo = form.cleaned_data.get('tipo_origen')
            if tipo == 'local':
                factura.area_comun = None
            elif tipo == 'area_comun':
                factura.local = None
            cliente = factura.cliente
            
            # ---- Validación local ----
            if factura.local:
                local_cliente = factura.local.cliente
                if local_cliente and local_cliente != cliente:
                    conflicto = True
                    conflicto_tipo = "local"

            # ---- Validación área ----
            if factura.area_comun:
                area_cliente = factura.area_comun.cliente
                if area_cliente and area_cliente != cliente:
                    conflicto = True
                    conflicto_tipo = "area"

            # ---- Lógica de autorización ----
            if conflicto:
                # Si ya eres superusuario, pasa
                if request.user.is_superuser:
                    superuser_auth_ok = True
                    print("[DEBUG] Usuario actual es superuser, pasa conflicto.")
                # Si se dio password de superusuario
                elif superuser_username and superuser_password:
                    from django.contrib.auth import authenticate
                    superuser = authenticate(username=superuser_username, password=superuser_password)
                    if superuser and superuser.is_superuser:
                        superuser_auth_ok = True
                        print("[DEBUG] Autenticación por superuser exitosa.")
                    else:
                        form.add_error(None, "La autenticación de superusuario es incorrecta. No se creó la factura.")
                        print("[DEBUG] Autenticación superuser fallida.")
                else:
                    form.add_error(None, f"El cliente del {conflicto_tipo} seleccionado no coincide. Contacta al superusuario para autorizar el cambio.")
                    print("[DEBUG] Conflicto detectado sin autorización.")
            else:
                superuser_auth_ok = True  # Si no hay conflicto, siempre debe pasar

            if superuser_auth_ok:
                try:
                    with transaction.atomic():
                        # Asignar empresa
                        if request.user.is_superuser:
                            factura.empresa = factura.cliente.empresa
                        else:
                            factura.empresa = request.user.perfilusuario.empresa
                        
                        factura.estatus = 'pendiente'
                        # Asignar fecha_emision si no viene del formulario
                        if not factura.fecha_emision:
                            #factura.fecha_emision = now()
                            factura.fecha_emision = timezone.now().date()
                        factura.save()

                        # Folio único
                        count = Factura.objects.filter(fecha_emision__year=now().year).count() + 1
                        if tipo == 'local':
                            factura.folio = f"CM-F{count:05d}"
                        elif tipo == 'area_comun':
                            factura.folio = f"AC-F{count:05d}"
                        factura.save()

                        # Asignar cliente a local/área si está vacío o si hay conflicto autorizado
                        if factura.local and (factura.local.cliente is None or request.user.is_superuser or superuser_auth_ok):
                            factura.local.cliente = cliente
                            factura.local.save()
                        if factura.area_comun and (factura.area_comun.cliente is None or request.user.is_superuser or superuser_auth_ok):
                            factura.area_comun.cliente = cliente
                            factura.area_comun.save()

                        # Auditoría
                        for field in form.fields:
                            if hasattr(factura, field):
                                valor = getattr(factura, field)
                                AuditoriaCambio.objects.create(
                                    modelo='factura',
                                    objeto_id=factura.pk,
                                    usuario=request.user,
                                    campo=field,
                                    valor_anterior='--CREADA--',
                                    valor_nuevo=valor
                                )

                        messages.success(request, "Registro Exitoso.")
                       
                        return redirect('lista_facturas')
                except Exception as e:
                    form.add_error(None, f"Error al guardar: {e}")
                    
        else:
            ''
    else:
        form = FacturaForm(user=request.user)

    return render(request, 'facturacion/crear_factura.html', {
        'form': form,
        'pedir_superuser': conflicto and not superuser_auth_ok and request.method == 'POST'
    })

@login_required
def lista_facturas(request):
    empresa_id = request.GET.get('empresa')
    local_id = request.GET.get('local_id')
    area_id = request.GET.get('area_id')

    if request.user.is_superuser:
        facturas = Factura.objects.all().order_by('-fecha_vencimiento')
        empresas = Empresa.objects.all()
        locales = LocalComercial.objects.filter(activo=True)
        areas = AreaComun.objects.filter(activo=True)
    else:
        empresa = request.user.perfilusuario.empresa
        facturas = Factura.objects.filter(empresa=empresa).order_by('-fecha_vencimiento')
        empresas = None
        locales = LocalComercial.objects.filter(empresa=empresa, activo=True)
        areas = AreaComun.objects.filter(empresa=empresa, activo=True)

    if empresa_id:
        facturas = facturas.filter(empresa_id=empresa_id).order_by('-fecha_vencimiento')
    if local_id:
        facturas = facturas.filter(local_id=local_id).order_by('-fecha_vencimiento')
    if area_id:
        facturas = facturas.filter(area_comun_id=area_id).order_by('-fecha_vencimiento')

    facturas = facturas.select_related('cliente', 'empresa', 'local', 'area_comun').prefetch_related('pagos')    
    # Paginación
    paginator = Paginator(facturas, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'facturacion/lista_facturas.html', {
        'facturas': facturas,
        'empresas': empresas,
        'empresa_seleccionada': int(empresa_id) if empresa_id else None,
        'locales': locales,
        'areas': areas,
        'local_id': local_id,
        'area_id': area_id,
        'facturas': page_obj,
    })

from django.db.models import Max

@login_required
def facturar_mes_actual(request, facturar_locales=True, facturar_areas=True):
    # Permitir seleccionar año y mes por GET o POST (solo superusuario puede facturar meses anteriores)
    if request.method == 'POST':
        año = int(request.POST.get('anio', datetime.now().year))
        mes = int(request.POST.get('mes', datetime.now().month))
    else:
        año = int(request.GET.get('anio', datetime.now().year))
        mes = int(request.GET.get('mes', datetime.now().month))
    hoy = date.today()

    # Solo superusuario puede facturar meses distintos al actual
    # if (año != hoy.year or mes != hoy.month) and not request.user.is_superuser:
    #     messages.error(request, "Solo el superusuario puede generar facturas de meses anteriores.")
    #     return redirect('confirmar_facturacion')

    facturas_creadas = 0
    facturas_a_crear = []

    if request.user.is_superuser:
        locales = LocalComercial.objects.filter(activo=True, cliente__isnull=False) if facturar_locales else []
        areas = AreaComun.objects.filter(activo=True, cliente__isnull=False) if facturar_areas else []
    else:
        empresa = request.user.perfilusuario.empresa
        locales = LocalComercial.objects.filter(empresa=empresa, activo=True, cliente__isnull=False) if facturar_locales else []
        areas = AreaComun.objects.filter(empresa=empresa, activo=True, cliente__isnull=False) if facturar_areas else []

    fecha_factura = date(año, mes, 1)

    # Pre-carga los folios existentes por empresa y tipo para evitar consultas repetidas
    folios_locales = set(
        Factura.objects.filter(
            empresa__in=[l.empresa for l in locales],
            folio__startswith="CM-F"
        ).values_list('empresa_id', 'folio')
    )
    folios_areas = set(
        Factura.objects.filter(
            empresa__in=[a.empresa for a in areas],
            folio__startswith="AC-F"
        ).values_list('empresa_id', 'folio')
    )
    folios_deposito = set(
        Factura.objects.filter(
            empresa__in=[a.empresa for a in areas],
            folio__startswith="DG-F"
        ).values_list('empresa_id', 'folio')
    )

    # Locales
    max_folio_local = {}
    for local in locales:
        existe = Factura.objects.filter(
            cliente=local.cliente,
            local=local,
            fecha_emision__year=año,
            fecha_emision__month=mes
        ).exists()
        if not existe:
            empresa_id = local.empresa_id
            if empresa_id not in max_folio_local:
                max_folio = Factura.objects.filter(
                    empresa_id=empresa_id,
                    folio__startswith="CM-F"
                ).aggregate(max_f=Max('folio'))['max_f']
                if max_folio:
                    try:
                        last_num = int(max_folio.replace("CM-F", ""))
                    except Exception:
                        last_num = 0
                else:
                    last_num = 0
                max_folio_local[empresa_id] = last_num
            max_folio_local[empresa_id] += 1
            folio = f"CM-F{max_folio_local[empresa_id]:05d}"
            facturas_a_crear.append(Factura(
                empresa=local.empresa,
                cliente=local.cliente,
                local=local,
                folio=folio,
                fecha_emision=fecha_factura,
                fecha_vencimiento=fecha_factura,
                monto=local.cuota,
                tipo_cuota='mantenimiento',
                estatus='pendiente',
                observaciones='emision mensual'
            ))
            facturas_creadas += 1

    # Áreas
    max_folio_area = {}
    max_folio_deposito = {}
    for area in areas:
        existe = Factura.objects.filter(
            cliente=area.cliente,
            area_comun=area,
            fecha_emision__year=año,
            fecha_emision__month=mes
        ).exists()
        if not existe:
            empresa_id = area.empresa_id
            if empresa_id not in max_folio_area:
                max_folio = Factura.objects.filter(
                    empresa_id=empresa_id,
                    folio__startswith="AC-F"
                ).aggregate(max_f=Max('folio'))['max_f']
                if max_folio:
                    try:
                        last_num = int(max_folio.replace("AC-F", ""))
                    except Exception:
                        last_num = 0
                else:
                    last_num = 0
                max_folio_area[empresa_id] = last_num
            max_folio_area[empresa_id] += 1
            folio = f"AC-F{max_folio_area[empresa_id]:05d}"
            facturas_a_crear.append(Factura(
                empresa=area.empresa,
                cliente=area.cliente,
                area_comun=area,
                folio=folio,
                fecha_emision=fecha_factura,
                fecha_vencimiento=fecha_factura,
                monto=area.cuota,
                tipo_cuota='renta',
                estatus='pendiente',
                observaciones='emision mensual'
            ))
            facturas_creadas += 1

        # Depósito en garantía por única vez
        if area.deposito and area.deposito > 0:
            existe_deposito = Factura.objects.filter(
                cliente=area.cliente,
                area_comun=area,
                tipo_cuota='deposito',
            ).exists()
            if not existe_deposito:
                empresa_id = area.empresa_id
                if empresa_id not in max_folio_deposito:
                    max_folio = Factura.objects.filter(
                        empresa_id=empresa_id,
                        folio__startswith="DG-F"
                    ).aggregate(max_f=Max('folio'))['max_f']
                    if max_folio:
                        try:
                            last_num = int(max_folio.replace("DG-F", ""))
                        except Exception:
                            last_num = 0
                    else:
                        last_num = 0
                    max_folio_deposito[empresa_id] = last_num
                max_folio_deposito[empresa_id] += 1
                folio_deposito = f"DG-F{max_folio_deposito[empresa_id]:05d}"
                facturas_a_crear.append(Factura(
                    empresa=area.empresa,
                    cliente=area.cliente,
                    area_comun=area,
                    folio=folio_deposito,
                    fecha_emision=fecha_factura,
                    fecha_vencimiento=fecha_factura,
                    monto=area.deposito,
                    tipo_cuota='deposito',
                    estatus='pendiente',
                    observaciones='Depósito en garantía'
                ))

    # Bulk create
    if facturas_a_crear:
        Factura.objects.bulk_create(facturas_a_crear, batch_size=50)

    messages.success(request, f"{facturas_creadas} facturas generadas para {fecha_factura.strftime('%B %Y')}")
    return redirect('lista_facturas')

@login_required
def confirmar_facturacion(request):
    hoy = now().date()
    año, mes = hoy.year, hoy.month

    empresa = None
    if not request.user.is_superuser:
        empresa = request.user.perfilusuario.empresa

    # FILTRAR locales y áreas activos con cliente
    if request.user.is_superuser:
        locales = LocalComercial.objects.filter(activo=True, cliente__isnull=False)
        areas = AreaComun.objects.filter(activo=True, cliente__isnull=False)
    else:
        locales = LocalComercial.objects.filter(empresa=empresa, activo=True, cliente__isnull=False)
        areas = AreaComun.objects.filter(empresa=empresa, activo=True, cliente__isnull=False)

    # CONTAR locales no facturados aún este mes
    total_locales = sum(
        not Factura.objects.filter(cliente=l.cliente, local=l,
                                   fecha_emision__year=año, fecha_emision__month=mes).exists()
        for l in locales
    )

    total_areas = sum(
        not Factura.objects.filter(cliente=a.cliente, area_comun=a,
                                   fecha_emision__year=año, fecha_emision__month=mes).exists()
        for a in areas
    )

    if request.method == 'POST':
        facturar_locales = 'locales' in request.POST
        facturar_areas = 'areas' in request.POST
        return facturar_mes_actual(request, facturar_locales, facturar_areas)

    return render(request, 'facturacion/confirmar_facturacion.html', {
        'total_locales': total_locales,
        'total_areas': total_areas,
        'año': año,
        'mes': mes
    })

@login_required
@transaction.atomic
def registrar_pago(request, factura_id):
    factura = get_object_or_404(Factura, pk=factura_id)

    if factura.estatus == 'pagada' or factura.saldo_pendiente <= 0:
        messages.warning(request, "La factura ya está completamente pagada. No se pueden registrar más pagos.")
        return redirect('lista_facturas')

    if request.method == 'POST':
        form = PagoForm(request.POST,request.FILES)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.factura = factura           # SIEMPRE antes de save()
            pago.registrado_por = request.user

            if pago.forma_pago == 'nota_credito':
                #if factura.total_pagado > 0:
                    #form.add_error('monto', "No se puede registrar una nota de crédito si la factura tiene cobros asignados.")
                #else:    
                pago.save()
                factura.estatus = 'cancelada'
                factura.monto = 0  # Saldo pendiente a 0
                factura.save()
                messages.success(request, "La factura ha sido cancelada por nota de crédito. el saldo pendiente es $0.00")
                return redirect('lista_facturas')

            if pago.monto > factura.saldo_pendiente:
                form.add_error('monto', f"El monto excede el saldo pendiente (${factura.saldo_pendiente:.2f}).")
            else:
                pago.save()
                pagos_validos = factura.pagos.exclude(forma_pago='nota_credito')
                total_pagado = sum([p.monto for p in pagos_validos])
                if total_pagado >= factura.monto:
                    factura.estatus = 'pagada'
                else:
                    factura.estatus = 'pendiente'
                factura.save()
                factura.actualizar_estatus()  # Actualiza el estatus de la factura
                messages.success(request, f"Cobro registrado. Saldo restante: ${factura.saldo_pendiente:.2f}")
                return redirect('lista_facturas')
    else:
        form = PagoForm()

    return render(request, 'facturacion/registrar_pago.html', {
        'form': form,
        'factura': factura,
        'saldo': factura.saldo_pendiente,
    })

@login_required
def facturas_detalle(request, pk):
    factura = get_object_or_404(Factura, pk=pk)
    cobros = factura.pagos.all().order_by('fecha_pago')

    return render(request, 'facturacion/facturas_detalle.html', {
        'factura': factura,
        'cobros': cobros,
    })


@login_required
#pagos_por_origen.html
#reporte cobros cuotas
def pagos_por_origen(request):
    empresa_id = request.GET.get('empresa')
    local_id = request.GET.get('local_id')
    area_id = request.GET.get('area_id')
    #tipo = request.GET.get('tipo')  # 'local' o 'area'
    #pagos = Pago.objects.select_related('factura', 'factura__cliente', 'factura__empresa')

    if request.user.is_superuser:
        pagos = Pago.objects.select_related('factura', 'factura__empresa', 'factura__local', 'factura__area_comun', 'factura__cliente').all().order_by('-fecha_pago')
        empresas = Empresa.objects.all()
        locales = LocalComercial.objects.filter(activo=True)
        areas = AreaComun.objects.filter(activo=True)
    else:
        empresa = request.user.perfilusuario.empresa
        pagos = Pago.objects.select_related('factura').filter(factura__empresa=empresa).order_by('-fecha_pago')
        empresas = None
        locales = LocalComercial.objects.filter(empresa=empresa, activo=True)
        areas = AreaComun.objects.filter(empresa=empresa, activo=True)

    if empresa_id:
        pagos = pagos.filter(factura__empresa_id=empresa_id).order_by('fecha_pago')
    if local_id:
        pagos = pagos.filter(factura__local_id=local_id).order_by('fecha_pago')
    if area_id:
        pagos = pagos.filter(factura__area_comun_id=area_id).order_by('fecha_pago')

   
    pagos_validos = pagos.exclude(forma_pago='nota_credito')
    total_pagos = pagos_validos.aggregate(total=Sum('monto'))['total'] or 0

    #paginacion
    paginator = Paginator(pagos, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)



    return render(request, 'facturacion/pagos_por_origen.html', {
        'pagos': pagos,
        #'tipo': tipo,
        'total_pagos': total_pagos,
        'empresas': empresas,
        'empresa_seleccionada': int(empresa_id) if empresa_id else None,
        'locales': locales,
        'areas': areas,
        'local_id': local_id,
        'area_id': area_id,
        'pagos': page_obj,
    })

@login_required
#saldos.html
#dashboard cartera vencida
def dashboard_saldos(request):
    hoy = timezone.now().date()
    cliente_id = request.GET.get('cliente')
    origen = request.GET.get('origen')
    if not origen:
        origen = 'todos'
        
    es_super = request.user.is_superuser

    # Empresa según usuario
    if es_super:
        empresas = Empresa.objects.all()
        if not empresas.exists():
            #messages.error(request, "No hay empresas registradas en el sistema.")
            return render(request, 'dashboard/saldos.html', {'empresas': [], 'facturas': []})
        empresa_id = request.GET.get('empresa')
        if not empresa_id or empresa_id == "todas":
            filtro_empresa = Q()
        else:
            filtro_empresa = Q(empresa_id=empresa_id)
    else:
        try:
            empresa = request.user.perfilusuario.empresa
            empresas = Empresa.objects.filter(id=empresa.id)
            empresa_id = empresa.id
            filtro_empresa = Q(empresa_id=empresa_id)
        except Exception:
            messages.error(request, "No tienes una empresa asignada. Contacta al administrador.")
            return render(request, 'dashboard/saldos.html', {'empresas': [], 'facturas': []})   

    # Filtrado base de facturas pendientes
    facturas = Factura.objects.filter(estatus='pendiente').filter(filtro_empresa)
    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)
    if origen == 'local':
        facturas = facturas.filter(local__isnull=False)
        
    elif origen == 'area':
        facturas = facturas.filter(area_comun__isnull=False)
        
    # Subconsulta: total pagado por factura
    pagos_subquery = Pago.objects.filter(factura=OuterRef('pk')) \
        .values('factura') \
        .annotate(total_pagado_dash=Coalesce(Sum('monto'), Value(0, output_field=DecimalField()))) \
        .values('total_pagado_dash')
    facturas = facturas.annotate(
        total_pagado_dash=Coalesce(Subquery(pagos_subquery), Value(0, output_field=DecimalField())),
        saldo_pendiente_dash=ExpressionWrapper(
            F('monto') - Coalesce(Subquery(pagos_subquery), Value(0, output_field=DecimalField())),
            output_field=DecimalField()
        )
    )
    # --- Facturas otros ingresos ---
    facturas_otros = FacturaOtrosIngresos.objects.filter(estatus='pendiente', activo=True)
    if not es_super:
        facturas_otros = facturas_otros.filter(empresa=empresa)
    if empresa_id:
        facturas_otros = facturas_otros.filter(empresa_id=empresa_id)
    if cliente_id:
        facturas_otros = facturas_otros.filter(cliente_id=cliente_id)

    cobros_subquery = CobroOtrosIngresos.objects.filter(factura=OuterRef('pk')) \
        .values('factura') \
        .annotate(total_cobrado_dash=Coalesce(Sum('monto'), Value(0, output_field=DecimalField()))) \
        .values('total_cobrado_dash')
    facturas_otros = facturas_otros.annotate(
        total_cobrado_dash=Coalesce(Subquery(cobros_subquery), Value(0, output_field=DecimalField())),
        saldo_pendiente_dash=ExpressionWrapper(
            F('monto') - Coalesce(Subquery(cobros_subquery), Value(0, output_field=DecimalField())),
            output_field=DecimalField()
        )   
    )
    
    saldo_0_30 = facturas.filter(fecha_vencimiento__gt=hoy - timedelta(days=30)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0 
    saldo_31_60 = facturas.filter(fecha_vencimiento__gt=hoy - timedelta(days=60), fecha_vencimiento__lte=hoy - timedelta(days=30)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0 
    saldo_61_90 = facturas.filter(fecha_vencimiento__gt=hoy - timedelta(days=90), fecha_vencimiento__lte=hoy - timedelta(days=60)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0
    saldo_91_180 = facturas.filter(fecha_vencimiento__gt=hoy - timedelta(days=180), fecha_vencimiento__lte=hoy - timedelta(days=90)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0
    saldo_181_mas = facturas.filter(fecha_vencimiento__lte=hoy - timedelta(days=180)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0
    
    saldo_0_30_otros = facturas_otros.filter(fecha_vencimiento__gt=hoy - timedelta(days=30)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0
    saldo_31_60_otros = facturas_otros.filter(fecha_vencimiento__gt=hoy - timedelta(days=60), fecha_vencimiento__lte=hoy - timedelta(days=30)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0
    saldo_61_90_otros = facturas_otros.filter(fecha_vencimiento__gt=hoy - timedelta(days=90), fecha_vencimiento__lte=hoy - timedelta(days=60)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0
    saldo_91_180_otros = facturas_otros.filter(fecha_vencimiento__gt=hoy - timedelta(days=180), fecha_vencimiento__lte=hoy - timedelta(days=90)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0
    saldo_181_mas_otros = facturas_otros.filter(fecha_vencimiento__lte=hoy - timedelta(days=180)).aggregate(total=Sum('saldo_pendiente_dash'))['total'] or 0


    if origen == "todos":
        saldo_0_30_total = (saldo_0_30 or 0) + (saldo_0_30_otros or 0)
        saldo_31_60_total = (saldo_31_60 or 0) + (saldo_31_60_otros or 0)
        saldo_61_90_total = (saldo_61_90 or 0) + (saldo_61_90_otros or 0)
        saldo_91_180_total = (saldo_91_180 or 0) + (saldo_91_180_otros or 0)
        saldo_181_mas_total = (saldo_181_mas or 0) + (saldo_181_mas_otros or 0)
    else:
        saldo_0_30_total = saldo_0_30 if origen != "otros" else saldo_0_30_otros
        saldo_31_60_total = saldo_31_60 if origen != "otros" else saldo_31_60_otros
        saldo_61_90_total = saldo_61_90 if origen != "otros" else saldo_61_90_otros
        saldo_91_180_total = saldo_91_180 if origen != "otros" else saldo_91_180_otros
        saldo_181_mas_total = saldo_181_mas if origen != "otros" else saldo_181_mas_otros

        # --- Top 10 adeudos por local/área/otros ingresos ---
    if origen == "otros":
        top_adeudos = (
            facturas_otros
            .annotate(
                nombre_otro=Coalesce(
                    F('tipo_ingreso__nombre'),
                    Value('Otro ingreso'),
                    output_field=CharField()
                ),
                nombre_cliente=F('cliente__nombre')
            )
            .values('nombre_otro', 'nombre_cliente')
            .annotate(total=Sum('saldo_pendiente_dash'))
            .order_by('-total')[:10]
        )
        top_labels = [x['nombre_otro'] for x in top_adeudos]
        top_data = [float(x['total']) for x in top_adeudos]
        top_clientes = [x['nombre_cliente'] for x in top_adeudos]
    elif origen == "todos":
        # Top locales/áreas
        top_local_area = (
            facturas
            .annotate(
                nombre_local_area=Coalesce(
                    F('local__numero'),
                    F('area_comun__numero'),
                    output_field=CharField()
                ),
                tipo_origen=Case(
                    When(local__isnull=False, then=Value('Local')),
                    When(area_comun__isnull=False, then=Value('Área')),
                    default=Value(''),
                    output_field=CharField()
                ),
                nombre_cliente=F('cliente__nombre')
            )
            .values('nombre_local_area', 'tipo_origen', 'nombre_cliente')
            .annotate(total=Sum('saldo_pendiente_dash'))
        )
        # Top otros ingresos
        top_otros = (
            facturas_otros
            .annotate(
                nombre_otro=Coalesce(
                    F('tipo_ingreso__nombre'),
                    Value('Otro ingreso'),
                    output_field=CharField()
                ),
                nombre_cliente=F('cliente__nombre')
            )
            .values('nombre_otro', 'nombre_cliente')
            .annotate(total=Sum('saldo_pendiente_dash'))
        )
        # Unir ambos y ordenar
        top_combined = [
            {'label': f"{x['tipo_origen']} {x['nombre_local_area']}".strip(), 'total': float(x['total']), 'cliente': x['nombre_cliente']}
            for x in top_local_area
        ] + [
            {'label': x['nombre_otro'], 'total': float(x['total']), 'cliente': x['nombre_cliente']}
            for x in top_otros
        ]
        top_combined = sorted(top_combined, key=lambda x: x['total'], reverse=True)[:10]
        top_labels = [x['label'] for x in top_combined]
        top_data = [x['total'] for x in top_combined]
        top_clientes = [x['cliente'] for x in top_combined]
    else:
        top_adeudos = (
            facturas
            .annotate(
                nombre_local_area=Coalesce(
                    F('local__numero'),
                    F('area_comun__numero'),
                    output_field=CharField()
                ),
                tipo_origen=Case(
                    When(local__isnull=False, then=Value('Local')),
                    When(area_comun__isnull=False, then=Value('Área')),
                    default=Value(''),
                    output_field=CharField()
                ),
                nombre_cliente=F('cliente__nombre')
            )
            .values('nombre_local_area', 'tipo_origen', 'nombre_cliente')
            .annotate(total=Sum('saldo_pendiente_dash'))
            .order_by('-total')[:10]
        )
        top_labels = [
            f"{x['tipo_origen']} {x['nombre_local_area']}" if x['nombre_local_area'] else x['tipo_origen']
            for x in top_adeudos
        ]
        top_data = [float(x['total']) for x in top_adeudos]
        top_clientes = [x['nombre_cliente'] for x in top_adeudos]
    
    #clientes = Cliente.objects.filter(empresa__in=empresas)
    if origen == "otros":
        clientes = Cliente.objects.filter(
            id__in=facturas_otros.values_list('cliente_id', flat=True).distinct()
        )
    else:
        clientes = Cliente.objects.filter(empresa__in=empresas)

    return render(request, 'dashboard/saldos.html', {
        'facturas': facturas,
        'clientes': clientes,
        'empresas': empresas,
        'empresa_id': str(empresa_id) if empresa_id else "",
        'cliente_id': int(cliente_id) if cliente_id else None,
        'saldo_0_30': saldo_0_30,
        'saldo_31_60': saldo_31_60,
        'saldo_61_90': saldo_61_90,
        'saldo_91_180': saldo_91_180,
        'saldo_181_mas': saldo_181_mas,
        'saldo_0_30_otros': saldo_0_30_otros,
        'saldo_31_60_otros': saldo_31_60_otros,
        'saldo_61_90_otros': saldo_61_90_otros,
        'saldo_91_180_otros': saldo_91_180_otros,
        'saldo_181_mas_otros': saldo_181_mas_otros,
        'origen': origen,
        'es_super': es_super,
        'top_labels': top_labels,
        'top_data': top_data,
        'facturas_otros': facturas_otros,
        'top_clientes': top_clientes,
        'saldo_0_30_total': saldo_0_30_total,
        'saldo_31_60_total': saldo_31_60_total,
        'saldo_61_90_total': saldo_61_90_total,
        'saldo_91_180_total': saldo_91_180_total,
        'saldo_181_mas_total': saldo_181_mas_total,
    })

@login_required
#pagos.html
#dashboard cuotas
@login_required
def dashboard_pagos(request):
    anio_actual = datetime.now().year
    anio = request.GET.get('anio')
    anio_seleccionado = request.GET.get('anio', anio_actual)
    if not anio:
        anio = anio_actual
    es_super = request.user.is_superuser

    if es_super:
        empresas = Empresa.objects.all()
        empresa_id = request.GET.get('empresa')
        if not empresa_id or empresa_id == "todas":
            filtro_empresa = Q()
            empresa = None
        else:
            filtro_empresa = Q(factura__empresa_id=empresa_id)
            empresa = Empresa.objects.get(pk=empresa_id)
    else:
        empresa = request.user.perfilusuario.empresa
        empresas = Empresa.objects.filter(id=empresa.id)
        empresa_id = empresa.id
        filtro_empresa = Q(factura__empresa_id=empresa_id)

    cliente_id = request.GET.get('cliente')
    origen = request.GET.get('origen')
    anio = request.GET.get('anio')
    mes = request.GET.get('mes')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

   

    filtro = Q(factura__activo=True) & filtro_empresa

    if cliente_id:
        filtro &= Q(factura__cliente_id=cliente_id)
    if origen == 'local':
        filtro &= Q(factura__local__isnull=False)
    elif origen == 'area':
        filtro &= Q(factura__area_comun__isnull=False)

    pagos = Pago.objects.exclude(forma_pago='nota_credito').filter(filtro)

    # Cobros de otros ingresos
    otros_cobros = CobroOtrosIngresos.objects.select_related(
    'factura', 'factura__empresa', 'factura__cliente', 'factura__tipo_ingreso').all()
    #otros_cobros = CobroOtrosIngresos.objects.all()
    if not request.user.is_superuser:
        otros_cobros = otros_cobros.filter(factura__empresa=request.user.perfilusuario.empresa)
    if empresa_id:
        otros_cobros = otros_cobros.filter(factura__empresa_id=empresa_id)
    if cliente_id:
        otros_cobros = otros_cobros.filter(factura__cliente_id=cliente_id)
    if anio:
        otros_cobros = otros_cobros.filter(fecha_cobro__year=anio)
    if mes:
        otros_cobros = otros_cobros.filter(fecha_cobro__month=mes)
    if fecha_inicio and fecha_fin:
        try:
            fecha_i = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            fecha_f = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            otros_cobros = otros_cobros.filter(fecha_cobro__range=[fecha_i, fecha_f])
        except:
            pass

    
                
    # --- AJUSTE CLAVE ---
    if origen == 'local' or origen == 'area':
        otros_cobros = CobroOtrosIngresos.objects.none()
        otros_por_mes = []
        otros_por_anio = []
    else:
        otros_por_mes = otros_cobros.annotate(mes=TruncMonth('fecha_cobro')).values('mes').annotate(total=Sum('monto')).order_by('mes')
        otros_por_anio = otros_cobros.annotate(anio=TruncYear('fecha_cobro')).values('anio').annotate(total=Sum('monto')).order_by('anio')    

    # Filtros de fechas para pagos
    if anio:
        pagos = pagos.filter(fecha_pago__year=anio)
    if mes:
        pagos = pagos.filter(fecha_pago__month=mes)
    if fecha_inicio and fecha_fin:
        try:
            fecha_i = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            fecha_f = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            pagos = pagos.filter(fecha_pago__range=[fecha_i, fecha_f])
        except:
            pass

    if origen == 'otros':
        pagos = Pago.objects.none() 
        pagos_por_mes = []
        pagos_por_anio = []
    else:
        pagos_por_mes = pagos.annotate(mes=TruncMonth('fecha_pago')).values('mes').annotate(total=Sum('monto')).order_by('mes')
        pagos_por_anio = pagos.annotate(anio=TruncYear('fecha_pago')).values('anio').annotate(total=Sum('monto')).order_by('anio')

    # Otros ingresos por mes y año para los gráficos
    otros_por_mes = otros_cobros.annotate(mes=TruncMonth('fecha_cobro')).values('mes').annotate(total=Sum('monto')).order_by('mes')
    otros_por_anio = otros_cobros.annotate(anio=TruncYear('fecha_cobro')).values('anio').annotate(total=Sum('monto')).order_by('anio')

    clientes = Cliente.objects.filter(empresa__in=empresas)

    # Suma total de pagos y otros ingresos
    total_pagos = pagos.aggregate(total=Sum('monto'))['total'] or 0
    total_otros = otros_cobros.aggregate(total=Sum('monto'))['total'] or 0
    total_general = total_pagos + total_otros

    # Unifica meses de ambos queryset
    meses_cuotas = {p['mes']: p['total'] for p in pagos_por_mes}
    meses_otros = {o['mes']: o['total'] for o in otros_por_mes}
    todos_los_meses = sorted(set(list(meses_cuotas.keys()) + list(meses_otros.keys())))

    # Prepara datos alineados
    labels_meses = [DateFormat(m).format('F Y') for m in todos_los_meses]
    data_cuotas = [meses_cuotas.get(m, 0) for m in todos_los_meses]
    data_otros = [meses_otros.get(m, 0) for m in todos_los_meses]

    # --- PRESUPUESTO DE INGRESOS POR MES ---
    
    presup_qs = PresupuestoIngreso.objects.all()
    if anio:
        presup_qs = presup_qs.filter(anio=anio)
    if empresa:
        presup_qs = presup_qs.filter(empresa=empresa)
    
    presup_dict = {}
    for p in presup_qs:
        key = (p.anio, p.mes)
        presup_dict.setdefault(key, 0)
        if origen == 'local' and p.origen == 'local':
            presup_dict[key] += float(p.monto_presupuestado)
        elif origen == 'area' and p.origen == 'area':
            presup_dict[key] += float(p.monto_presupuestado)
        elif origen == 'otros' and p.origen == 'otros':
            presup_dict[key] += float(p.monto_presupuestado)
        elif origen in (None, '', 'todos', 'Todo', 'Todos'):  # Todos los orígenes
            presup_dict[key] += float(p.monto_presupuestado)
    # Prepara datos de presupuesto alineados
    data_presupuesto = []
    for m in todos_los_meses:
        key = (m.year, m.month)
        data_presupuesto.append(presup_dict.get(key, 0))

        # Obtén todos los años presentes en pagos y presupuesto
    anios_pagos = pagos.values_list('fecha_pago__year', flat=True).distinct()
    anios_otros = otros_cobros.values_list('fecha_cobro__year', flat=True).distinct()
    anios_presupuesto = PresupuestoIngreso.objects.values_list('anio', flat=True).distinct()
    todos_los_anios = sorted(set(list(anios_pagos) + list(anios_otros) + list(anios_presupuesto)))
    #todos_los_anios = sorted(set(anios_pagos + anios_otros + anios_presupuesto))

    # Suma ingresos reales por año
    data_cuotas_anio = { (p['anio'].year if hasattr(p['anio'], 'year') else p['anio']): float(p['total']) for p in pagos_por_anio }
    data_otros_anio = { (o['anio'].year if hasattr(o['anio'], 'year') else o['anio']): float(o['total']) for o in otros_por_anio }

    # Suma presupuesto anual por año
    presup_anual_qs = PresupuestoIngreso.objects.all()
    if empresa:
        presup_anual_qs = presup_anual_qs.filter(empresa=empresa)
    presup_anual_dict = {}
    for p in presup_anual_qs:
        key = (p.anio, p.origen, p.tipo_otro or "")
        presup_anual_dict.setdefault(p.anio, {}).setdefault((p.origen, p.tipo_otro or ""), 0)
        presup_anual_dict[p.anio][(p.origen, p.tipo_otro or "")] += float(p.monto_presupuestado)

    data_presupuesto_anio = []
    for anio_ in todos_los_anios:
        if origen == 'local':
            total_presup = presup_anual_dict.get(anio_, {}).get(('local', ''), 0)
        elif origen == 'area':
            total_presup = presup_anual_dict.get(anio_, {}).get(('area', ''), 0)
        elif origen == 'otros':
            total_presup = sum(
                v for (o, _), v in presup_anual_dict.get(anio_, {}).items() if o == 'otros'
            )
        else:  
            presup_local = presup_anual_dict.get(anio_, {}).get(('local', ''), 0)
            presup_area = presup_anual_dict.get(anio_, {}).get(('area', ''), 0)
            presup_otros = sum(
                v for (o, _), v in presup_anual_dict.get(anio_, {}).items() if o == 'otros'
            )
            total_presup = presup_local + presup_area + presup_otros
        data_presupuesto_anio.append(total_presup)

    labels_anios = [str(a) for a in todos_los_anios]
    data_cuotas_anio_list = [data_cuotas_anio.get(a, 0) for a in todos_los_anios]
    data_otros_anio_list = [data_otros_anio.get(a, 0) for a in todos_los_anios]    

        # --- FACTURAS POR COBRAR POR MES ---
    facturas_pendientes = Factura.objects.filter(estatus='pendiente', activo=True)
    if empresa:
        facturas_pendientes = facturas_pendientes.filter(empresa=empresa)
    if cliente_id:
        facturas_pendientes = facturas_pendientes.filter(cliente_id=cliente_id)
    if anio:
        facturas_pendientes = facturas_pendientes.filter(fecha_emision__year=anio)
    if mes:
        facturas_pendientes = facturas_pendientes.filter(fecha_emision__month=mes)
    if origen == 'local':
        facturas_pendientes = facturas_pendientes.filter(local__isnull=False)
    elif origen == 'area':
        facturas_pendientes = facturas_pendientes.filter(area_comun__isnull=False)
        
        # Anota el total pagado por factura
    facturas_pendientes = facturas_pendientes.annotate(
    pagado=Coalesce(Sum('pagos__monto'), 0.0, output_field=DecimalField())
).annotate(
    saldo_pendiente_db=ExpressionWrapper(
        F('monto') - F('pagado'),
        output_field=DecimalField()
    )
)
        
        # Obtén todos los meses/años presentes
    meses_facturas = sorted(set(f.fecha_vencimiento.replace(day=1) for f in facturas_pendientes))
    anios_facturas = sorted(set(f.fecha_vencimiento.year for f in facturas_pendientes))



        # Por mes: solo facturas emitidas en ese mes, no vencidas
    meses_por_cobrar = {}
    for mes in todos_los_meses:
        total = sum(
            float(f.monto) - float(f.pagado)
            for f in facturas_pendientes
            if f.fecha_vencimiento.year == mes.year and f.fecha_vencimiento.month == mes.month
        )
        meses_por_cobrar[mes] = total
    data_por_cobrar = [meses_por_cobrar.get(m, 0) for m in todos_los_meses]

    # Por año: solo facturas emitidas en ese año
    anios_por_cobrar = {}
    for anio in todos_los_anios:
        total = sum(
            float(f.monto) - float(f.pagado)
            for f in facturas_pendientes
            if f.fecha_vencimiento.year == anio
        )
        anios_por_cobrar[anio] = total
    data_por_cobrar_anio = [anios_por_cobrar.get(a, 0) for a in todos_los_anios]

    otros_tipos_por_mes = [defaultdict(float) for _ in range(len(todos_los_meses))]
    for cobro in otros_cobros:
        for idx, mes in enumerate(todos_los_meses):
            if cobro.fecha_cobro.year == mes.year and cobro.fecha_cobro.month == mes.month:

                tipo = getattr(cobro.factura, 'tipo_ingreso', None) or 'Otro'
              
                if hasattr(cobro.factura, 'get_tipo_ingreso_display'):
                    tipo = cobro.factura.get_tipo_ingreso_display()
                otros_tipos_por_mes[idx][tipo] += float(cobro.monto)
                break

    # Convierte a lista de listas de strings para el tooltip
    otros_tipos_tooltip = [
        [f"{tipo}: ${monto:,.2f}" for tipo, monto in tipos.items()]
        for tipos in otros_tipos_por_mes
]

    return render(request, 'dashboard/pagos.html', {
        'pagos': pagos,
        'empresas': empresas,
        'empresa_id': str(empresa_id) if empresa_id else "",
        'clientes': clientes,
        'cliente_id': int(cliente_id) if cliente_id else None,
        'origen': origen,
        'es_super': es_super,
        'pagos_por_mes': pagos_por_mes,
        'pagos_por_anio': pagos_por_anio,
        'otros_por_mes': otros_por_mes,
        'otros_por_anio': otros_por_anio,
        'anio': anio,
        'mes': mes,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'otros_cobros': otros_cobros,
        'total_pagos': total_pagos,
        'total_otros': total_otros,
        'total_general': total_general,
        'labels_meses': labels_meses,
        'data_cuotas': data_cuotas, 
        'data_otros': data_otros,
        'data_presupuesto': data_presupuesto,
        'data_cuotas_anio': data_cuotas_anio_list,
        'data_otros_anio': data_otros_anio_list,
        'data_presupuesto_anio': data_presupuesto_anio,
        'labels_anios': labels_anios,
        'data_por_cobrar': data_por_cobrar,
        'data_por_cobrar_anio': data_por_cobrar_anio,
        'meses_facturas': meses_facturas,
        'anios_facturas': anios_facturas,
        'otros_tipos_tooltip': json.dumps(otros_tipos_tooltip),
        'anio_actual': anio_actual,
        'anio_seleccionado': anio_seleccionado,
    })


@login_required
def cartera_vencida(request):
    hoy = timezone.now().date()
    filtro = request.GET.get('rango')
    origen = request.GET.get('origen')

    # Optimiza con select_related para evitar N+1 queries
    facturas = Factura.objects.filter(
        estatus='pendiente',
        fecha_vencimiento__lt=hoy,
        activo=True
    ).select_related('cliente', 'empresa', 'local', 'area_comun')

    facturas_otros = FacturaOtrosIngresos.objects.filter(
        estatus='pendiente',
        fecha_vencimiento__lt=hoy,
        activo=True
    ).select_related('cliente', 'empresa', 'tipo_ingreso')
    # facturas = Factura.objects.filter(
    #     estatus='pendiente',
    #     fecha_vencimiento__lt=hoy,
    #     activo=True
    # )

    # facturas_otros = FacturaOtrosIngresos.objects.filter(
    #     estatus='pendiente',
    #     fecha_vencimiento__lt=hoy,
    #     activo=True
    # )

    # Filtrar por empresa
    if not request.user.is_superuser and hasattr(request.user, 'perfilusuario'):
        empresa = request.user.perfilusuario.empresa
        facturas = facturas.filter(empresa=empresa)
        facturas_otros = facturas_otros.filter(empresa=empresa)
        clientes = Cliente.objects.filter(empresa=empresa)
    else:
        if request.GET.get('empresa'):
            facturas = facturas.filter(empresa_id=request.GET['empresa'])
            facturas_otros = facturas_otros.filter(empresa_id=request.GET['empresa'])
            clientes = Cliente.objects.filter(empresa_id=request.GET['empresa'])
        else:
            clientes = Cliente.objects.all()
    
    # Filtrar por cliente
    cliente_id = request.GET.get('cliente')
    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)
        facturas_otros = facturas_otros.filter(cliente_id=cliente_id)

    # Filtrar por origen
    if origen == 'local':
        facturas = facturas.filter(local__isnull=False)
        facturas_otros = facturas_otros.none()
    elif origen == 'area':
        facturas = facturas.filter(area_comun__isnull=False)
        facturas_otros = facturas_otros.none()
    elif origen == 'otros':
        facturas = facturas.none()

    # Filtros de rango de días vencidos
    if filtro == 'menor30':
        facturas = facturas.filter(fecha_vencimiento__gt=hoy - timedelta(days=30))
        facturas_otros = facturas_otros.filter(fecha_vencimiento__gt=hoy - timedelta(days=30))
    elif filtro == '30a60':
        facturas = facturas.filter(
            fecha_vencimiento__lte=hoy - timedelta(days=30),
            fecha_vencimiento__gt=hoy - timedelta(days=60)
        )
        facturas_otros = facturas_otros.filter(
            fecha_vencimiento__lte=hoy - timedelta(days=30),
            fecha_vencimiento__gt=hoy - timedelta(days=60)
        )
    elif filtro == '60a90':
        facturas = facturas.filter(
            fecha_vencimiento__lte=hoy - timedelta(days=60),
            fecha_vencimiento__gt=hoy - timedelta(days=90)
        )
        facturas_otros = facturas_otros.filter(
            fecha_vencimiento__lte=hoy - timedelta(days=60),
            fecha_vencimiento__gt=hoy - timedelta(days=90)
        )
    elif filtro == '90a180':
        facturas = facturas.filter(
            fecha_vencimiento__lte=hoy - timedelta(days=90),
            fecha_vencimiento__gt=hoy - timedelta(days=180)
        )
        facturas_otros = facturas_otros.filter(
            fecha_vencimiento__lte=hoy - timedelta(days=90),
            fecha_vencimiento__gt=hoy - timedelta(days=180)
        )
    elif filtro == 'mas180':
        facturas = facturas.filter(fecha_vencimiento__lte=hoy - timedelta(days=180))
        facturas_otros = facturas_otros.filter(fecha_vencimiento__lte=hoy - timedelta(days=180))

    # Días de atraso y tipo
    for f in facturas:
        f.dias_vencidos = (hoy - f.fecha_vencimiento).days
        f.tipo_origen = 'local' if getattr(f, 'local_id', None) else 'area' if getattr(f, 'area_comun_id', None) else 'cuota'
        f.es_otro = False
    for f in facturas_otros:
        f.dias_vencidos = (hoy - f.fecha_vencimiento).days
        f.tipo_origen = 'otros'
        f.es_otro = True

    # Unir ambos queryset y ordenar
    facturas_todas = list(facturas) + list(facturas_otros)
    facturas_todas.sort(key=lambda x: x.fecha_vencimiento, reverse=True)

    # Paginar la lista combinada
    paginator = Paginator(facturas_todas, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'facturacion/cartera_vencida.html', {
        'facturas': page_obj,
        'hoy': hoy,
        'empresas': Empresa.objects.all(),
        'clientes': clientes,
        'rango_seleccionado': filtro
    })

@login_required
def exportar_cartera_excel(request):
    cliente_id = request.GET.get('cliente')
    hoy = timezone.now().date()
    origen = request.GET.get('origen')

    # Facturas cuotas
    facturas = Factura.objects.filter(
        estatus='pendiente',
        fecha_vencimiento__lt=hoy,
        activo=True
    )
    # Facturas otros ingresos
    facturas_otros = FacturaOtrosIngresos.objects.filter(
        estatus='pendiente',
        fecha_vencimiento__lt=hoy,
        activo=True
    )

    if not request.user.is_superuser and hasattr(request.user, 'perfilusuario'):
        empresa = request.user.perfilusuario.empresa
        facturas = facturas.filter(empresa=empresa)
        facturas_otros = facturas_otros.filter(empresa=empresa)
    elif request.GET.get('empresa'):
        empresa_id = request.GET.get('empresa')
        facturas = facturas.filter(empresa_id=empresa_id)
        facturas_otros = facturas_otros.filter(empresa_id=empresa_id)

    if origen == 'local':
        facturas = facturas.filter(local__isnull=False)
        facturas_otros = facturas_otros.none()
    elif origen == 'area':
        facturas = facturas.filter(area_comun__isnull=False)
        facturas_otros = facturas_otros.none()
    elif origen == 'otros':
        facturas = facturas.none()
        # facturas_otros ya trae todos los otros ingresos

    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)
        facturas_otros = facturas_otros.filter(cliente_id=cliente_id)

    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cartera Vencida"

    # Encabezados
    ws.append([
        'Folio', 'Cliente', 'Empresa', 'Origen', 'Monto',
        'Saldo Pendiente', 'Fecha Vencimiento', 'Días Vencidos'
    ])

    # Contenido: Facturas cuotas
    for factura in facturas:
        dias_vencidos = (hoy - factura.fecha_vencimiento).days
        origen_str = f"Local {factura.local.numero}" if factura.local else f"Área {factura.area_comun.numero}" if factura.area_comun else "-"
        ws.append([
            factura.folio,
            factura.cliente.nombre,
            factura.empresa.nombre,
            origen_str,
            float(factura.monto),
            float(factura.saldo_pendiente),
            str(factura.fecha_vencimiento),
            dias_vencidos
        ])

    # Contenido: Facturas otros ingresos
    for factura in facturas_otros:
        dias_vencidos = (hoy - factura.fecha_vencimiento).days
        tipo_ingreso = factura.get_tipo_ingreso_display() if hasattr(factura, 'get_tipo_ingreso_display') else (factura.tipo_ingreso.nombre if hasattr(factura.tipo_ingreso, 'nombre') else 'Otro ingreso')
        ws.append([
            factura.folio,
            factura.cliente.nombre,
            factura.empresa.nombre,
            tipo_ingreso,
            float(factura.monto),
            float(getattr(factura, 'saldo', factura.saldo_pendiente if hasattr(factura, 'saldo_pendiente') else factura.monto)),
            str(factura.fecha_vencimiento),
            dias_vencidos
        ])

    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=cartera_vencida.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_pagos_excel(request):
    empresa_id = request.GET.get('empresa')
    local_id = request.GET.get('local_id')
    area_id = request.GET.get('area_id')

    pagos = Pago.objects.select_related('factura', 'factura__empresa', 'factura__local', 'factura__area_comun', 'factura__cliente').all()
    if not request.user.is_superuser:
        pagos = pagos.filter(factura__empresa=request.user.perfilusuario.empresa)
    if empresa_id:
        pagos = pagos.filter(factura__empresa_id=empresa_id)
    if local_id:
        pagos = pagos.filter(factura__local_id=local_id)
    if area_id:
        pagos = pagos.filter(factura__area_comun_id=area_id)
    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ingresos"

    # Encabezados
    ws.append([
        'Local/Área','Cliente','Monto Cobro','Forma de Cobro','Folio Factura', 'Empresa',  
        'Fecha Cobro', 
    ])

    # Contenido
    for pago in pagos:
        factura = pago.factura
        local_area = factura.local.numero if factura.local else factura.area_comun.numero if factura.area_comun else '-'
        ws.append([
            local_area,
            factura.cliente.nombre,
            float(pago.monto),
            pago.forma_pago,
            factura.folio,
            factura.empresa.nombre,
            pago.fecha_pago.strftime('%Y-%m-%d') 
        ])

    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=pagos.xlsx'
    wb.save(response)
    return response


def buscar_por_id_o_nombre(modelo, valor, campo='nombre'):
    """Busca por ID, si falla busca por nombre (sin acentos, insensible a mayúsculas y espacios)."""
    if not valor:
        return None
    val = unidecode(str(valor)).strip().lower()
    try:
        return modelo.objects.get(pk=int(val))
    except (ValueError, modelo.DoesNotExist):
        todos = modelo.objects.all()
        candidatos = [
            obj for obj in todos
            if unidecode(str(getattr(obj, campo))).strip().lower() == val
        ]
        if len(candidatos) == 1:
            return candidatos[0]
        elif len(candidatos) > 1:
            conflicto = "; ".join([f"ID={obj.pk}, {campo}='{getattr(obj, campo)}'" for obj in candidatos])
            raise Exception(f"Conflicto: '{valor}' coincide con varios registros en {modelo.__name__}: {conflicto}")
        raise Exception(f"No se encontró '{valor}' en {modelo.__name__}")


@login_required
def plantilla_facturas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Facturas"

    # Encabezados (ajusta según tu modelo)
    ws.append([
        'folio', 'condominio', 'cliente', 'Num.local', 'Num.area',  'tipo cuota',
        'monto','fecha emision', 'fecha vencimiento', 'observaciones'
    ])
    # Fila de ejemplo (puedes poner valores ficticios)
    ws.append([
        'FAC001', 'Condominio Torre Reforma AC', 'Juan Pérez', 'L-101', '','mantenimiento', '1500.00', '2025-06-10', '2025-07-10', 'carga inicial'
    ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_facturas.xlsx'
    wb.save(response)
    return response        

@login_required
def editar_factura(request, factura_id):
    factura = get_object_or_404(Factura, pk=factura_id)
    
     # Bloqueo si la factura está pagada
    if factura.estatus == 'pagada':
        messages.warning(request, "Esta factura ya está pagada y no puede ser editada.")
        return redirect('lista_facturas')    

    if request.method == 'POST':
        form = FacturaEditForm(request.POST, instance=factura)
        if form.is_valid():
            factura_original = Factura.objects.get(pk=factura_id)
            factura_modificada = form.save(commit=False)

            # Comparar y guardar auditoría
            for field in form.changed_data:
                valor_anterior = getattr(factura_original, field)
                valor_nuevo = getattr(factura_modificada, field)
                if str(valor_anterior) != str(valor_nuevo):
                    AuditoriaCambio.objects.create(
                        modelo='factura',
                        objeto_id=factura.pk,
                        campo=field,
                        valor_anterior=valor_anterior,
                        valor_nuevo=valor_nuevo,
                        usuario=request.user,
                    )
            factura_modificada.save()
            return redirect('lista_facturas')
    else:
        form = FacturaEditForm(instance=factura)
    return render(request, 'facturacion/editar_factura.html', {
        'form': form,
        'factura': factura,
    })

@login_required
def exportar_lista_facturas_excel(request):
    empresa_id = request.GET.get('empresa')
    local_id = request.GET.get('local_id')
    area_id = request.GET.get('area_id')
    
    facturas = Factura.objects.all().select_related('cliente', 'empresa', 'local', 'area_comun')
    
    if not request.user.is_superuser:
        facturas = facturas.filter(empresa=request.user.perfilusuario.empresa)  
    
    if empresa_id:
        facturas = facturas.filter(empresa_id=empresa_id)
    if local_id:
        facturas = facturas.filter(local_id=local_id)
    if area_id:
        facturas = facturas.filter(area_comun_id=area_id)
    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Facturas"
    # Encabezados
    ws.append([
        'Folio', 'Empresa', 'Cliente', 'Local/Área', 'Monto',
        'Saldo', 'Periodo', 'Estatus', 'Observaciones'
    ])
    # Contenido
    for factura in facturas:
        local_area = factura.local.numero if factura.local else (factura.area_comun.numero if factura.area_comun else '-')
        ws.append([
            factura.folio,
            factura.empresa.nombre,
            factura.cliente.nombre,
            local_area,
            float(factura.monto),
            float(factura.saldo_pendiente),
            #factura.fecha_emision.strftime('%Y-%m-%d'),
            factura.fecha_vencimiento.strftime('%Y-%m-%d'),
            factura.estatus,
            factura.observaciones or ''
        ])
    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=lista_facturas.xlsx'
    wb.save(response)
    return response

@staff_member_required
def carga_masiva_facturas_cobradas(request):
    if request.method == 'POST':
        form = FacturaCargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            errores = []
            COLUMNAS_ESPERADAS = 10  # Cambia según tus columnas
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None:
                    continue
                if len(row) != COLUMNAS_ESPERADAS:
                    errores.append(f"Fila {i}: número de columnas incorrecto ({len(row)} en vez de {COLUMNAS_ESPERADAS})")
                    continue
                folio, empresa_val, cliente_val, local_val, area_val, tipo_cuota, monto, fecha_emision, fecha_vencimiento, observaciones = row
                try:
                    empresa = buscar_por_id_o_nombre(Empresa, empresa_val)
                    if not empresa:
                        errores.append(f"Fila {i}: No se encontró la empresa '{empresa_val}'")
                        continue

                    # Validar folio único por empresa
                    if Factura.objects.filter(folio=str(folio), empresa=empresa).exists():
                        errores.append(f"Fila {i}: El folio '{folio}' ya existe para la empresa '{empresa}'.")
                        continue

                    # Validar local o área
                    local = buscar_por_id_o_nombre(LocalComercial, local_val, campo='numero') if local_val else None
                    area = buscar_por_id_o_nombre(AreaComun, area_val, campo='numero') if area_val else None
                    if local_val and not local:
                        errores.append(f"Fila {i}: No se encontró el local '{local_val}'")
                        continue
                    if area_val and not area:
                        errores.append(f"Fila {i}: No se encontró el área '{area_val}'")
                        continue

                    # Buscar o crear cliente
                    #cliente, _ = Cliente.objects.get_or_create(
                     #   nombre=cliente_val,
                     # 3  empresa=empresa
                    #)

                    # Buscar o crear cliente (manejo de duplicados)
                    clientes = Cliente.objects.filter(nombre=cliente_val, empresa=empresa)
                    #if clientes.count() == 1:
                    if clientes.exists():    
                        cliente = clientes.first()
                    #elif clientes.count() == 0:
                     #   cliente = Cliente.objects.create(nombre=cliente_val, empresa=empresa)
                    else:
                        cliente = Cliente.objects.create(nombre=cliente_val, empresa=empresa)

                    # Validar y convertir monto
                    try:
                        cuota_decimal = Decimal(monto)
                    except (InvalidOperation, TypeError, ValueError):
                        errores.append(f"Fila {i}: El valor de monto '{monto}' no es válido.")
                        continue

                    factura=Factura.objects.create(
                        folio=str(folio),
                        empresa=empresa,
                        cliente=cliente,
                        local=local,
                        area_comun=area,
                        tipo_cuota=tipo_cuota,
                        monto=cuota_decimal,
                        fecha_emision=fecha_emision,
                        fecha_vencimiento=fecha_vencimiento,
                        observaciones=observaciones or "",
                        estatus='cobrada',  # Establecer estatus como 'cobrada'
                    )
                    Pago.objects.create(
                        factura=Factura.objects.get(folio=str(folio), empresa=empresa),
                        monto=cuota_decimal,
                        forma_pago='transferencia', 
                        fecha_pago=fecha_emision,
                        registrado_por=request.user,
                        observaciones=observaciones or "",
                    )
                    
                    #factura.actualizar_estatus()  # ✅ Correcto

                    print(f"[DEBUG] Factura creada: {folio} para {cliente.nombre} ({empresa.nombre})")
                except ValueError as ve:
                    errores.append(f"Fila {i}: Error de valor - {str(ve)}")     
                except Exception as e:
                    import traceback
                    errores.append(f"Fila {i}: {str(e) or repr(e)}<br>{traceback.format_exc()}")

            if errores:
                messages.error(request, "Algunas facturas no se cargaron:<br>" + "<br>".join(errores))
            else:
                messages.success(request, "¡Facturas cargadas exitosamente!")
            return redirect('carga_masiva_facturas_cobradas')
    else:
        form = FacturaCargaMasivaForm()
    return render(request, 'facturacion/carga_masiva_facturas_cobradas.html', {'form': form})

@login_required
#carga masiva cuentas x cobrar (cartera vencida)
def carga_masiva_facturas(request):
    if request.method == 'POST':
        form = FacturaCargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            errores = []
            COLUMNAS_ESPERADAS = 10  # Cambia según tus columnas
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None:
                    continue
                if len(row) != COLUMNAS_ESPERADAS:
                    errores.append(f"Fila {i}: número de columnas incorrecto ({len(row)} en vez de {COLUMNAS_ESPERADAS})")
                    continue
                folio, empresa_val, cliente_val, local_val, area_val, tipo_cuota, monto, fecha_emision, fecha_vencimiento, observaciones = row
                try:
                    empresa = buscar_por_id_o_nombre(Empresa, empresa_val)
                    if not empresa:
                        errores.append(f"Fila {i}: No se encontró la empresa '{empresa_val}'")
                        continue

                    # Validar folio único por empresa
                    if Factura.objects.filter(folio=str(folio), empresa=empresa).exists():
                        errores.append(f"Fila {i}: El folio '{folio}' ya existe para la empresa '{empresa}'.")
                        continue

                    # Validar local o área
                    local = buscar_por_id_o_nombre(LocalComercial, local_val, campo='numero') if local_val else None
                    area = buscar_por_id_o_nombre(AreaComun, area_val, campo='numero') if area_val else None
                    if local_val and not local:
                        errores.append(f"Fila {i}: No se encontró el local '{local_val}'")
                        continue
                    if area_val and not area:
                        errores.append(f"Fila {i}: No se encontró el área '{area_val}'")
                        continue

                    clientes = Cliente.objects.filter(nombre=cliente_val, empresa=empresa)
                    #if clientes.count() == 1:
                    if clientes.exists():
                        cliente = clientes.first()
                    #elif clientes.count() == 0:
                     #   cliente = Cliente.objects.create(nombre=cliente_val, empresa=empresa)
                    else:
                        cliente = Cliente.objects.create(nombre=cliente_val, empresa=empresa)   

                    # Validar y convertir monto
                    try:
                        cuota_decimal = Decimal(monto)
                    except (InvalidOperation, TypeError, ValueError):
                        errores.append(f"Fila {i}: El valor de monto '{monto}' no es válido.")
                        continue

                    factura=Factura.objects.create(
                        folio=str(folio),
                        empresa=empresa,
                        cliente=cliente,
                        local=local,
                        area_comun=area,
                        tipo_cuota=tipo_cuota,
                        monto=cuota_decimal,
                        fecha_emision=fecha_emision,
                        fecha_vencimiento=fecha_vencimiento,
                        observaciones=observaciones or "",
                        estatus='pendiente',
                    )
                    #factura.actualizar_estatus()  # ✅ Correcto    
                except ValueError as ve:
                    errores.append(f"Fila {i}: Error de valor - {str(ve)}")     
                except Exception as e:
                    import traceback
                    errores.append(f"Fila {i}: {str(e) or repr(e)}<br>{traceback.format_exc()}")

            if errores:
                messages.error(request, "Algunas facturas no se cargaron:<br>" + "<br>".join(errores))
            else:
                messages.success(request, "¡Facturas cargadas exitosamente!")
            return redirect('carga_masiva_facturas')
    else:
        form = FacturaCargaMasivaForm()
    return render(request, 'facturacion/carga_masiva_facturas.html', {'form': form})

@login_required
def crear_factura_otros_ingresos(request):
    if request.method == 'POST':
        form = FacturaOtrosIngresosForm(request.POST,request.FILES,user=request.user)
        if form.is_valid():
            factura = form.save(commit=False)
            # Asignar empresa según el cliente seleccionado
            factura.empresa = request.user.perfilusuario.empresa
            # Generar folio único
            count = FacturaOtrosIngresos.objects.filter(fecha_emision__year=now().year).count() + 1
            factura.folio = f"OI-F{count:05d}"
            factura.save()
            messages.success(request, "Registro Exitoso.")
            return redirect('lista_facturas_otros_ingresos')
    else:
        form = FacturaOtrosIngresosForm(user=request.user)
    try:
        tipos_ingreso = TipoOtroIngreso.objects.filter(empresa=request.user.perfilusuario.empresa)
    except Exception:
        
        tipos_ingreso = []

    return render(request, 'otros_ingresos/crear_factura.html', {'form': form, 'tipos_ingreso': tipos_ingreso})

@login_required
def lista_facturas_otros_ingresos(request):
    #facturas = FacturaOtrosIngresos.objects.all().order_by('-fecha_emision')
    facturas = FacturaOtrosIngresos.objects.select_related('cliente', 'empresa', 'tipo_ingreso').all().order_by('-fecha_emision')
    # Filtrar por empresa si no es superusuario
    if not request.user.is_superuser:
        facturas = facturas.filter(empresa=request.user.perfilusuario.empresa)

    # Paginación
    paginator = Paginator(facturas, 25)  # 25 facturas por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'otros_ingresos/lista_facturas.html', {'facturas': page_obj})

@login_required
def registrar_cobro_otros_ingresos(request, factura_id):
    factura = get_object_or_404(FacturaOtrosIngresos, pk=factura_id)

    if request.method == 'POST':
        form = CobroForm (request.POST,request.FILES)
        if form.is_valid():
            cobro = form.save(commit=False)
            cobro.factura = factura
            cobro.registrado_por = request.user
            if cobro.monto > factura.saldo:
                messages.error(request, "El monto del cobro no puede ser mayor al saldo pendiente de la factura.")
            else:    
                cobro.save()
                # Actualiza estatus de la factura si ya está pagada
                total_cobrado = sum([c.monto for c in factura.cobros.all()])
                if total_cobrado >= factura.monto:
                    factura.estatus = 'cobrada'
                    factura.save()
                messages.success(request, "Cobro registrado correctamente.")
                return redirect('lista_facturas_otros_ingresos')
    else:
        form = CobroForm()

    return render(request, 'otros_ingresos/registrar_cobro.html', {
        'form': form,
        'factura': factura,
    })

@login_required
def detalle_factura_otros_ingresos(request, factura_id):
    factura = get_object_or_404(FacturaOtrosIngresos, pk=factura_id)
    cobros = factura.cobros.all().order_by('fecha_cobro')
    return render(request, 'otros_ingresos/detalle_factura.html', {
        'factura': factura,
        'cobros': cobros,
    })

@login_required
def reporte_cobros_otros_ingresos(request):
    
    empresa_id = request.GET.get('empresa')
    cliente_id = request.GET.get('cliente')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_ingreso = request.GET.get('tipo_ingreso')
    tipo_ingreso_id = request.GET.get('tipo_ingreso')    

    cobros = CobroOtrosIngresos.objects.select_related('factura', 'factura__empresa', 'factura__cliente')

    # Filtros
    if not request.user.is_superuser:
        cobros = cobros.filter(factura__empresa=request.user.perfilusuario.empresa)
    if empresa_id:
        cobros = cobros.filter(factura__empresa_id=empresa_id)
    if cliente_id:
        cobros = cobros.filter(factura__cliente_id=cliente_id)
    if fecha_inicio and fecha_fin:
        cobros = cobros.filter(fecha_cobro__range=[fecha_inicio, fecha_fin])
        
    
    #if tipo_ingreso:
    if tipo_ingreso_id and tipo_ingreso_id.isdigit():
        cobros = cobros.filter(factura__tipo_ingreso_id=tipo_ingreso_id)    

    total_cobrado = cobros.aggregate(total=Sum('monto'))['total'] or 0

    # Paginación
    paginator = Paginator(cobros.order_by('-fecha_cobro'), 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    empresas = Empresa.objects.all() if request.user.is_superuser else Empresa.objects.filter(id=request.user.perfilusuario.empresa.id)
    clientes = Cliente.objects.filter(empresa__in=empresas)
    tipos_ingreso = TipoOtroIngreso.objects.filter(empresa__in=empresas)

    return render(request, 'otros_ingresos/reporte_cobros.html', {
        'cobros': page_obj,
        'empresas': empresas,
        'clientes': clientes,
        'empresa_id': empresa_id,
        'cliente_id': cliente_id,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'total_cobrado': total_cobrado,
        'tipo_ingreso': tipo_ingreso,
        'tipo_ingreso_id': tipo_ingreso_id,
        'tipos_ingreso': tipos_ingreso,
    })


@login_required
def exportar_cobros_otros_ingresos_excel(request):
    empresa_id = request.GET.get('empresa')
    cliente_id = request.GET.get('cliente')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    cobros = CobroOtrosIngresos.objects.select_related('factura', 'factura__empresa', 'factura__cliente')

    if not request.user.is_superuser:
        cobros = cobros.filter(factura__empresa=request.user.perfilusuario.empresa)
    if empresa_id and empresa_id.isdigit():
        cobros = cobros.filter(factura__empresa_id=empresa_id)
    if cliente_id and cliente_id.isdigit():
        cobros = cobros.filter(factura__cliente_id=cliente_id)
    # Validar fechas antes de filtrar
    if fecha_inicio and fecha_fin:
        try:
            fecha_i = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            fecha_f = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            cobros = cobros.filter(fecha_cobro__range=[fecha_i, fecha_f])
        except ValueError:
            pass  # Ignora el filtro si las fechas no son válidas

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cobros Otros Ingresos"

    headers = [
        "Fecha cobro", "Empresa", "Cliente", "Tipo ingreso", "Monto", "Forma cobro",
        "Factura", "Comprobante", "Observaciones"
    ]
    ws.append(headers)

    for cobro in cobros:
        ws.append([
            cobro.fecha_cobro.strftime('%Y-%m-%d'),
            cobro.factura.empresa.nombre,
            cobro.factura.cliente.nombre,
            cobro.factura.get_tipo_ingreso_display(),
            float(cobro.monto),
            cobro.get_forma_cobro_display(),
            cobro.factura.folio,
            cobro.comprobante.url if cobro.comprobante else '',
            cobro.observaciones or ''
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="cobros_otros_ingresos.xlsx"'
    return response

@login_required
def exportar_lista_facturas_otros_ingresos_excel(request):
    empresa_id = request.GET.get('empresa')
    cliente_id = request.GET.get('cliente')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    facturas = FacturaOtrosIngresos.objects.select_related('empresa', 'cliente').all()

    if not request.user.is_superuser:
        facturas = facturas.filter(empresa=request.user.perfilusuario.empresa)
    if empresa_id:
        facturas = facturas.filter(empresa_id=empresa_id)
    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)
    if fecha_inicio and fecha_fin:
        try:
            fecha_i = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            fecha_f = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            facturas = facturas.filter(fecha_emision__range=[fecha_i, fecha_f])
        except ValueError:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas Otros Ingresos"

    ws.append([
        'Folio', 'Empresa', 'Cliente', 'Tipo ingreso', 'Monto','Saldo',
         'Periodo', 'Estatus', 'Observaciones'
    ])

    for factura in facturas:
        ws.append([
            factura.folio,
            factura.empresa.nombre,
            factura.cliente.nombre,
            str(factura.tipo_ingreso) if factura.tipo_ingreso else '',
            #factura.get_tipo_ingreso_display() if hasattr(factura, 'get_tipo_ingreso_display') else factura.tipo_ingreso,
            float(factura.monto),
            float(factura.saldo),
            #factura.fecha_emision.strftime('%Y-%m-%d'),
            factura.fecha_vencimiento.strftime('%Y-%m-%d') if factura.fecha_vencimiento else '',
            factura.estatus,
            factura.observaciones or ''
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=lista_facturas_otros_ingresos.xlsx'
    wb.save(response)
    return response


@login_required
def crear_tipo_otro_ingreso(request):
    if request.method == 'POST':
        form = TipoOtroIngresoForm(request.POST)
        if form.is_valid():
            tipo_ingreso = form.save(commit=False)
            tipo_ingreso.empresa = request.user.perfilusuario.empresa
            tipo_ingreso.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'id': tipo_ingreso.id, 'nombre': tipo_ingreso.nombre})
            messages.success(request, "Tipo de ingreso creado correctamente.")
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Error al crear el tipo de ingreso.'})
            messages.error(request, "Error al crear el tipo de ingreso.")
    return redirect(request.META.get('HTTP_REFERER', 'crear_factura_otros_ingresos'))


@login_required
def tipos_otro_ingreso_json(request):
    tipos = TipoOtroIngreso.objects.filter(empresa=request.user.perfilusuario.empresa)
    data = [{'id': t.id, 'nombre': t.nombre} for t in tipos]
    return JsonResponse({'tipos': data})